"""
Phase 37B — Excel Integration for Policy Events.

Extends Phase 30/37A Excel output with policy events and notification status.
"""

from __future__ import annotations

import os
from dataclasses import dataclass, field
from datetime import datetime, date, time, timedelta
from typing import Any, Dict, List, Optional

import pytz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.attendance.daily_excel import (
    DailyExcelExporter,
    DailyExportRequest,
    DailyExportResult,
    HEADER_FONT,
    CELL_FONT,
    HEADER_FILL,
    BORDER_THIN,
    ALIGN_CENTER,
    ALIGN_LEFT,
    STATE_COLORS,
)
from app.attendance.contract import AttendanceRecord, IdentityCertainty
from app.attendance.timetable import AttendanceState, Timetable, TimetableEntry, SessionDay
from app.attendance.repository import AttendanceRepository
from app.attendance.calendar import CalendarEngine
from app.attendance.daily_resolver import DailyExpectedResolver, ExpectedStudent, ExpectedSession
from app.attendance.policy_engine.contract import PolicyEvent, PolicyType, PolicyEventState


# Additional colors for policy events
POLICY_EVENT_COLORS = {
    PolicyType.MORNING_ABSENCE: "FFC7CE",    # Light red
    PolicyType.LONG_EXIT: "FFEB9C",          # Light yellow
    PolicyType.MISSING_CHECKOUT: "FFC7CE",   # Light red
    PolicyType.SHORT_EXIT: "C6EFCE",         # Light green (filtered)
}

NOTIFICATION_STATUS_COLORS = {
    "pending": "FFF2CC",      # Light yellow
    "sending": "FFEB9C",      # Light yellow
    "sent": "C6EFCE",         # Light green
    "retry": "FFEB9C",        # Light yellow
    "failed": "FFC7CE",       # Light red
    "disabled": "F2F2F2",     # Light gray
    "no_recipient": "F2F2F2", # Light gray
    "rate_limited": "FFEB9C", # Light yellow
    "deduplicated": "DDEBF7", # Light blue
    "ignored": "F2F2F2",      # Light gray
}


@dataclass
class PolicyExcelExporterConfig:
    """Configuration for policy Excel export."""
    include_policy_events_sheet: bool = True
    include_notification_status_sheet: bool = True
    include_policy_summary_sheet: bool = True


class PolicyExcelExporter:
    """
    Extended Excel exporter that includes policy events and notification status.
    
    Adds sheets:
    - POLICY_EVENTS: All policy events for the date
    - NOTIFICATION_STATUS: Notification delivery status
    - POLICY_SUMMARY: Summary of policy events by type
    """
    
    def __init__(
        self,
        repository: Optional[AttendanceRepository] = None,
        config: Optional[PolicyExcelExporterConfig] = None,
    ):
        self.base_exporter = DailyExcelExporter(repository)
        self.config = config or PolicyExcelExporterConfig()
        self.timezone = pytz.timezone("Asia/Bangkok")
    
    def export_daily_with_policy(
        self,
        request: DailyExportRequest,
        policy_events: List[PolicyEvent],
        notification_records: List[Dict[str, Any]],
    ) -> DailyExportResult:
        """
        Export daily attendance with policy events and notification status.
        
        Args:
            request: Base export request
            policy_events: List of PolicyEvents for the date
            notification_records: List of notification records (from queue)
            
        Returns:
            DailyExportResult
        """
        try:
            # First, run base export
            base_result = self.base_exporter.export_daily_attendance(request)
            
            if not base_result.success:
                return base_result
            
            # Load the workbook to add policy sheets
            from openpyxl import load_workbook
            wb = load_workbook(request.output_path)
            
            # Add policy sheets
            sheets_added = []
            
            if self.config.include_policy_events_sheet and policy_events:
                self._create_policy_events_sheet(wb, policy_events, request)
                sheets_added.append("POLICY_EVENTS")
            
            if self.config.include_notification_status_sheet and notification_records:
                self._create_notification_status_sheet(wb, notification_records, request)
                sheets_added.append("NOTIFICATION_STATUS")
            
            if self.config.include_policy_summary_sheet and policy_events:
                self._create_policy_summary_sheet(wb, policy_events, request)
                sheets_added.append("POLICY_SUMMARY")
            
            # Save updated workbook
            wb.save(request.output_path)
            
            # Update result
            base_result.sheets_created.extend(sheets_added)
            
            return base_result
            
        except Exception as e:
            return DailyExportResult(
                success=False,
                error=f"Policy export failed: {str(e)}",
                output_path=request.output_path,
            )
    
    def _create_policy_events_sheet(
        self,
        wb: Workbook,
        policy_events: List[PolicyEvent],
        request: DailyExportRequest,
    ) -> None:
        """Create the policy events sheet."""
        ws = wb.create_sheet("POLICY_EVENTS")
        
        headers = [
            "No.",
            "Event ID",
            "Student ID",
            "Student Name",
            "Policy Type",
            "Occurred At",
            "Effective At",
            "Source Attendance Event ID",
            "Global Observation ID",
            "State",
            "Evidence Summary",
            "Idempotency Key",
            "Created At",
            # Semantic fields (Phase 37D)
            "Session Type",
            "Subject",
            "Location",
            "Expected Location",
            "Outside Allowed",
            "Semantic State",
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BORDER_THIN
            cell.alignment = ALIGN_CENTER
            ws.column_dimensions[get_column_letter(col_num)].width = 18
        
        ws.freeze_panes = "A2"
        
        # Sort by occurred_at
        sorted_events = sorted(policy_events, key=lambda e: e.occurred_at)
        
        row_num = 2
        for idx, event in enumerate(sorted_events, 1):
            # Format timestamps
            occurred_dt = datetime.fromtimestamp(event.occurred_at, self.timezone)
            effective_dt = datetime.fromtimestamp(event.effective_at, self.timezone)
            
            # Create evidence summary
            evidence_summary = self._format_evidence_summary(event.evidence)
            
            # Extract semantic fields from evidence
            session_type = event.evidence.get("session_type", "N/A")
            subject = event.evidence.get("subject", "N/A")
            location = event.evidence.get("location", "N/A")
            expected_location = event.evidence.get("expected_location", "N/A")
            outside_allowed = event.evidence.get("outside_allowed", False)
            semantic_state = event.evidence.get("semantic_state", "N/A")
            
            row_data = [
                idx,
                event.event_id,
                event.student_id,
                event.evidence.get("student_name", f"Student {event.student_id}"),
                event.policy_type.value,
                occurred_dt.strftime("%Y-%m-%d %H:%M:%S"),
                effective_dt.strftime("%Y-%m-%d %H:%M:%S"),
                event.source_attendance_event_id,
                event.source_global_observation_id or "N/A",
                event.state.value,
                evidence_summary,
                event.idempotency_key,
                event.created_at,
                # Semantic fields (Phase 37D)
                session_type,
                subject,
                location,
                expected_location,
                "Yes" if outside_allowed else "No",
                semantic_state,
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.font = CELL_FONT
                cell.border = BORDER_THIN
                cell.alignment = ALIGN_CENTER
                
                # Color code by policy type
                if col_num == 5:  # Policy Type column
                    color = POLICY_EVENT_COLORS.get(event.policy_type, "FFFFFF")
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                
                # Color code by state
                if col_num == 10:  # State column
                    color = NOTIFICATION_STATUS_COLORS.get(event.state.value, "FFFFFF")
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                
                # Color code by semantic state
                if col_num == 18:  # Semantic State column
                    if semantic_state == "EXPECTED_OUTSIDE":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
                    elif semantic_state == "EXPECTED_INSIDE":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
            
            row_num += 1
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 12), 50)
            ws.column_dimensions[col[0].column_letter].width = adjusted_width
    
    def _create_notification_status_sheet(
        self,
        wb: Workbook,
        notification_records: List[Dict[str, Any]],
        request: DailyExportRequest,
    ) -> None:
        """Create the notification status sheet."""
        ws = wb.create_sheet("NOTIFICATION_STATUS")
        
        headers = [
            "No.",
            "Notification ID",
            "Idempotency Key",
            "Event ID",
            "Student ID",
            "Parent ID",
            "Telegram Chat ID",
            "Notification Type",
            "Status",
            "Attempts",
            "Max Attempts",
            "Created At",
            "Sent At",
            "Last Error",
            "Last Attempt At",
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BORDER_THIN
            cell.alignment = ALIGN_CENTER
            ws.column_dimensions[get_column_letter(col_num)].width = 18
        
        ws.freeze_panes = "A2"
        
        # Sort by created_at
        sorted_records = sorted(notification_records, key=lambda r: r.get("created_at", ""))
        
        row_num = 2
        for idx, record in enumerate(sorted_records, 1):
            row_data = [
                idx,
                record.get("notification_id", ""),
                record.get("idempotency_key", ""),
                record.get("event_id", ""),
                record.get("student_id", ""),
                record.get("parent_id", ""),
                record.get("telegram_chat_id", ""),
                record.get("notification_type", ""),
                record.get("status", ""),
                record.get("attempts", 0),
                record.get("max_attempts", 3),
                record.get("created_at", ""),
                record.get("sent_at", "") or "N/A",
                record.get("last_error", "") or "N/A",
                record.get("last_attempt_at", "") or "N/A",
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.font = CELL_FONT
                cell.border = BORDER_THIN
                cell.alignment = ALIGN_CENTER
                
                # Color code by status
                if col_num == 9:  # Status column
                    status = str(value).lower()
                    color = NOTIFICATION_STATUS_COLORS.get(status, "FFFFFF")
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            row_num += 1
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 12), 50)
            ws.column_dimensions[col[0].column_letter].width = adjusted_width
    
    def _create_policy_summary_sheet(
        self,
        wb: Workbook,
        policy_events: List[PolicyEvent],
        request: DailyExportRequest,
    ) -> None:
        """Create the policy summary sheet."""
        ws = wb.create_sheet("POLICY_SUMMARY")
        
        # Count by policy type
        type_counts = {}
        state_counts = {}
        student_counts = {}
        
        for event in policy_events:
            ptype = event.policy_type.value
            type_counts[ptype] = type_counts.get(ptype, 0) + 1
            
            state = event.state.value
            state_counts[state] = state_counts.get(state, 0) + 1
            
            student_id = event.student_id
            if student_id not in student_counts:
                student_counts[student_id] = {}
            student_counts[student_id][ptype] = student_counts[student_id].get(ptype, 0) + 1
        
        # Write summary
        ws.append(["POLICY EVENT SUMMARY"])
        ws.append(["Date:", request.date.strftime("%Y-%m-%d")])
        ws.append(["Export Time:", datetime.now(self.timezone).strftime("%Y-%m-%d %H:%M:%S")])
        ws.append([])
        
        ws.append(["BY POLICY TYPE:"])
        ws.append(["Policy Type", "Count"])
        for ptype, count in sorted(type_counts.items()):
            ws.append([ptype.replace("_", " ").title(), count])
        ws.append([])
        
        ws.append(["BY STATE:"])
        ws.append(["State", "Count"])
        for state, count in sorted(state_counts.items()):
            ws.append([state.replace("_", " ").title(), count])
        ws.append([])
        
        ws.append(["BY STUDENT:"])
        ws.append(["Student ID", "Morning Absence", "Long Exit", "Missing Checkout", "Short Exit", "Total"])
        for student_id, counts in sorted(student_counts.items()):
            total = sum(counts.values())
            ws.append([
                student_id,
                counts.get("morning_absence", 0),
                counts.get("long_exit", 0),
                counts.get("missing_checkout", 0),
                counts.get("short_exit", 0),
                total,
            ])
        
        # Format
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
            for cell in row:
                cell.font = HEADER_FONT if cell.column == 1 else CELL_FONT
                cell.alignment = ALIGN_LEFT
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 12), 30)
            ws.column_dimensions[col[0].column_letter].width = adjusted_width
    
    def _format_evidence_summary(self, evidence: Dict[str, Any]) -> str:
        """Create a concise evidence summary."""
        parts = []
        
        for key in ["check_time", "expected_entry_time", "out_time", "elapsed_seconds", 
                    "threshold_seconds", "expected_departure_time", "duration_seconds", "status"]:
            if key in evidence:
                value = evidence[key]
                if key.endswith("_seconds") and isinstance(value, (int, float)):
                    value = f"{int(value // 60)}m {int(value % 60)}s"
                parts.append(f"{key}: {value}")
        
        return "; ".join(parts) if parts else "N/A"


def create_policy_excel_exporter(
    repository: Optional[AttendanceRepository] = None,
    config: Optional[PolicyExcelExporterConfig] = None,
) -> PolicyExcelExporter:
    """Factory function to create PolicyExcelExporter."""
    return PolicyExcelExporter(repository, config)