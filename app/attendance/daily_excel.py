"""
Phase 30 — Daily Excel Export.

Deterministic, auditable Excel export layer for attendance records.
Consumes canonical attendance data from Phase 25 repository and Phase 26 decisions.
"""

from __future__ import annotations

import hashlib
import os
from dataclasses import dataclass, field
from datetime import datetime, time, timedelta
from enum import Enum
from typing import Any, Dict, List, Optional, Set, Tuple

import pytz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.attendance.contract import AttendanceRecord, IdentityCertainty
from app.attendance.engine import AttendanceEngine
from app.attendance.policy import AttendanceDecision
from app.attendance.timetable import AttendanceState, Timetable, TimetableEntry, SessionDay
from app.attendance.repository import AttendanceRepository
from app.attendance.calendar import CalendarEngine, CalendarDay, DayType
from app.attendance.daily_resolver import DailyExpectedResolver, ExpectedStudent, ExpectedSession, ExpectedStatus

# Constants for Excel formatting
HEADER_FONT = Font(name="Calibri", size=11, bold=True)
CELL_FONT = Font(name="Calibri", size=10)
HEADER_FILL = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
BORDER_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")

# Color semantics for attendance states
STATE_COLORS = {
    AttendanceState.PRESENT: "C6EFCE",  # Light green
    AttendanceState.LATE: "FFEB9C",     # Light yellow
    AttendanceState.LEFT: "DDEBF7",     # Light blue
    AttendanceState.ABSENT: "FFC7CE",   # Light red
    AttendanceState.UNKNOWN: "F2F2F2",  # Light gray
    AttendanceState.EXPECTED: "E2EFDA", # Very light green
}

@dataclass
class DailyExportRequest:
    """
    Request for daily attendance export.

    Fields:
    - date: Date to export (datetime.date)
    - output_path: Path to save the Excel file
    - timezone: Timezone for date interpretation (default: 'Asia/Bangkok')
    - export_version: Version of the export format (default: '1.0')
    - timetable: Optional timetable for session information
    - include_events_sheet: Include events sheet (default: True)
    - include_provenance_sheet: Include provenance sheet (default: True)
    - include_summary_sheet: Include summary sheet (default: True)
    """
    date: datetime.date
    output_path: str
    timezone: str = "Asia/Bangkok"
    export_version: str = "1.0"
    timetable: Optional[Timetable] = None
    include_events_sheet: bool = True
    include_provenance_sheet: bool = True
    include_summary_sheet: bool = True

@dataclass
class DailyExportResult:
    """Result of daily export operation."""
    success: bool
    output_path: Optional[str] = None
    error: Optional[str] = None
    records_processed: int = 0
    records_exported: int = 0
    sheets_created: List[str] = field(default_factory=list)
    export_id: Optional[str] = None

class DailyExcelExporter:
    """
    Daily Excel exporter for attendance records.

    Core principles:
    - Read-only: Does not modify source data
    - Deterministic: Same input = same output
    - Auditable: Preserves full provenance
    - Professional: Clean, readable Excel format
    - Bounded: Only loads required date range
    """

    def __init__(self, repository: Optional[AttendanceRepository] = None):
        """
        Initialize exporter.

        Args:
            repository: Attendance repository (optional, will create if None)
        """
        self.repository = repository or AttendanceRepository()
        self.timezone = pytz.timezone("Asia/Bangkok")

    def export_daily_attendance(self, request: DailyExportRequest) -> DailyExportResult:
        """
        Export daily attendance to Excel workbook.

        Args:
            request: DailyExportRequest with export parameters

        Returns:
            DailyExportResult with export outcome
        """
        try:
            # Validate request
            validation_error = self._validate_export_request(request)
            if validation_error:
                return DailyExportResult(success=False, error=validation_error)

            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Generate export ID
            export_id = self._generate_export_id(request)

            # Query attendance records for the date
            records = self._query_daily_records(request)

            # Create daily resolver if timetable is provided
            daily_resolver = None
            calendar_engine = None
            if request.timetable is not None:
                calendar_engine = CalendarEngine()
                daily_resolver = DailyExpectedResolver(request.timetable, calendar_engine)

            # Create sheets
            sheets_created = []

            # Sheet 1: Daily Attendance (always created)
            self._create_daily_attendance_sheet(wb, records, request)
            sheets_created.append("DAILY_ATTENDANCE")

            # Sheet 2: Expected Schedule (if timetable provided)
            if request.timetable is not None and daily_resolver is not None:
                self._create_expected_schedule_sheet(wb, request, daily_resolver)
                sheets_created.append("EXPECTED_SCHEDULE")

            # Sheet 3: Events (optional)
            if request.include_events_sheet:
                self._create_events_sheet(wb, records, request)
                sheets_created.append("EVENTS")

            # Sheet 4: Summary (optional)
            if request.include_summary_sheet:
                self._create_summary_sheet(wb, records, request)
                sheets_created.append("SUMMARY")

            # Sheet 5: Provenance (optional)
            if request.include_provenance_sheet:
                self._create_provenance_sheet(wb, records, request)
                sheets_created.append("PROVENANCE")

            # Save workbook
            os.makedirs(os.path.dirname(request.output_path), exist_ok=True)
            wb.save(request.output_path)

            return DailyExportResult(
                success=True,
                output_path=request.output_path,
                records_processed=len(records),
                records_exported=len([r for r in records if self._should_export_record(r)]),
                sheets_created=sheets_created,
                export_id=export_id
            )

        except Exception as e:
            return DailyExportResult(
                success=False,
                error=f"Export failed: {str(e)}",
                output_path=request.output_path if 'request' in locals() else None
            )

    def _validate_export_request(self, request: DailyExportRequest) -> Optional[str]:
        """Validate export request."""
        if not request.date:
            return "Date is required"
        if not request.output_path:
            return "Output path is required"
        if not os.path.isabs(request.output_path):
            return "Output path must be absolute"
        if not request.output_path.lower().endswith('.xlsx'):
            return "Output path must have .xlsx extension"
        return None

    def _generate_export_id(self, request: DailyExportRequest) -> str:
        """Generate deterministic export ID."""
        content = f"EXPORT:{request.date.isoformat()}:{request.export_version}"
        hash_suffix = hashlib.sha256(content.encode()).hexdigest()[:16]
        return f"EXPORT-{hash_suffix}"

    def _query_daily_records(self, request: DailyExportRequest) -> List[AttendanceRecord]:
        """Query attendance records for the requested date."""
        # Convert date to timestamp range in the specified timezone
        tz = pytz.timezone(request.timezone)
        date_start = tz.localize(datetime.combine(request.date, time.min))
        date_end = tz.localize(datetime.combine(request.date, time.max))

        # Convert to UTC timestamps for query
        start_timestamp = date_start.astimezone(pytz.UTC).timestamp()
        end_timestamp = date_end.astimezone(pytz.UTC).timestamp()

        # Query records for the date range
        records = self.repository.query_by_time_range(
            start_timestamp=start_timestamp,
            end_timestamp=end_timestamp
        )

        return records

    def _should_export_record(self, record: AttendanceRecord) -> bool:
        """Determine if a record should be exported."""
        # Always export valid records
        return True

    def _create_daily_attendance_sheet(
        self,
        wb: Workbook,
        records: List[AttendanceRecord],
        request: DailyExportRequest
    ) -> None:
        """Create the main daily attendance sheet."""
        ws = wb.create_sheet("DAILY_ATTENDANCE")

        # Set up headers
        headers = [
            "No.",
            "Person ID",
            "Name",
            "Identity Certainty",
            "State",
            "IN Time",
            "OUT Time",
            "Duration",
            "Camera",
            "Global Observation",
            "Status"
        ]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BORDER_THIN
            cell.alignment = ALIGN_CENTER
            ws.column_dimensions[get_column_letter(col_num)].width = 15

        # Freeze header row
        ws.freeze_panes = "A2"

        # Sort records by timestamp for deterministic ordering
        sorted_records = sorted(records, key=lambda r: r.event_timestamp)
        
        # Write data rows
        row_num = 2
        for record in sorted_records:
            if not self._should_export_record(record):
                continue

            # Get attendance state from record
            attendance_state = self._get_attendance_state_from_record(record)

            # Format IN/OUT times
            in_time_str = self._format_time(record.event_timestamp) if record.is_in else "N/A"
            out_time_str = "N/A"  # Will be populated from matching OUT record

            # Find matching OUT record for IN records
            if record.is_in:
                out_record = self._find_matching_out_record(record, records)
                if out_record:
                    out_time_str = self._format_time(out_record.event_timestamp)

            # Calculate duration
            duration_str = self._calculate_duration(record, out_record) if record.is_in else "N/A"

            # Write row
            row_data = [
                row_num - 1,  # No.
                self._get_person_id(record),
                self._get_person_name(record),
                record.identity_certainty.value.upper(),
                attendance_state.value.upper(),
                in_time_str,
                out_time_str,
                duration_str,
                record.camera_id,
                record.global_observation_id or "N/A",
                self._get_status_text(attendance_state)
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.font = CELL_FONT
                cell.border = BORDER_THIN
                cell.alignment = ALIGN_CENTER

                # Apply color formatting for attendance state
                if col_num == 5 and attendance_state in STATE_COLORS:  # State column
                    cell.fill = PatternFill(
                        start_color=STATE_COLORS[attendance_state],
                        end_color=STATE_COLORS[attendance_state],
                        fill_type="solid"
                    )

            row_num += 1

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            # Ensure minimum width of 8 for readability
            ws.column_dimensions[col[0].column_letter].width = max(min(adjusted_width, 30), 8)

    def _create_events_sheet(
        self,
        wb: Workbook,
        records: List[AttendanceRecord],
        request: DailyExportRequest
    ) -> None:
        """Create the chronological events sheet."""
        ws = wb.create_sheet("EVENTS")

        # Set up headers
        headers = [
            "Timestamp",
            "Person ID",
            "Name",
            "Identity Certainty",
            "Direction",
            "Attendance State",
            "Camera",
            "Local Track",
            "Global Observation",
            "Source Resolution ID"
        ]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BORDER_THIN
            cell.alignment = ALIGN_CENTER
            ws.column_dimensions[get_column_letter(col_num)].width = 15

        # Freeze header row
        ws.freeze_panes = "A2"

        # Sort records by timestamp
        sorted_records = sorted(records, key=lambda r: r.event_timestamp)

        # Write data rows
        row_num = 2
        for record in sorted_records:
            if not self._should_export_record(record):
                continue

            attendance_state = self._get_attendance_state_from_record(record)

            row_data = [
                self._format_datetime(record.event_timestamp),
                self._get_person_id(record),
                self._get_person_name(record),
                record.identity_certainty.value.upper(),
                record.direction.upper(),
                attendance_state.value.upper(),
                record.camera_id,
                record.local_track_id,
                record.global_observation_id or "N/A",
                record.source_resolution_id
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.font = CELL_FONT
                cell.border = BORDER_THIN
                cell.alignment = ALIGN_CENTER

                # Apply color formatting for direction
                if col_num == 5:  # Direction column
                    if record.direction == "in":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            row_num += 1

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col[0].column_letter].width = min(adjusted_width, 30)

    def _create_summary_sheet(
        self,
        wb: Workbook,
        records: List[AttendanceRecord],
        request: DailyExportRequest
    ) -> None:
        """Create the summary sheet."""
        ws = wb.create_sheet("SUMMARY")

        # Calculate summary statistics
        total_records = len([r for r in records if self._should_export_record(r)])
        present_count = len([r for r in records if self._get_attendance_state_from_record(r) == AttendanceState.PRESENT])
        late_count = len([r for r in records if self._get_attendance_state_from_record(r) == AttendanceState.LATE])
        left_count = len([r for r in records if self._get_attendance_state_from_record(r) == AttendanceState.LEFT])
        absent_count = len([r for r in records if self._get_attendance_state_from_record(r) == AttendanceState.ABSENT])
        unknown_count = len([r for r in records if self._get_attendance_state_from_record(r) == AttendanceState.UNKNOWN])

        # Unique people count
        unique_people = set()
        for record in records:
            if self._should_export_record(record):
                person_id = self._get_person_id(record)
                if person_id and person_id != "UNKNOWN":
                    unique_people.add(person_id)
        unique_people_count = len(unique_people)

        # Write summary information
        ws.append(["ATTENDANCE SUMMARY"])
        ws.append(["Date:", request.date.strftime("%Y-%m-%d")])
        ws.append(["Export Time:", datetime.now(self.timezone).strftime("%Y-%m-%d %H:%M:%S")])
        ws.append(["Timezone:", request.timezone])
        ws.append([])

        ws.append(["TOTAL RECORDS:", total_records])
        ws.append(["UNIQUE PEOPLE:", unique_people_count])
        ws.append([])

        ws.append(["ATTENDANCE STATES:"])
        ws.append(["PRESENT:", present_count])
        ws.append(["LATE:", late_count])
        ws.append(["LEFT:", left_count])
        ws.append(["ABSENT:", absent_count])
        ws.append(["UNKNOWN:", unknown_count])
        ws.append([])

        # Format summary
        for row in ws.iter_rows(min_row=1, max_row=15, min_col=1, max_col=2):
            for cell in row:
                cell.font = HEADER_FONT if cell.column == 1 else CELL_FONT
                cell.alignment = ALIGN_LEFT

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col[0].column_letter].width = min(adjusted_width, 30)

    def _create_provenance_sheet(
        self,
        wb: Workbook,
        records: List[AttendanceRecord],
        request: DailyExportRequest
    ) -> None:
        """Create the provenance sheet."""
        ws = wb.create_sheet("PROVENANCE")

        # Set up headers
        headers = [
            "Attendance Record ID",
            "Source Resolution ID",
            "Source Raw Event ID",
            "Source Crossing Event ID",
            "Global Observation ID",
            "Camera ID",
            "Local Track ID",
            "Event Timestamp",
            "Geometry Version",
            "Geometry Config Hash",
            "Resolver Version",
            "Resolver Config Hash",
            "Timetable ID",
            "Timetable Version",
            "Attendance Schema Version"
        ]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BORDER_THIN
            cell.alignment = ALIGN_CENTER
            ws.column_dimensions[get_column_letter(col_num)].width = 20

        # Freeze header row
        ws.freeze_panes = "A2"

        # Write data rows
        row_num = 2
        for record in records:
            if not self._should_export_record(record):
                continue

            row_data = [
                record.attendance_record_id,
                record.source_resolution_id,
                record.source_raw_event_id,
                record.source_crossing_event_id or "N/A",
                record.global_observation_id or "N/A",
                record.camera_id,
                record.local_track_id,
                self._format_datetime(record.event_timestamp),
                record.geometry_version,
                record.geometry_config_hash,
                record.resolver_version,
                record.resolver_config_hash,
                "N/A",  # Timetable ID - not available in AttendanceRecord
                "N/A",  # Timetable Version - not available in AttendanceRecord
                record.attendance_schema_version
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.font = CELL_FONT
                cell.border = BORDER_THIN
                cell.alignment = ALIGN_CENTER

            row_num += 1

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col[0].column_letter].width = min(adjusted_width, 40)

    def _create_expected_schedule_sheet(
        self,
        wb: Workbook,
        request: DailyExportRequest,
        daily_resolver: DailyExpectedResolver
    ) -> None:
        """Create the expected schedule sheet from timetable."""
        ws = wb.create_sheet("EXPECTED_SCHEDULE")

        # Resolve expected students for the date
        expected_result = daily_resolver.resolve_for_date(request.date)

        # Set up headers
        headers = [
            "No.",
            "Student ID",
            "Name",
            "Status",
            "Session ID",
            "Class Name",
            "Session Type",
            "Expected Entry",
            "Entry Window Start",
            "Entry Window End",
            "Late Tolerance",
            "Expected Exit",
            "Exit Window Start",
            "Exit Window End",
            "Exception",
        ]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BORDER_THIN
            cell.alignment = ALIGN_CENTER
            ws.column_dimensions[get_column_letter(col_num)].width = 18

        # Freeze header row
        ws.freeze_panes = "A2"

        # Write data rows
        row_num = 2
        for student in expected_result.expected_students:
            if not student.sessions:
                # Student with no sessions
                row_data = [
                    row_num - 1,
                    student.student_id,
                    student.person_name or "N/A",
                    student.status.value.upper(),
                    "N/A", "N/A", "N/A",
                    "N/A", "N/A", "N/A", "N/A",
                    "N/A", "N/A", "N/A",
                    "No sessions scheduled"
                ]
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.font = CELL_FONT
                    cell.border = BORDER_THIN
                    cell.alignment = ALIGN_CENTER
                row_num += 1
                continue

            for session in student.sessions:
                exception_str = "None"
                if session.exception:
                    exception_str = f"{session.exception.exception_type.value}: {session.exception.description}"
                    if session.exception.new_entry_time is not None:
                        exception_str += f" (Entry: {self._format_seconds(session.exception.new_entry_time)})"
                    if session.exception.new_exit_time is not None:
                        exception_str += f" (Exit: {self._format_seconds(session.exception.new_exit_time)})"

                row_data = [
                    row_num - 1,
                    student.student_id,
                    student.person_name or "N/A",
                    student.status.value.upper(),
                    session.session_id,
                    session.class_name,
                    session.session_type.value,
                    self._format_seconds(session.effective_entry_time),
                    self._format_seconds(session.entry_window_start),
                    self._format_seconds(session.entry_window_end),
                    f"{session.late_tolerance}s",
                    self._format_seconds(session.effective_exit_time),
                    self._format_seconds(session.exit_window_start),
                    self._format_seconds(session.exit_window_end),
                    exception_str,
                ]

                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.font = CELL_FONT
                    cell.border = BORDER_THIN
                    cell.alignment = ALIGN_CENTER

                    # Color code by status
                    if col_num == 4:  # Status column
                        status_colors = {
                            "SCHEDULED": "C6EFCE",      # Light green
                            "LATER_START": "FFEB9C",      # Light yellow
                            "EARLIER_DEPARTURE": "DDEBF7", # Light blue
                            "CANCELLED": "FFC7CE",        # Light red
                            "EXCEPTION": "E2EFDA",        # Very light green
                            "NOT_SCHEDULED": "F2F2F2",    # Light gray
                            "HOLIDAY": "F2F2F2",          # Light gray
                        }
                        color = status_colors.get(student.status.value.upper(), "FFFFFF")
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

                row_num += 1

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col[0].column_letter].width = min(max(adjusted_width, 12), 40)

    def _format_seconds(self, seconds: int) -> str:
        """Format seconds from midnight as HH:MM:SS."""
        if seconds <= 0:
            return "N/A"
        hours = seconds // 3600
        minutes = (seconds % 3600) // 60
        secs = seconds % 60
        return f"{hours:02d}:{minutes:02d}:{secs:02d}"

    def _get_attendance_state_from_record(self, record: AttendanceRecord) -> AttendanceState:
        """
        Determine attendance state from AttendanceRecord.

        Note: This is a simplified mapping. In a real implementation,
        we would need to query the AttendanceDecision for the full state.
        """
        # Map derived states to attendance states
        if record.new_state == "inside":
            if record.is_in:
                return AttendanceState.PRESENT
            else:
                return AttendanceState.LEFT
        elif record.new_state == "outside":
            if record.is_out:
                return AttendanceState.LEFT
            else:
                return AttendanceState.ABSENT
        else:
            return AttendanceState.UNKNOWN

    def _find_matching_out_record(
        self,
        in_record: AttendanceRecord,
        all_records: List[AttendanceRecord]
    ) -> Optional[AttendanceRecord]:
        """Find matching OUT record for an IN record."""
        if not in_record.is_in:
            return None

        # Find OUT record with same global_observation_id and camera_id
        for record in all_records:
            if (record.is_out and
                record.global_observation_id == in_record.global_observation_id and
                record.camera_id == in_record.camera_id):
                return record

        return None

    def _calculate_duration(
        self,
        in_record: AttendanceRecord,
        out_record: Optional[AttendanceRecord]
    ) -> str:
        """Calculate duration between IN and OUT events."""
        if not out_record:
            return "N/A"

        duration_seconds = out_record.event_timestamp - in_record.event_timestamp
        hours = int(duration_seconds // 3600)
        minutes = int((duration_seconds % 3600) // 60)
        return f"{hours}h{minutes:02d}m"

    def _get_person_id(self, record: AttendanceRecord) -> str:
        """Get person ID from record."""
        if record.identity_certainty == IdentityCertainty.KNOWN and record.identity_candidate:
            return record.identity_candidate
        elif record.identity_certainty == IdentityCertainty.AMBIGUOUS:
            return "AMBIGUOUS"
        else:
            return "UNKNOWN"

    def _get_person_name(self, record: AttendanceRecord) -> str:
        """Get person name from record."""
        # In a real implementation, this would come from a person database
        # For now, we'll use the person ID or a placeholder
        person_id = self._get_person_id(record)
        if person_id == "UNKNOWN":
            return "Unknown"
        elif person_id == "AMBIGUOUS":
            return "Ambiguous Identity"
        else:
            return f"Person {person_id}"

    def _get_status_text(self, state: AttendanceState) -> str:
        """Get status text for attendance state."""
        status_map = {
            AttendanceState.PRESENT: "Present",
            AttendanceState.LATE: "Late",
            AttendanceState.LEFT: "Left Early",
            AttendanceState.ABSENT: "Absent",
            AttendanceState.UNKNOWN: "Unknown",
            AttendanceState.EXPECTED: "Expected"
        }
        return status_map.get(state, "Unknown")

    def _format_time(self, timestamp: float) -> str:
        """Format timestamp as HH:MM:SS."""
        dt = datetime.fromtimestamp(timestamp, self.timezone)
        return dt.strftime("%H:%M:%S")

    def _format_datetime(self, timestamp: float) -> str:
        """Format timestamp as YYYY-MM-DD HH:MM:SS."""
        dt = datetime.fromtimestamp(timestamp, self.timezone)
        return dt.strftime("%Y-%m-%d %H:%M:%S")

    def close(self) -> None:
        """Close the exporter."""
        if self.repository:
            self.repository.close()

    def __enter__(self) -> "DailyExcelExporter":
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        self.close()

def create_daily_excel_exporter(repository: Optional[AttendanceRepository] = None) -> DailyExcelExporter:
    """Factory function to create DailyExcelExporter."""
    return DailyExcelExporter(repository)