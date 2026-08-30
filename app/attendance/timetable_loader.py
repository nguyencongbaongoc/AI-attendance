"""
Phase 37A — Timetable Data Loader.

Production timetable data layer built on top of the existing Phase 26 contract.
Parses external timetable sources (Excel) into Phase 26 Timetable objects.
"""

from __future__ import annotations

import hashlib
import logging
from dataclasses import dataclass
from datetime import datetime, time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.attendance.timetable import (
    Timetable,
    TimetableEntry,
    SessionDay,
    SessionType,
    AttendanceState,
    generate_timetable_id,
    validate_timetable_entry,
)

logger = logging.getLogger(__name__)


# =============================================================================
# TIMETABLE COLUMN DEFINITIONS
# =============================================================================

# Canonical timetable columns (required)
REQUIRED_COLUMNS = [
    "student_id",      # Canonical student/business identifier
    "class_name",      # Class/session name
    "day",             # Day of week (monday, tuesday, etc.)
    "session_type",    # Session type (classroom, break, outside_lesson, lab, other, morning, afternoon, full_day, evening)
    "entry_time",      # Expected entry time (HH:MM:SS or seconds from midnight)
    "exit_time",       # Expected exit time (HH:MM:SS or seconds from midnight)
]

# Optional columns with defaults
OPTIONAL_COLUMNS = {
    "session_id": None,              # Auto-generated if not provided
    "person_name": None,             # Human-readable name
    "subject": None,                 # Subject name (e.g., Toán, GDTC, Hóa thực hành)
    "location": None,                # Physical location (e.g., Phòng 101, Sân thể dục)
    "expected_location": None,       # Expected location for outside lessons
    "outside_allowed": False,        # Whether students are allowed outside classroom
    "entry_window_start": None,      # Defaults to entry_time - 300s (5 min)
    "entry_window_end": None,        # Defaults to entry_time + 300s (5 min)
    "late_tolerance": 600,           # Default 10 minutes (seconds)
    "exit_window_start": None,       # Defaults to exit_time - 300s (5 min)
    "exit_window_end": None,         # Defaults to exit_time + 300s (5 min)
    "timetable_version": "1.0",      # Version string
}

# All supported columns in order
ALL_COLUMNS = REQUIRED_COLUMNS + list(OPTIONAL_COLUMNS.keys())


# =============================================================================
# VALIDATION ERRORS
# =============================================================================

@dataclass(frozen=True)
class TimetableValidationError:
    """Validation error for a timetable row."""
    row_number: int
    column: str
    value: Any
    message: str
    severity: str = "error"  # "error" or "warning"


@dataclass
class TimetableLoadResult:
    """Result of loading a timetable from a file."""
    success: bool
    timetable: Optional[Timetable] = None
    errors: List[TimetableValidationError] = None
    warnings: List[TimetableValidationError] = None
    rows_processed: int = 0
    rows_valid: int = 0
    rows_invalid: int = 0
    
    def __post_init__(self):
        if self.errors is None:
            self.errors = []
        if self.warnings is None:
            self.warnings = []


# =============================================================================
# TIME PARSING UTILITIES
# =============================================================================

def parse_time_value(value: Any) -> int:
    """
    Parse a time value to seconds from midnight.
    
    Accepts:
    - int/float: seconds from midnight
    - str: "HH:MM:SS", "HH:MM", or "H:MM:SS"
    - datetime.time: time object
    
    Returns:
        Seconds from midnight (int)
    """
    if value is None:
        return 0
    
    if isinstance(value, (int, float)):
        return int(value)
    
    if isinstance(value, time):
        return value.hour * 3600 + value.minute * 60 + value.second
    
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return 0
        
        # Try HH:MM:SS format
        parts = value.split(":")
        if len(parts) == 3:
            try:
                h, m, s = int(parts[0]), int(parts[1]), int(parts[2])
                if not (0 <= h <= 23 and 0 <= m <= 59 and 0 <= s <= 59):
                    raise ValueError(f"Invalid time components: {h}:{m}:{s}")
                return h * 3600 + m * 60 + s
            except ValueError:
                pass
        
        # Try HH:MM format
        if len(parts) == 2:
            try:
                h, m = int(parts[0]), int(parts[1])
                if not (0 <= h <= 23 and 0 <= m <= 59):
                    raise ValueError(f"Invalid time components: {h}:{m}")
                return h * 3600 + m * 60
            except ValueError:
                pass
        
        # Try plain seconds
        try:
            return int(float(value))
        except ValueError:
            pass
    
    raise ValueError(f"Cannot parse time value: {value}")


def parse_day_value(value: Any) -> SessionDay:
    """Parse a day value to SessionDay enum."""
    if isinstance(value, SessionDay):
        return value
    
    if isinstance(value, str):
        day_str = value.strip().lower()
        try:
            return SessionDay(day_str)
        except ValueError:
            # Try common abbreviations
            day_map = {
                "mon": SessionDay.MONDAY,
                "tue": SessionDay.TUESDAY,
                "wed": SessionDay.WEDNESDAY,
                "thu": SessionDay.THURSDAY,
                "fri": SessionDay.FRIDAY,
                "sat": SessionDay.SATURDAY,
                "sun": SessionDay.SUNDAY,
            }
            if day_str in day_map:
                return day_map[day_str]
            raise ValueError(f"Invalid day: {value}")
    
    raise ValueError(f"Cannot parse day value: {value}")


def parse_session_type_value(value: Any) -> SessionType:
    """Parse a session type value to SessionType enum."""
    if isinstance(value, SessionType):
        return value
    
    if isinstance(value, str):
        session_str = value.strip().lower()
        try:
            return SessionType(session_str)
        except ValueError:
            # Try common variations
            session_map = {
                "am": SessionType.MORNING,
                "pm": SessionType.AFTERNOON,
                "full": SessionType.FULL_DAY,
                "eve": SessionType.EVENING,
                "evening": SessionType.EVENING,
            }
            if session_str in session_map:
                return session_map[session_str]
            raise ValueError(f"Invalid session_type: {value}")
    
    raise ValueError(f"Cannot parse session_type value: {value}")


# =============================================================================
# TIMETABLE LOADER
# =============================================================================

class TimetableLoader:
    """
    Production timetable loader.
    
    Parses Excel files into Phase 26 Timetable objects with full validation.
    """
    
    def __init__(
        self,
        enrollment_person_ids: Optional[List[str]] = None,
        default_entry_window_seconds: int = 300,
        default_exit_window_seconds: int = 300,
        default_late_tolerance_seconds: int = 600,
    ):
        """
        Initialize the timetable loader.
        
        Args:
            enrollment_person_ids: List of valid person_ids from enrollment database.
                                   If provided, validates that all timetable student_ids exist.
            default_entry_window_seconds: Default entry window (±seconds from entry_time)
            default_exit_window_seconds: Default exit window (±seconds from exit_time)
            default_late_tolerance_seconds: Default late tolerance (seconds after entry_time)
        """
        self.enrollment_person_ids = set(enrollment_person_ids) if enrollment_person_ids else None
        self.default_entry_window_seconds = default_entry_window_seconds
        self.default_exit_window_seconds = default_exit_window_seconds
        self.default_late_tolerance_seconds = default_late_tolerance_seconds
    
    def load_from_excel(self, file_path: str) -> TimetableLoadResult:
        """
        Load timetable from Excel file.
        
        Args:
            file_path: Path to Excel file (.xlsx)
            
        Returns:
            TimetableLoadResult with timetable or errors
        """
        path = Path(file_path)
        if not path.exists():
            return TimetableLoadResult(
                success=False,
                errors=[TimetableValidationError(0, "file", file_path, "File not found")]
            )
        
        if not path.suffix.lower() in (".xlsx", ".xls"):
            return TimetableLoadResult(
                success=False,
                errors=[TimetableValidationError(0, "file", file_path, "Unsupported file format. Use .xlsx")]
            )
        
        try:
            workbook = openpyxl.load_workbook(path, data_only=True)
            sheet = workbook.active
            
            if sheet is None:
                return TimetableLoadResult(
                    success=False,
                    errors=[TimetableValidationError(0, "file", file_path, "No active sheet found")]
                )
            
            return self._parse_worksheet(sheet)
            
        except Exception as e:
            logger.exception(f"Failed to load timetable from {file_path}")
            return TimetableLoadResult(
                success=False,
                errors=[TimetableValidationError(0, "file", file_path, f"Load failed: {str(e)}")]
            )
    
    def _parse_worksheet(self, sheet: Worksheet) -> TimetableLoadResult:
        """Parse a worksheet into a Timetable."""
        # Find header row (first row with required columns)
        header_row = self._find_header_row(sheet)
        if header_row is None:
            return TimetableLoadResult(
                success=False,
                errors=[TimetableValidationError(0, "header", None, "Required columns not found in any row")]
            )
        
        # Map column names to indices
        column_map = self._build_column_map(sheet, header_row)
        
        # Validate required columns present
        missing_columns = [col for col in REQUIRED_COLUMNS if col not in column_map]
        if missing_columns:
            return TimetableLoadResult(
                success=False,
                errors=[TimetableValidationError(
                    header_row, "header", None, f"Missing required columns: {missing_columns}"
                )]
            )
        
        # Parse data rows
        entries = []
        errors = []
        warnings = []
        rows_processed = 0
        rows_valid = 0
        rows_invalid = 0
        
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            row_data = self._read_row(sheet, row_idx, column_map)
            
            # Skip empty rows
            if self._is_empty_row(row_data):
                continue
            
            rows_processed += 1
            
            # Parse and validate entry
            entry, entry_errors, entry_warnings = self._parse_entry(row_data, row_idx)
            
            if entry_errors:
                errors.extend(entry_errors)
                rows_invalid += 1
            else:
                entries.append(entry)
                rows_valid += 1
                warnings.extend(entry_warnings)
        
        # Cross-entry validation (duplicates, conflicts)
        cross_errors = self._validate_cross_entries(entries)
        errors.extend(cross_errors)
        
        # Enrollment validation
        if self.enrollment_person_ids is not None:
            enrollment_errors = self._validate_enrollment(entries)
            errors.extend(enrollment_errors)
        
        # Build timetable if no errors
        if not errors and entries:
            timetable = self._build_timetable(entries, sheet.title)
            return TimetableLoadResult(
                success=True,
                timetable=timetable,
                errors=errors,
                warnings=warnings,
                rows_processed=rows_processed,
                rows_valid=rows_valid,
                rows_invalid=rows_invalid,
            )
        elif not entries:
            errors.append(TimetableValidationError(0, "data", None, "No valid timetable entries found"))
            return TimetableLoadResult(
                success=False,
                errors=errors,
                warnings=warnings,
                rows_processed=rows_processed,
                rows_valid=rows_valid,
                rows_invalid=rows_invalid,
            )
        else:
            return TimetableLoadResult(
                success=False,
                errors=errors,
                warnings=warnings,
                rows_processed=rows_processed,
                rows_valid=rows_valid,
                rows_invalid=rows_invalid,
            )
    
    def _find_header_row(self, sheet: Worksheet) -> Optional[int]:
        """Find the header row containing required columns."""
        for row_idx in range(1, min(sheet.max_row + 1, 20)):  # Check first 20 rows
            row_values = []
            for cell in sheet[row_idx]:
                if cell.value is not None:
                    row_values.append(str(cell.value).strip().lower())
            
            # Check if all required columns are present
            if all(col in row_values for col in REQUIRED_COLUMNS):
                return row_idx
        
        return None
    
    def _build_column_map(self, sheet: Worksheet, header_row: int) -> Dict[str, int]:
        """Build mapping from column name to column index."""
        column_map = {}
        for cell in sheet[header_row]:
            if cell.value is not None:
                col_name = str(cell.value).strip().lower()
                column_map[col_name] = cell.column
        return column_map
    
    def _read_row(self, sheet: Worksheet, row_idx: int, column_map: Dict[str, int]) -> Dict[str, Any]:
        """Read a row of data using the column map."""
        row_data = {}
        for col_name, col_idx in column_map.items():
            cell = sheet.cell(row=row_idx, column=col_idx)
            row_data[col_name] = cell.value
        return row_data
    
    def _is_empty_row(self, row_data: Dict[str, Any]) -> bool:
        """Check if a row is empty (all required fields are None/empty)."""
        for col in REQUIRED_COLUMNS:
            value = row_data.get(col)
            if value is not None and str(value).strip() != "":
                return False
        return True
    
    def _parse_entry(
        self,
        row_data: Dict[str, Any],
        row_number: int,
    ) -> Tuple[Optional[TimetableEntry], List[TimetableValidationError], List[TimetableValidationError]]:
        """Parse a single row into a TimetableEntry."""
        errors = []
        warnings = []
        
        # Required fields
        student_id = row_data.get("student_id")
        class_name = row_data.get("class_name")
        day_value = row_data.get("day")
        session_type_value = row_data.get("session_type")
        entry_time_value = row_data.get("entry_time")
        exit_time_value = row_data.get("exit_time")
        
        # Validate required fields
        if not student_id or str(student_id).strip() == "":
            errors.append(TimetableValidationError(row_number, "student_id", student_id, "student_id is required"))
        
        if not class_name or str(class_name).strip() == "":
            errors.append(TimetableValidationError(row_number, "class_name", class_name, "class_name is required"))
        
        # Parse day
        day = None
        if day_value is not None:
            try:
                day = parse_day_value(day_value)
            except ValueError as e:
                errors.append(TimetableValidationError(row_number, "day", day_value, str(e)))
        else:
            errors.append(TimetableValidationError(row_number, "day", day_value, "day is required"))
        
        # Parse session_type
        session_type = SessionType.FULL_DAY
        if session_type_value is not None:
            try:
                session_type = parse_session_type_value(session_type_value)
            except ValueError as e:
                errors.append(TimetableValidationError(row_number, "session_type", session_type_value, str(e)))
        
        # Parse times
        entry_time = 0
        if entry_time_value is not None:
            try:
                entry_time = parse_time_value(entry_time_value)
            except ValueError as e:
                errors.append(TimetableValidationError(row_number, "entry_time", entry_time_value, str(e)))
        else:
            errors.append(TimetableValidationError(row_number, "entry_time", entry_time_value, "entry_time is required"))
        
        exit_time = 0
        if exit_time_value is not None:
            try:
                exit_time = parse_time_value(exit_time_value)
            except ValueError as e:
                errors.append(TimetableValidationError(row_number, "exit_time", exit_time_value, str(e)))
        else:
            errors.append(TimetableValidationError(row_number, "exit_time", exit_time_value, "exit_time is required"))
        
        # Validate entry < exit
        if entry_time >= exit_time:
            errors.append(TimetableValidationError(
                row_number, "entry_time/exit_time", f"{entry_time}/{exit_time}",
                "entry_time must be < exit_time"
            ))
        
        # Optional fields with defaults
        session_id = row_data.get("session_id")
        if not session_id or str(session_id).strip() == "":
            # Generate session_id from class_name and day
            session_id = f"{class_name}_{day.value}" if day else f"{class_name}_unknown"
            warnings.append(TimetableValidationError(
                row_number, "session_id", session_id,
                "session_id auto-generated from class_name and day",
                severity="warning"
            ))
        
        person_name = row_data.get("person_name")
        
        # Semantic fields (Phase 37D)
        subject = row_data.get("subject")
        location = row_data.get("location")
        expected_location = row_data.get("expected_location")
        outside_allowed = row_data.get("outside_allowed")
        if outside_allowed is not None:
            if isinstance(outside_allowed, str):
                outside_allowed = outside_allowed.strip().lower() in ("true", "1", "yes", "y")
            else:
                outside_allowed = bool(outside_allowed)
        else:
            outside_allowed = False
        
        # Entry window
        entry_window_start = row_data.get("entry_window_start")
        if entry_window_start is not None:
            try:
                entry_window_start = parse_time_value(entry_window_start)
            except ValueError as e:
                errors.append(TimetableValidationError(row_number, "entry_window_start", entry_window_start, str(e)))
        else:
            entry_window_start = max(0, entry_time - self.default_entry_window_seconds)
        
        entry_window_end = row_data.get("entry_window_end")
        if entry_window_end is not None:
            try:
                entry_window_end = parse_time_value(entry_window_end)
            except ValueError as e:
                errors.append(TimetableValidationError(row_number, "entry_window_end", entry_window_end, str(e)))
        else:
            entry_window_end = entry_time + self.default_entry_window_seconds
        
        # Late tolerance
        late_tolerance = row_data.get("late_tolerance")
        if late_tolerance is not None:
            try:
                late_tolerance = int(late_tolerance)
            except (ValueError, TypeError):
                errors.append(TimetableValidationError(row_number, "late_tolerance", late_tolerance, "Must be integer seconds"))
        else:
            late_tolerance = self.default_late_tolerance_seconds
        
        # Exit window
        exit_window_start = row_data.get("exit_window_start")
        if exit_window_start is not None:
            try:
                exit_window_start = parse_time_value(exit_window_start)
            except ValueError as e:
                errors.append(TimetableValidationError(row_number, "exit_window_start", exit_window_start, str(e)))
        else:
            exit_window_start = max(0, exit_time - self.default_exit_window_seconds)
        
        exit_window_end = row_data.get("exit_window_end")
        if exit_window_end is not None:
            try:
                exit_window_end = parse_time_value(exit_window_end)
            except ValueError as e:
                errors.append(TimetableValidationError(row_number, "exit_window_end", exit_window_end, str(e)))
        else:
            exit_window_end = exit_time + self.default_exit_window_seconds
        
        # Validate windows
        if entry_window_start > entry_window_end:
            errors.append(TimetableValidationError(
                row_number, "entry_window", f"{entry_window_start}/{entry_window_end}",
                "entry_window_start must be <= entry_window_end"
            ))
        
        if exit_window_start > exit_window_end:
            errors.append(TimetableValidationError(
                row_number, "exit_window", f"{exit_window_start}/{exit_window_end}",
                "exit_window_start must be <= exit_window_end"
            ))
        
        # Timetable version
        timetable_version = str(row_data.get("timetable_version", "1.0")).strip()
        
        # Generate entry_id
        entry_id = self._generate_entry_id(student_id, session_id, day.value if day else "unknown", row_number)
        
        if errors:
            return None, errors, warnings
        
        # Create TimetableEntry
        try:
            entry = TimetableEntry(
                entry_id=entry_id,
                person_id=str(student_id).strip(),
                session_id=str(session_id).strip(),
                session_type=session_type,
                day=day,
                class_name=str(class_name).strip(),
                person_name=str(person_name).strip() if person_name else None,
                # Semantic fields (Phase 37D)
                subject=str(subject).strip() if subject else None,
                location=str(location).strip() if location else None,
                expected_location=str(expected_location).strip() if expected_location else None,
                outside_allowed=outside_allowed,
                entry_time=entry_time,
                exit_time=exit_time,
                entry_window_start=entry_window_start,
                entry_window_end=entry_window_end,
                late_tolerance=late_tolerance,
                exit_window_start=exit_window_start,
                exit_window_end=exit_window_end,
                timetable_version=timetable_version,
            )
            
            # Validate using Phase 26 validator
            validation_error = validate_timetable_entry(entry)
            if validation_error:
                errors.append(TimetableValidationError(row_number, "entry", entry_id, validation_error))
                return None, errors, warnings
            
            return entry, errors, warnings
            
        except Exception as e:
            errors.append(TimetableValidationError(row_number, "entry", entry_id, f"Failed to create entry: {str(e)}"))
            return None, errors, warnings
    
    def _generate_entry_id(self, student_id: str, session_id: str, day: str, row_number: int) -> str:
        """Generate deterministic entry ID."""
        content = f"{student_id}:{session_id}:{day}:{row_number}"
        hash_suffix = hashlib.md5(content.encode()).hexdigest()[:8]
        return f"ENT-{hash_suffix}"
    
    def _validate_cross_entries(self, entries: List[TimetableEntry]) -> List[TimetableValidationError]:
        """Validate cross-entry constraints (duplicates, conflicts)."""
        errors = []
        
        # Check for duplicate entries (same person_id, day, session_id)
        seen = {}
        for entry in entries:
            key = (entry.person_id, entry.day, entry.session_id)
            if key in seen:
                errors.append(TimetableValidationError(
                    0, "duplicate", key,
                    f"Duplicate timetable entry for person_id={entry.person_id}, day={entry.day.value}, session_id={entry.session_id}"
                ))
            else:
                seen[key] = entry
        
        # Check for conflicting times (same person_id, day, overlapping sessions)
        person_day_entries = {}
        for entry in entries:
            key = (entry.person_id, entry.day)
            if key not in person_day_entries:
                person_day_entries[key] = []
            person_day_entries[key].append(entry)
        
        for (person_id, day), day_entries in person_day_entries.items():
            if len(day_entries) > 1:
                # Check for time overlaps
                for i, e1 in enumerate(day_entries):
                    for e2 in day_entries[i+1:]:
                        # Check if time ranges overlap
                        if not (e1.exit_time <= e2.entry_time or e2.exit_time <= e1.entry_time):
                            errors.append(TimetableValidationError(
                                0, "conflict", f"{e1.entry_id}/{e2.entry_id}",
                                f"Overlapping sessions for person_id={person_id} on {day.value}: "
                                f"{e1.session_id} ({e1.entry_time}-{e1.exit_time}) overlaps "
                                f"{e2.session_id} ({e2.entry_time}-{e2.exit_time})"
                            ))
        
        return errors
    
    def _validate_enrollment(self, entries: List[TimetableEntry]) -> List[TimetableValidationError]:
        """Validate that all student_ids exist in enrollment database."""
        errors = []
        
        if self.enrollment_person_ids is None:
            return errors
        
        for entry in entries:
            if entry.person_id not in self.enrollment_person_ids:
                errors.append(TimetableValidationError(
                    0, "enrollment", entry.person_id,
                    f"student_id '{entry.person_id}' not found in enrollment database"
                ))
        
        return errors
    
    def _build_timetable(self, entries: List[TimetableEntry], source_name: str) -> Timetable:
        """Build Timetable from validated entries."""
        # Generate deterministic timetable_id
        content = f"{source_name}:{len(entries)}:{sorted(e.person_id for e in entries)}"
        hash_suffix = hashlib.sha256(content.encode()).hexdigest()[:16]
        timetable_id = f"TTB-v1.0-{hash_suffix}"
        
        return Timetable(
            timetable_id=timetable_id,
            timetable_version="1.0",
            entries=entries,
        )


# =============================================================================
# CONVENIENCE FUNCTIONS
# =============================================================================

def load_timetable_from_excel(
    file_path: str,
    enrollment_person_ids: Optional[List[str]] = None,
) -> TimetableLoadResult:
    """
    Convenience function to load timetable from Excel.
    
    Args:
        file_path: Path to Excel file
        enrollment_person_ids: Optional list of valid person_ids from enrollment
        
    Returns:
        TimetableLoadResult
    """
    loader = TimetableLoader(enrollment_person_ids=enrollment_person_ids)
    return loader.load_from_excel(file_path)


def create_sample_timetable_excel(output_path: str) -> None:
    """Create a sample timetable Excel file for reference."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timetable"
    
    # Headers
    headers = ALL_COLUMNS
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
    
    # Sample data
    sample_data = [
        {
            "student_id": "HS001",
            "class_name": "Math 101",
            "day": "monday",
            "session_type": "morning",
            "entry_time": "07:30:00",
            "exit_time": "12:00:00",
            "session_id": "MATH101_MON",
            "person_name": "Student One",
            "late_tolerance": 600,
        },
        {
            "student_id": "HS002",
            "class_name": "Physics 101",
            "day": "monday",
            "session_type": "morning",
            "entry_time": "08:30:00",
            "exit_time": "13:00:00",
            "session_id": "PHYS101_MON",
            "person_name": "Student Two",
            "late_tolerance": 600,
        },
        {
            "student_id": "HS001",
            "class_name": "English 101",
            "day": "tuesday",
            "session_type": "afternoon",
            "entry_time": "13:00:00",
            "exit_time": "17:30:00",
            "session_id": "ENG101_TUE",
            "person_name": "Student One",
            "late_tolerance": 600,
        },
    ]
    
    for row_idx, data in enumerate(sample_data, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=data.get(header))
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max(max_length + 2, 12), 30)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width
    
    wb.save(output_path)
    logger.info(f"Sample timetable created at {output_path}")


# =============================================================================
# TIMETABLE VERSIONING
# =============================================================================

@dataclass(frozen=True)
class TimetableVersionInfo:
    """Version information for a loaded timetable."""
    timetable_id: str
    timetable_version: str
    source_file: str
    source_hash: str
    loaded_at: str
    entry_count: int
    person_ids: List[str]
    days_covered: List[str]
    sessions_covered: List[str]


def compute_timetable_version_info(timetable: Timetable, source_file: str) -> TimetableVersionInfo:
    """Compute version info for a loaded timetable."""
    # Compute source file hash
    source_hash = ""
    try:
        with open(source_file, "rb") as f:
            source_hash = hashlib.sha256(f.read()).hexdigest()[:16]
    except Exception:
        source_hash = "unknown"
    
    person_ids = sorted(set(e.person_id for e in timetable.entries))
    days_covered = sorted(set(e.day.value for e in timetable.entries))
    sessions_covered = sorted(set(e.session_id for e in timetable.entries))
    
    return TimetableVersionInfo(
        timetable_id=timetable.timetable_id,
        timetable_version=timetable.timetable_version,
        source_file=source_file,
        source_hash=source_hash,
        loaded_at=datetime.utcnow().isoformat() + "Z",
        entry_count=len(timetable.entries),
        person_ids=person_ids,
        days_covered=days_covered,
        sessions_covered=sessions_covered,
    )


def detect_timetable_change(
    current_version_info: TimetableVersionInfo,
    source_file: str,
) -> Tuple[bool, Optional[str]]:
    """
    Detect if timetable source file has changed.
    
    Returns:
        (changed, new_hash) - True if changed, new hash if computed
    """
    try:
        with open(source_file, "rb") as f:
            new_hash = hashlib.sha256(f.read()).hexdigest()[:16]
        
        changed = new_hash != current_version_info.source_hash
        return changed, new_hash
    except Exception:
        return False, None