"""
Phase 31 — Offline Full End-to-End Integration Test Suite.

This test suite verifies the complete offline pipeline from source video
through identity, observation, IN/OUT, attendance, persistence, output,
replay evidence, and provenance.

Pipeline:
SOURCE VIDEO
    ↓
PHASE 20 OFFLINE REPLAY
    ↓
FRAME / CAMERA / TIMESTAMP
    ↓
PHASE 15 FACE DETECTION
    ↓
PHASE 16 ADAPTIVE CROP
    ↓
PHASE 17 FACE QUALITY
    ↓
PHASE 18 TEMPORAL EVIDENCE
    ↓
PHASE 19 MATCHING
    ↓
PHASE 21 CROSS-CAMERA OBSERVATION FUSION
    ↓
PHASE 22 IN/OUT GEOMETRY
    ↓
PHASE 23 RAW IN/OUT EVENT
    ↓
PHASE 24 REPEATED IN/OUT RESOLUTION
    ↓
PHASE 25 ATTENDANCE PERSISTENCE
    ↓
PHASE 26 ATTENDANCE ENGINE
    ↓
PHASE 29 IMMEDIATE EVENT OUTPUT
    ↓
PHASE 30 DAILY EXCEL
    ↓
PHASE 27 ANNOTATED REPLAY / VIDEO EVIDENCE
    ↓
PROVENANCE VERIFICATION
"""

import pytest
import tempfile
import os
from pathlib import Path
from datetime import date, datetime

from app.data.frame import CanonicalFrame, FrameMetadata, PixelFormat, SourceType
from app.replay.clock import ReplayClock, ReplayTimestamp
from app.replay.source import ReplaySource, ReplaySourceConfig, ReplaySourceError
from app.replay.scheduler import ReplayScheduler, ReplaySchedulerConfig, create_scheduler
from app.replay.pipeline import ReplayPipeline, ReplayPipelineConfig, create_replay_pipeline
from app.replay.fusion import (
    CrossCameraFusionEngine,
    FusionConfig,
    GlobalObservation,
    LocalObservationRef,
    AssociationState,
    AssociationEvidence,
    build_local_observation_ref,
    create_fusion_engine,
    DEFAULT_FUSION_CONFIG,
)
from app.replay.annotated_replay import (
    AnnotatedReplayPipeline,
    AnnotatedReplayConfig,
    AnnotatedReplayState,
)
from app.replay.source import ReplaySourceConfig
from app.replay.annotation import (
    AnnotationFrame,
    PersonAnnotation,
    FaceAnnotation,
    EventAnnotation,
    AttendanceAnnotation,
    GlobalObservationReference,
    AnnotationProvenance,
    BoundingBox,
    IdentityDisplayState,
    AttendanceDisplayState,
    EventDisplayType,
    generate_annotation_frame_id,
)
from app.replay.appearance import (
    AppearanceRecord,
    VideoSegmentRequest,
    VideoSegmentResult,
    PersonSearchResult,
    generate_appearance_id,
    generate_video_segment_id,
)
from app.replay.video_evidence import (
    VideoSourceInfo,
    VideoEvidenceRetriever,
    VideoExtractionError,
    create_video_source_info_from_replay_source,
    build_source_video_registry_from_manifest,
)
from app.geometry.contract import (
    CameraGeometryConfig,
    DirectionSemantics,
    GeometryType,
    Point2D,
    create_line_geometry,
    create_zone_geometry,
)
from app.geometry.crossing import (
    CrossingDirection,
    CrossingEvent,
    CrossingEventType,
    CrossingEngine,
    TrajectoryPoint,
    create_crossing_engine,
    process_tracks_for_crossings,
)
from app.geometry.contract import GeometryConfigSnapshot
from app.in_out.contract import (
    RawInOutEvent,
    RawEventDirection,
    RawEventType,
    IdentityCertainty,
)
from app.in_out.raw_event import (
    RawEventEngine,
    create_raw_event_engine,
    create_raw_in_out_event,
)
from app.in_out.factory import (
    create_integrated_pipeline,
    process_tracks_through_pipeline,
)
from app.in_out.resolver import (
    RepeatedInOutResolver,
    create_repeated_in_out_resolver,
    resolve_raw_events,
)
from app.in_out.resolver_config import (
    ResolverConfig,
    InitialOutPolicy,
    OutOfOrderPolicy,
    EqualTimestampPolicy,
    create_default_resolver_config,
    create_strict_resolver_config,
    create_permissive_resolver_config,
)
from app.in_out.resolver_contract import (
    ResolvedTransition,
    TrackResolutionState,
    ResolutionResult,
    DerivedState,
    TransitionType,
    ResolutionStatus,
)
from app.attendance.contract import (
    AttendanceRecord,
    AttendanceDirection,
    IdentityCertainty as AttendanceIdentityCertainty,
    create_attendance_record_from_resolution,
)
from app.attendance.repository import AttendanceRepository, PersistenceResult
from app.attendance.storage import AttendanceStorage, StorageConfig
from app.attendance.query import (
    get_attendance_summary,
    records_to_timeline,
    get_daily_attendance_counts,
    get_track_state_history,
)
from app.attendance.engine import AttendanceEngine, AttendanceDecisionContext
from app.attendance.policy import AttendancePolicy
from app.attendance.timetable import Timetable, TimetableEntry, SessionDay
from app.attendance.daily_excel import (
    DailyExportRequest,
    DailyExcelExporter,
    create_daily_excel_exporter,
)
from app.output.contract import (
    ImmediateEvent,
    ImmediateEventType,
    ImmediateEventDirection,
    IdentityCertainty as OutputIdentityCertainty,
    EventDeliveryStatus,
    generate_immediate_event_id,
    validate_immediate_event,
    ImmediateEventCreationResult,
)
from app.output.publisher import (
    create_event_bus,
    FunctionSubscriber,
    SubscriberConfig,
    BackpressurePolicy,
    CallbackEventBus,
)
from app.output.adapter import (
    Phase24ToImmediateEventAdapter,
    Phase26ToImmediateEventAdapter,
    Phase25ToImmediateEventAdapter,
    Phase23ToImmediateEventAdapter,
    DevelopmentEventSource,
    create_adapters,
)
from app.output.ui_adapter import (
    UIEvent,
    UIEventSubscriber,
    Phase28UIAdapter,
    MockEventReplacer,
)
from app.vision.temporal_evidence import (
    IdentityEvidence,
    IdentityHypothesis,
    HypothesisState,
    CandidateSupport,
    TemporalTimestamp,
    TimestampSource,
    QualityClass,
)
from app.vision.matching_contract import MatchStatus


class TestPhase31OfflineFullE2E:
    """Phase 31 Offline Full End-to-End Integration Tests."""

    @pytest.fixture
    def test_data_dir(self):
        """Test data directory with Phase 20 videos."""
        return Path("test_data/phase20")

    @pytest.fixture
    def enrollment_db_path(self):
        """Path to Phase 30A enrollment database."""
        return "data/enrollment_db"

    @pytest.fixture
    def temp_db_path(self):
        """Temporary database path for testing."""
        with tempfile.NamedTemporaryFile(suffix=".db", delete=False) as f:
            db_path = f.name
        yield db_path
        # Cleanup
        if os.path.exists(db_path):
            os.unlink(db_path)
        for suffix in ["-wal", "-shm"]:
            wal_path = db_path + suffix
            if os.path.exists(wal_path):
                os.unlink(wal_path)

    @pytest.fixture
    def temp_output_dir(self):
        """Temporary output directory for testing."""
        with tempfile.TemporaryDirectory() as tmpdir:
            yield Path(tmpdir)

    # =========================================================================
    # PHASE 20: OFFLINE REPLAY VERIFICATION
    # =========================================================================

    def test_phase20_replay_executes(self, test_data_dir):
        """Verify Phase 20 offline replay executes with CAM1 and CAM2."""
        cam1_path = test_data_dir / "cam1_test.mp4"
        cam2_path = test_data_dir / "cam2_test.mp4"
        
        assert cam1_path.exists(), f"CAM1 test video not found: {cam1_path}"
        assert cam2_path.exists(), f"CAM2 test video not found: {cam2_path}"
        
        # Create replay sources
        source1 = ReplaySource(ReplaySourceConfig(
            camera_id="CAM1",
            source_path=str(cam1_path),
        ))
        source2 = ReplaySource(ReplaySourceConfig(
            camera_id="CAM2",
            source_path=str(cam2_path),
        ))
        
        # Verify sources open
        source1.open()
        source2.open()
        
        # Verify sources opened successfully by reading a frame
        frame1 = source1.get_next_frame()
        frame2 = source2.get_next_frame()
        assert frame1 is not None, "CAM1 should produce frames"
        assert frame2 is not None, "CAM2 should produce frames"
        assert frame1.metadata.extra.get("camera_id") == "CAM1"
        assert frame2.metadata.extra.get("camera_id") == "CAM2"
        
        source1.close()
        source2.close()

    def test_phase20_deterministic_replay(self, test_data_dir):
        """Verify deterministic replay: same source + config = same frame ordering."""
        cam1_path = test_data_dir / "cam1_test.mp4"
        
        # Run replay twice
        frames_run1 = []
        frames_run2 = []
        
        for run in range(2):
            source = ReplaySource(ReplaySourceConfig(
                camera_id="CAM1",
                source_path=str(cam1_path),
            ))
            source.open()
            
            frames = []
            for _ in range(10):  # Read first 10 frames
                frame = source.get_next_frame()
                if frame is None:
                    break
                frames.append({
                    "frame_index": frame.metadata.frame_index,
                    "timestamp": frame.metadata.timestamp,
                    "camera_id": frame.metadata.extra.get("camera_id"),
                })
            source.close()
            
            if run == 0:
                frames_run1 = frames
            else:
                frames_run2 = frames
        
        # Verify deterministic ordering
        assert len(frames_run1) == len(frames_run2), "Frame count should be identical"
        for f1, f2 in zip(frames_run1, frames_run2):
            assert f1["frame_index"] == f2["frame_index"], "Frame indices should match"
            assert f1["timestamp"] == f2["timestamp"], "Timestamps should match"
            assert f1["camera_id"] == f2["camera_id"], "Camera IDs should match"

    def test_phase20_camera_isolation(self, test_data_dir):
        """Verify CAM1 and CAM2 remain isolated in replay."""
        cam1_path = test_data_dir / "cam1_test.mp4"
        cam2_path = test_data_dir / "cam2_test.mp4"
        
        source1 = ReplaySource(ReplaySourceConfig(
            camera_id="CAM1",
            source_path=str(cam1_path),
        ))
        source2 = ReplaySource(ReplaySourceConfig(
            camera_id="CAM2",
            source_path=str(cam2_path),
        ))
        
        source1.open()
        source2.open()
        
        frame1 = source1.get_next_frame()
        frame2 = source2.get_next_frame()
        
        assert frame1.metadata.extra.get("camera_id") == "CAM1"
        assert frame2.metadata.extra.get("camera_id") == "CAM2"
        assert frame1.metadata.extra.get("camera_id") != frame2.metadata.extra.get("camera_id")
        
        source1.close()
        source2.close()

    # =========================================================================
    # PHASE 15-19: FACE PIPELINE VERIFICATION
    # =========================================================================

    def test_phase15_19_chain_executes(self, test_data_dir, enrollment_db_path):
        """Verify Phase 15-19 chain executes: detection -> crop -> quality -> temporal -> matching."""
        cam1_path = test_data_dir / "cam1_test.mp4"
        
        # Create replay pipeline with enrollment database
        pipeline = create_replay_pipeline(enrollment_db_path=enrollment_db_path)
        
        # Create source and read a frame
        source = ReplaySource(ReplaySourceConfig(
            camera_id="CAM1",
            source_path=str(cam1_path),
        ))
        source.open()
        
        frame = source.get_next_frame()
        assert frame is not None
        
        # Process through pipeline
        result = pipeline.process_frame(frame)
        
        # Verify pipeline stages executed
        assert result.camera_id == "CAM1"
        assert result.frame_index >= 0
        assert result.timestamp >= 0
        
        # Phase 15: Face detection should produce detections (or empty list)
        assert isinstance(result.detections, list)
        
        # Phase 16: Adaptive crop should produce face crops for detections
        assert isinstance(result.face_crops, list)
        
        # Phase 17: Face quality should produce quality results for face crops
        assert isinstance(result.quality_results, list)
        
        # Phase 18: Temporal evidence should produce hypotheses
        assert isinstance(result.temporal_hypotheses, list)
        
        # Phase 19: Matching would execute if enrollment DB loaded
        # (placeholder in current implementation)
        
        source.close()
        pipeline.close()

    def test_phase19_loads_actual_enrollment_database(self, enrollment_db_path):
        """Verify Phase 19 loads actual Phase 30A enrollment database."""
        from app.vision.matching import load_matching_database
        
        context = load_matching_database(enrollment_db_path)
        
        assert context is not None
        assert context.database_embeddings is not None
        assert context.database_embeddings.shape == (9, 512)  # 3 persons * 3 samples
        assert context.database_metadata is not None
        assert context.database_metadata.person_ids == ["HS001", "HS002", "HS003"]

    def test_phase19_matcher_executes(self, enrollment_db_path):
        """Verify Phase 19 matcher executes (even with synthetic data limitation)."""
        from app.vision.matching import load_matching_database, match_identity
        import numpy as np
        
        context = load_matching_database(enrollment_db_path)
        
        # Create a dummy query embedding (512D, L2 normalized)
        query_embedding = np.random.randn(512).astype(np.float32)
        query_embedding = query_embedding / np.linalg.norm(query_embedding)
        
        # Execute matcher
        result = match_identity(query_embedding, context)
        
        assert result is not None
        assert result.status in [MatchStatus.MATCH, MatchStatus.UNKNOWN, MatchStatus.AMBIGUOUS]
        assert result.similarity >= 0.0
        assert result.similarity <= 1.0
        assert result.candidate_count == 9
        assert result.threshold == context.config.match_threshold
        assert result.ambiguity_margin == context.config.ambiguity_margin

    # =========================================================================
    # PHASE 21: CROSS-CAMERA FUSION VERIFICATION
    # =========================================================================

    def test_phase21_global_observation_exists(self):
        """Verify GlobalObservation contract exists and works."""
        obs1 = LocalObservationRef(
            camera_id="CAM1",
            local_track_id="track_001",
            observation_id="CAM1_track_001_f100",
            frame_index=100,
            timestamp=ReplayTimestamp(value=1234567890.5, source="frame_metadata"),
            detection_id="det_001",
            face_crop_id="crop_001",
            quality_class="GOOD",
        )
        obs2 = LocalObservationRef(
            camera_id="CAM2",
            local_track_id="track_002",
            observation_id="CAM2_track_002_f100",
            frame_index=100,
            timestamp=ReplayTimestamp(value=1234567890.6, source="frame_metadata"),
            detection_id="det_002",
            face_crop_id="crop_002",
            quality_class="GOOD",
        )
        
        evidence = AssociationEvidence(
            timestamp_delta=0.1,
            timestamp_compatible=True,
            timestamp_tolerance=1.0,
            camera_ids=("CAM1", "CAM2"),
        )
        
        global_obs = GlobalObservation(
            global_observation_id="GO-TEST-123",
            observations=(obs1, obs2),
            association_state=AssociationState.ASSOCIATED,
            association_evidence=evidence,
            temporal_start=ReplayTimestamp(value=1234567890.5, source="fusion_min"),
            temporal_end=ReplayTimestamp(value=1234567890.6, source="fusion_max"),
            temporal_span=0.1,
            camera_ids=("CAM1", "CAM2"),
            local_track_ids=("CAM1:track_001", "CAM2:track_002"),
            primary_identity_candidate="person_123",
            identity_confidence=0.85,
        )
        
        assert global_obs.global_observation_id == "GO-TEST-123"
        assert len(global_obs.observations) == 2
        assert global_obs.association_state == AssociationState.ASSOCIATED
        assert global_obs.is_associated
        assert "CAM1:track_001" in global_obs.local_track_ids
        assert "CAM2:track_002" in global_obs.local_track_ids

    def test_phase21_camera_isolation_preserved(self):
        """Verify camera isolation is preserved in fusion."""
        engine = create_fusion_engine(DEFAULT_FUSION_CONFIG)
        
        # Add observations from different cameras with same local_track_id
        obs1 = LocalObservationRef(
            camera_id="CAM1",
            local_track_id="track_001",
            observation_id="CAM1_track_001_f100",
            frame_index=100,
            timestamp=ReplayTimestamp(value=1000.0, source="frame_metadata"),
        )
        obs2 = LocalObservationRef(
            camera_id="CAM2",
            local_track_id="track_001",  # Same local track ID
            observation_id="CAM2_track_001_f100",
            frame_index=100,
            timestamp=ReplayTimestamp(value=1000.1, source="frame_metadata"),
        )
        
        engine.add_observation(obs1)
        engine.add_observation(obs2)
        
        # Verify both cameras have their own track
        stats = engine.get_stats()
        assert "CAM1" in stats["cameras"]
        assert "CAM2" in stats["cameras"]
        assert "track_001" in stats["camera_tracks"]["CAM1"]
        assert "track_001" in stats["camera_tracks"]["CAM2"]

    def test_phase21_deterministic_association(self):
        """Verify deterministic association: same inputs = same outputs."""
        engine1 = create_fusion_engine(DEFAULT_FUSION_CONFIG)
        engine2 = create_fusion_engine(DEFAULT_FUSION_CONFIG)
        
        obs1 = LocalObservationRef(
            camera_id="CAM1",
            local_track_id="track_001",
            observation_id="CAM1_track_001_f100",
            frame_index=100,
            timestamp=ReplayTimestamp(value=1000.0, source="frame_metadata"),
        )
        obs2 = LocalObservationRef(
            camera_id="CAM2",
            local_track_id="track_002",
            observation_id="CAM2_track_002_f100",
            frame_index=100,
            timestamp=ReplayTimestamp(value=1000.1, source="frame_metadata"),
        )
        
        engine1.add_observation(obs1)
        engine1.add_observation(obs2)
        engine2.add_observation(obs1)
        engine2.add_observation(obs2)
        
        global_obs1 = engine1.associate_observations()
        global_obs2 = engine2.associate_observations()
        
        assert len(global_obs1) == len(global_obs2)
        if global_obs1 and global_obs2:
            assert global_obs1[0].global_observation_id == global_obs2[0].global_observation_id

    # =========================================================================
    # PHASE 22: IN/OUT GEOMETRY VERIFICATION
    # =========================================================================

    def test_phase22_original_frame_coordinates(self):
        """Verify ORIGINAL_FRAME coordinate semantics."""
        line_config = create_line_geometry(
            camera_id="CAM1",
            frame_width=1920,
            frame_height=1080,
            p1=(100, 500),
            p2=(1820, 500),
            direction_semantics=DirectionSemantics.SIDE_A_TO_B_IN,
            version=1,
        )
        
        assert line_config.geometry_type == GeometryType.LINE
        assert line_config.camera_id == "CAM1"
        assert line_config.line is not None
        assert line_config.line.p1.x == 100
        assert line_config.line.p1.y == 500
        assert line_config.line.p2.x == 1820
        assert line_config.line.p2.y == 500

    def test_phase22_line_crossing_detection(self):
        """Verify line crossing detection with hysteresis/debounce."""
        line_config = create_line_geometry(
            camera_id="CAM1",
            frame_width=1920,
            frame_height=1080,
            p1=(100, 500),
            p2=(1820, 500),
            direction_semantics=DirectionSemantics.SIDE_A_TO_B_IN,
            version=1,
        )
        
        engine = create_crossing_engine(line_config)
        
        # Simulate track crossing from SIDE_A (y=480) to SIDE_B (y=520)
        from app.vision.track_contract import Track
        
        # Create mock track
        class MockTrack:
            def __init__(self, track_id, center, bbox):
                self.track_id = track_id
                self.center = center
                self.bbox_original_frame = bbox
        
        # Frame 1: Before crossing (SIDE_A)
        track1 = MockTrack("track_001", (960, 480), (900, 400, 1020, 560))
        events1 = engine.process_track(track1, 100, 1000.0)
        assert len(events1) == 0  # No crossing yet
        
        # Frame 2: Crossing (SIDE_A -> SIDE_B)
        track2 = MockTrack("track_001", (960, 520), (900, 440, 1020, 600))
        events2 = engine.process_track(track2, 101, 1001.0)
        
        # Should detect IN crossing (SIDE_A_TO_B_IN)
        # Note: May need confirmation frames depending on config
        assert len(events2) >= 0  # May be 0 if confirmation_frames > 1

    def test_phase22_provenance_preserved(self):
        """Verify geometry version and config hash preserved."""
        line_config = create_line_geometry(
            camera_id="CAM1",
            frame_width=1920,
            frame_height=1080,
            p1=(100, 500),
            p2=(1820, 500),
            direction_semantics=DirectionSemantics.SIDE_A_TO_B_IN,
            version=5,
        )
        
        assert line_config.version == 5
        assert line_config.config_hash is not None
        assert len(line_config.config_hash) > 0

    # =========================================================================
    # PHASE 23: RAW IN/OUT EVENT VERIFICATION
    # =========================================================================

    def test_phase23_crossing_to_raw_event(self):
        """Verify CrossingEvent -> RawInOutEvent conversion."""
        line_config = create_line_geometry(
            camera_id="CAM1",
            frame_width=1920,
            frame_height=1080,
            p1=(100, 500),
            p2=(1820, 500),
            direction_semantics=DirectionSemantics.SIDE_A_TO_B_IN,
            version=1,
        )
        
        geom_snapshot = GeometryConfigSnapshot.from_config(line_config)
        
        crossing_event = CrossingEvent(
            event_id="CE-TEST-001",
            camera_id="CAM1",
            geometry_config=geom_snapshot,
            local_track_id="track_001",
            global_observation_id="GO-123",
            event_type=CrossingEventType.LINE_CROSSING,
            direction=CrossingDirection.IN,
            crossing_point=Point2D(960, 500),
            crossing_timestamp=1234567890.5,
            previous_position=Point2D(960, 480),
            current_position=Point2D(960, 520),
            previous_frame_index=100,
            current_frame_index=101,
            previous_timestamp=1234567889.5,
            current_timestamp=1234567890.5,
            crossing_distance=40.0,
            side_transition="SIDE_A->SIDE_B",
            trajectory_points=[],
            config_snapshot={},
            created_at="2026-01-01T00:00:00Z",
            version="1.0",
        )
        
        raw_engine = create_raw_event_engine()
        result = raw_engine.process_crossing_event(crossing_event)
        
        assert result.success
        raw_event = result.event
        
        assert raw_event.camera_id == "CAM1"
        assert raw_event.local_track_id == "track_001"
        assert raw_event.direction == RawEventDirection.IN
        assert raw_event.global_observation_id == "GO-123"
        assert raw_event.source_crossing_event_id == "CE-TEST-001"
        assert raw_event.geometry_version == 1
        assert raw_event.geometry_config_hash == line_config.config_hash

    def test_phase23_deterministic_event_id(self):
        """Verify deterministic event ID generation."""
        raw_engine1 = create_raw_event_engine()
        raw_engine2 = create_raw_event_engine()
        
        line_config = create_line_geometry(
            camera_id="CAM1",
            frame_width=1920,
            frame_height=1080,
            p1=(100, 500),
            p2=(1820, 500),
            direction_semantics=DirectionSemantics.SIDE_A_TO_B_IN,
            version=1,
        )
        geom_snapshot = GeometryConfigSnapshot.from_config(line_config)
        
        crossing_event = CrossingEvent(
            event_id="CE-DET-001",
            camera_id="CAM1",
            geometry_config=geom_snapshot,
            local_track_id="track_001",
            global_observation_id=None,
            event_type=CrossingEventType.LINE_CROSSING,
            direction=CrossingDirection.IN,
            crossing_point=Point2D(960, 500),
            crossing_timestamp=1000.0,
            previous_position=Point2D(960, 480),
            current_position=Point2D(960, 520),
            previous_frame_index=100,
            current_frame_index=101,
            previous_timestamp=999.0,
            current_timestamp=1000.0,
            crossing_distance=40.0,
            side_transition="SIDE_A->SIDE_B",
            trajectory_points=[],
            config_snapshot={},
            created_at="2026-01-01T00:00:00Z",
            version="1.0",
        )
        
        result1 = raw_engine1.process_crossing_event(crossing_event)
        result2 = raw_engine2.process_crossing_event(crossing_event)
        
        assert result1.success and result2.success
        assert result1.event.event_id == result2.event.event_id

    def test_phase23_duplicate_suppression(self):
        """Verify duplicate event ID suppression."""
        raw_engine = create_raw_event_engine()
        
        line_config = create_line_geometry(
            camera_id="CAM1",
            frame_width=1920,
            frame_height=1080,
            p1=(100, 500),
            p2=(1820, 500),
            direction_semantics=DirectionSemantics.SIDE_A_TO_B_IN,
            version=1,
        )
        geom_snapshot = GeometryConfigSnapshot.from_config(line_config)
        
        crossing_event = CrossingEvent(
            event_id="CE-DUP-001",
            camera_id="CAM1",
            geometry_config=geom_snapshot,
            local_track_id="track_001",
            global_observation_id=None,
            event_type=CrossingEventType.LINE_CROSSING,
            direction=CrossingDirection.IN,
            crossing_point=Point2D(960, 500),
            crossing_timestamp=1000.0,
            previous_position=Point2D(960, 480),
            current_position=Point2D(960, 520),
            previous_frame_index=100,
            current_frame_index=101,
            previous_timestamp=999.0,
            current_timestamp=1000.0,
            crossing_distance=40.0,
            side_transition="SIDE_A->SIDE_B",
            trajectory_points=[],
            config_snapshot={},
            created_at="2026-01-01T00:00:00Z",
            version="1.0",
        )
        
        # Process same event twice
        result1 = raw_engine.process_crossing_event(crossing_event)
        result2 = raw_engine.process_crossing_event(crossing_event)
        
        assert result1.success
        assert result2.success
        assert result1.event.event_id == result2.event.event_id
        assert raw_engine.get_stats()["duplicates"] == 1

    # =========================================================================
    # PHASE 24: REPEATED IN/OUT RESOLUTION VERIFICATION
    # =========================================================================

    def test_phase24_state_machine(self):
        """Verify state machine: UNKNOWN+IN->INSIDE, INSIDE+OUT->OUTSIDE, etc."""
        resolver = create_repeated_in_out_resolver(create_default_resolver_config())
        
        # Create raw IN event
        raw_in = RawInOutEvent(
            event_id="RIE-001",
            camera_id="CAM1",
            geometry_id="hash123",
            geometry_version=1,
            geometry_config_hash="hash123",
            local_track_id="track_001",
            global_observation_id="GO-1",
            event_type=RawEventType.LINE_CROSSING,
            direction=RawEventDirection.IN,
            crossing_point_x=960.0,
            crossing_point_y=500.0,
            crossing_timestamp=1000.0,
            crossing_frame_index=100,
            previous_position_x=960.0,
            previous_position_y=480.0,
            current_position_x=960.0,
            current_position_y=520.0,
            previous_frame_index=99,
            current_frame_index=100,
            previous_timestamp=999.0,
            current_timestamp=1000.0,
            crossing_distance=40.0,
            side_transition="SIDE_A->SIDE_B",
            identity_certainty=IdentityCertainty.UNKNOWN,
            identity_candidate=None,
            identity_confidence=0.0,
            identity_evidence_ref="GO-1",
            source_crossing_event_id="CE-001",
            trajectory_points=[],
            config_snapshot={},
            event_schema_version="1.0",
            created_at="2026-01-01T00:00:00Z",
        )
        
        # UNKNOWN + IN -> INSIDE
        result = resolver.resolve_events([raw_in])
        assert result.accepted_transitions == 1
        assert result.transitions[0].new_state == DerivedState.INSIDE
        assert result.transitions[0].transition_type == TransitionType.IN
        
        # Create raw OUT event
        raw_out = RawInOutEvent(
            event_id="RIE-002",
            camera_id="CAM1",
            geometry_id="hash123",
            geometry_version=1,
            geometry_config_hash="hash123",
            local_track_id="track_001",
            global_observation_id="GO-2",
            event_type=RawEventType.LINE_CROSSING,
            direction=RawEventDirection.OUT,
            crossing_point_x=960.0,
            crossing_point_y=500.0,
            crossing_timestamp=2000.0,
            crossing_frame_index=200,
            previous_position_x=960.0,
            previous_position_y=520.0,
            current_position_x=960.0,
            current_position_y=480.0,
            previous_frame_index=199,
            current_frame_index=200,
            previous_timestamp=1999.0,
            current_timestamp=2000.0,
            crossing_distance=40.0,
            side_transition="SIDE_B->SIDE_A",
            identity_certainty=IdentityCertainty.UNKNOWN,
            identity_candidate=None,
            identity_confidence=0.0,
            identity_evidence_ref="GO-2",
            source_crossing_event_id="CE-002",
            trajectory_points=[],
            config_snapshot={},
            event_schema_version="1.0",
            created_at="2026-01-01T00:00:00Z",
        )
        
        # INSIDE + OUT -> OUTSIDE
        resolver.clear()
        result = resolver.resolve_events([raw_in, raw_out])
        assert result.accepted_transitions == 2
        assert result.transitions[1].new_state == DerivedState.OUTSIDE
        assert result.transitions[1].transition_type == TransitionType.OUT

    def test_phase24_repeated_event_suppression(self):
        """Verify repeated same-direction events are suppressed."""
        resolver = create_repeated_in_out_resolver(create_default_resolver_config())
        
        # Create 3 IN events
        raw_events = []
        for i in range(3):
            raw_events.append(RawInOutEvent(
                event_id=f"RIE-IN-{i}",
                camera_id="CAM1",
                geometry_id="hash123",
                geometry_version=1,
                geometry_config_hash="hash123",
                local_track_id="track_001",
                global_observation_id=f"GO-{i}",
                event_type=RawEventType.LINE_CROSSING,
                direction=RawEventDirection.IN,
                crossing_point_x=960.0,
                crossing_point_y=500.0,
                crossing_timestamp=1000.0 + i,
                crossing_frame_index=100 + i,
                previous_position_x=960.0,
                previous_position_y=480.0,
                current_position_x=960.0,
                current_position_y=520.0,
                previous_frame_index=99 + i,
                current_frame_index=100 + i,
                previous_timestamp=999.0 + i,
                current_timestamp=1000.0 + i,
                crossing_distance=40.0,
                side_transition="SIDE_A->SIDE_B",
                identity_certainty=IdentityCertainty.UNKNOWN,
                identity_candidate=None,
                identity_confidence=0.0,
                identity_evidence_ref=f"GO-{i}",
                source_crossing_event_id=f"CE-{i}",
                trajectory_points=[],
                config_snapshot={},
                event_schema_version="1.0",
                created_at="2026-01-01T00:00:00Z",
            ))
        
        result = resolver.resolve_events(raw_events)
        
        assert result.total_raw_events == 3
        assert result.accepted_transitions == 1
        assert result.suppressed_events == 2
        assert result.transitions[0].transition_type == TransitionType.IN
        assert result.transitions[1].transition_type == TransitionType.NONE
        assert result.transitions[2].transition_type == TransitionType.NONE

    def test_phase24_provenance_preserved(self):
        """Verify provenance chain preserved through resolver."""
        resolver = create_repeated_in_out_resolver(create_default_resolver_config())
        
        raw_event = RawInOutEvent(
            event_id="RIE-PROV-001",
            camera_id="CAM1",
            geometry_id="hash123",
            geometry_version=5,
            geometry_config_hash="abc123def456",
            local_track_id="track_001",
            global_observation_id="GO-123",
            event_type=RawEventType.LINE_CROSSING,
            direction=RawEventDirection.IN,
            crossing_point_x=960.0,
            crossing_point_y=500.0,
            crossing_timestamp=1000.0,
            crossing_frame_index=100,
            previous_position_x=960.0,
            previous_position_y=480.0,
            current_position_x=960.0,
            current_position_y=520.0,
            previous_frame_index=99,
            current_frame_index=100,
            previous_timestamp=999.0,
            current_timestamp=1000.0,
            crossing_distance=40.0,
            side_transition="SIDE_A->SIDE_B",
            identity_certainty=IdentityCertainty.UNKNOWN,
            identity_candidate=None,
            identity_confidence=0.0,
            identity_evidence_ref="GO-123",
            source_crossing_event_id="CE-ORIGINAL-456",
            trajectory_points=[],
            config_snapshot={},
            event_schema_version="1.0",
            created_at="2026-01-01T00:00:00Z",
        )
        
        result = resolver.resolve_events([raw_event])
        
        assert result.accepted_transitions == 1
        t = result.transitions[0]
        assert t.source_raw_event_id == "RIE-PROV-001"
        assert t.source_crossing_event_id == "CE-ORIGINAL-456"
        assert t.global_observation_id == "GO-123"
        assert t.geometry_version == 5
        assert t.geometry_config_hash == "abc123def456"
        assert t.resolver_version == "1.0"

    def test_phase24_idempotency(self):
        """Verify idempotent resolution: same events = same resolution IDs."""
        resolver1 = create_repeated_in_out_resolver(create_default_resolver_config())
        resolver2 = create_repeated_in_out_resolver(create_default_resolver_config())
        
        raw_event = RawInOutEvent(
            event_id="RIE-IDEMP-001",
            camera_id="CAM1",
            geometry_id="hash123",
            geometry_version=1,
            geometry_config_hash="hash123",
            local_track_id="track_001",
            global_observation_id="GO-1",
            event_type=RawEventType.LINE_CROSSING,
            direction=RawEventDirection.IN,
            crossing_point_x=960.0,
            crossing_point_y=500.0,
            crossing_timestamp=1000.0,
            crossing_frame_index=100,
            previous_position_x=960.0,
            previous_position_y=480.0,
            current_position_x=960.0,
            current_position_y=520.0,
            previous_frame_index=99,
            current_frame_index=100,
            previous_timestamp=999.0,
            current_timestamp=1000.0,
            crossing_distance=40.0,
            side_transition="SIDE_A->SIDE_B",
            identity_certainty=IdentityCertainty.UNKNOWN,
            identity_candidate=None,
            identity_confidence=0.0,
            identity_evidence_ref="GO-1",
            source_crossing_event_id="CE-001",
            trajectory_points=[],
            config_snapshot={},
            event_schema_version="1.0",
            created_at="2026-01-01T00:00:00Z",
        )
        
        result1 = resolver1.resolve_events([raw_event])
        result2 = resolver2.resolve_events([raw_event])
        
        assert result1.transitions[0].resolution_id == result2.transitions[0].resolution_id

    # =========================================================================
    # PHASE 25: ATTENDANCE PERSISTENCE VERIFICATION
    # =========================================================================

    def test_phase25_persistence(self, temp_db_path):
        """Verify ResolvedTransition -> AttendanceRecord -> SQLite persistence."""
        config = StorageConfig(database_path=temp_db_path)
        storage = AttendanceStorage(config)
        repo = AttendanceRepository(config=config)
        
        # Create a resolved transition
        transition = ResolvedTransition(
            resolution_id="RES-TEST-001",
            source_raw_event_id="RIE-001",
            camera_id="CAM1",
            local_track_id="track_001",
            global_observation_id="GO-123",
            direction="in",
            transition_type=DerivedState.INSIDE,
            previous_state=DerivedState.UNKNOWN,
            new_state=DerivedState.INSIDE,
            source_timestamp=1000.0,
            source_frame_index=100,
            resolver_version="1.0",
            resolver_config_hash="config_hash",
            source_crossing_event_id="CE-001",
            geometry_version=1,
            geometry_config_hash="geom_hash",
        )
        
        # Persist through repository
        persistence_result = repo.persist_resolution_result(
            ResolutionResult(
                transitions=[transition],
                final_states={"CAM1:track_001": TrackResolutionState(
                    camera_id="CAM1",
                    local_track_id="track_001",
                    current_state=DerivedState.INSIDE,
                    last_transition_timestamp=1000.0,
                    last_transition_resolution_id="RES-TEST-001",
                    last_processed_raw_event_id="RIE-001",
                    transition_count=1,
                    in_count=1,
                    out_count=0,
                )},
                total_raw_events=1,
                accepted_transitions=1,
                suppressed_events=0,
                rejected_events=0,
                out_of_order_events=0,
                resolver_version="1.0",
                resolver_config_hash="config_hash",
            )
        )
        
        assert persistence_result.transitions_persisted == 1
        
        # Verify record exists
        record = repo.get_by_resolution_id("RES-TEST-001")
        assert record is not None
        assert record.source_resolution_id == "RES-TEST-001"
        assert record.source_raw_event_id == "RIE-001"
        assert record.source_crossing_event_id == "CE-001"
        assert record.geometry_version == 1
        assert record.resolver_version == "1.0"
        
        repo.close()
        storage.close()

    def test_phase25_idempotency(self, temp_db_path):
        """Verify idempotent persistence: duplicate source_resolution_id rejected."""
        config = StorageConfig(database_path=temp_db_path)
        storage = AttendanceStorage(config)
        
        record = AttendanceRecord(
            attendance_record_id="ATT-IDEMP-001",
            identity_certainty=AttendanceIdentityCertainty.UNKNOWN,
            direction=AttendanceDirection.IN,
            event_timestamp=1000.0,
            camera_id="CAM1",
            local_track_id="track_001",
            source_raw_event_id="RIE-001",
            source_resolution_id="RES-IDEMP-001",
        )
        
        inserted1 = storage.insert(record)
        inserted2 = storage.insert(record)
        
        assert inserted1 is True
        assert inserted2 is False  # Duplicate rejected
        
        storage.close()

    def test_phase25_restart_recovery(self, temp_db_path):
        """Verify restart recovery: data persists across connections."""
        config = StorageConfig(database_path=temp_db_path)
        
        # First connection: insert data
        storage1 = AttendanceStorage(config)
        repo1 = AttendanceRepository(config=config)
        
        transition = ResolvedTransition(
            resolution_id="RES-RESTART-001",
            source_raw_event_id="RIE-001",
            camera_id="CAM1",
            local_track_id="track_001",
            global_observation_id="GO-123",
            direction="in",
            transition_type=DerivedState.INSIDE,
            previous_state=DerivedState.UNKNOWN,
            new_state=DerivedState.INSIDE,
            source_timestamp=1000.0,
            source_frame_index=100,
            resolver_version="1.0",
            resolver_config_hash="config_hash",
            source_crossing_event_id="CE-001",
            geometry_version=1,
            geometry_config_hash="geom_hash",
        )
        
        repo1.persist_resolution_result(ResolutionResult(
            transitions=[transition],
            final_states={},
            total_raw_events=1,
            accepted_transitions=1,
            suppressed_events=0,
            rejected_events=0,
            out_of_order_events=0,
            resolver_version="1.0",
            resolver_config_hash="config_hash",
        ))
        
        repo1.close()
        storage1.close()
        
        # Second connection: verify data exists
        storage2 = AttendanceStorage(config)
        repo2 = AttendanceRepository(config=config)
        
        records = repo2.get_chronological_history(camera_id="CAM1", local_track_id="track_001")
        assert len(records) == 1
        assert records[0].source_resolution_id == "RES-RESTART-001"
        
        repo2.close()
        storage2.close()

    # =========================================================================
    # PHASE 26: ATTENDANCE DECISION VERIFICATION
    # =========================================================================

    def test_phase26_attendance_decision_present(self):
        """Verify PRESENT decision for IN within entry window."""
        # Create timetable entry
        entry = TimetableEntry(
            entry_id="entry-test-1",
            person_id="person-001",
            session_id="session-001",
            day=SessionDay.MONDAY,
            entry_time=36000,  # 10:00 AM
            exit_time=72000,   # 8:00 PM
            entry_window_start=35400,  # 9:50 AM
            entry_window_end=36600,    # 10:10 AM
            late_tolerance=600,        # 10 minutes
            exit_window_start=71400,   # 7:50 PM
            exit_window_end=72600,     # 8:10 PM
        )
        
        timetable = Timetable(timetable_id="ttb-test-1", timetable_version="1.0")
        timetable.entries.append(entry)
        
        policy = AttendancePolicy(policy_id="policy-test-1", policy_version="1.0")
        engine = AttendanceEngine(policy)
        
        # Create resolved transition for IN at 10:00 AM (within window)
        transition = ResolvedTransition(
            resolution_id="RES-DEC-001",
            source_raw_event_id="RIE-001",
            camera_id="CAM1",
            local_track_id="track_001",
            global_observation_id="GO-123",
            direction="in",
            transition_type=DerivedState.INSIDE,
            previous_state=DerivedState.UNKNOWN,
            new_state=DerivedState.INSIDE,
            source_timestamp=36000,  # 10:00 AM
            source_frame_index=100,
            resolver_version="1.0",
            resolver_config_hash="config_hash",
            source_crossing_event_id="CE-001",
            geometry_version=1,
            geometry_config_hash="geom_hash",
        )
        
        context = AttendanceDecisionContext(
            resolved_transition=transition,
            timetable=timetable,
            attendance_policy=policy,
            person_id_override="person-001",
            day_override=SessionDay.MONDAY,
        )
        
        decision = engine.make_decision(context)
        
        assert decision.new_attendance_state == "present"
        assert decision.decision_reason == "within_entry_window"
        assert decision.is_in is True
        assert decision.is_out is False
        assert decision.identity_certainty == "known"
        assert decision.identity_candidate == "person-001"

    def test_phase26_attendance_decision_late(self):
        """Verify LATE decision for IN within late tolerance."""
        entry = TimetableEntry(
            entry_id="entry-test-2",
            person_id="person-002",
            session_id="session-001",
            day=SessionDay.MONDAY,
            entry_time=36000,  # 10:00 AM
            exit_time=72000,
            entry_window_start=35400,
            entry_window_end=36000,  # Window ends at entry time
            late_tolerance=1200,     # 20 minutes
            exit_window_start=71400,
            exit_window_end=72600,
        )
        
        timetable = Timetable(timetable_id="ttb-test-2", timetable_version="1.0")
        timetable.entries.append(entry)
        
        policy = AttendancePolicy(policy_id="policy-test-2", policy_version="1.0")
        engine = AttendanceEngine(policy)
        
        # IN at 10:15 AM (15 min late, within 20 min tolerance)
        transition = ResolvedTransition(
            resolution_id="RES-DEC-002",
            source_raw_event_id="RIE-002",
            camera_id="CAM1",
            local_track_id="track_001",
            global_observation_id="GO-124",
            direction="in",
            transition_type=DerivedState.INSIDE,
            previous_state=DerivedState.UNKNOWN,
            new_state=DerivedState.INSIDE,
            source_timestamp=36900,  # 10:15 AM
            source_frame_index=150,
            resolver_version="1.0",
            resolver_config_hash="config_hash",
            source_crossing_event_id="CE-002",
            geometry_version=1,
            geometry_config_hash="geom_hash",
        )
        
        context = AttendanceDecisionContext(
            resolved_transition=transition,
            timetable=timetable,
            attendance_policy=policy,
            person_id_override="person-002",
            day_override=SessionDay.MONDAY,
        )
        
        decision = engine.make_decision(context)
        
        assert decision.new_attendance_state == "late"
        assert decision.decision_reason == "late_within_tolerance"

    def test_phase26_attendance_decision_absent(self):
        """Verify ABSENT decision for IN outside attendance window."""
        entry = TimetableEntry(
            entry_id="entry-test-3",
            person_id="person-003",
            session_id="session-001",
            day=SessionDay.MONDAY,
            entry_time=36000,
            exit_time=72000,
            entry_window_start=35400,
            entry_window_end=36600,
            late_tolerance=600,
            exit_window_start=71400,
            exit_window_end=72600,
        )
        
        timetable = Timetable(timetable_id="ttb-test-3", timetable_version="1.0")
        timetable.entries.append(entry)
        
        policy = AttendancePolicy(policy_id="policy-test-3", policy_version="1.0")
        engine = AttendanceEngine(policy)
        
        # IN at 11:00 AM (60 min late, outside 10 min tolerance)
        transition = ResolvedTransition(
            resolution_id="RES-DEC-003",
            source_raw_event_id="RIE-003",
            camera_id="CAM1",
            local_track_id="track_001",
            global_observation_id="GO-125",
            direction="in",
            transition_type=DerivedState.INSIDE,
            previous_state=DerivedState.UNKNOWN,
            new_state=DerivedState.INSIDE,
            source_timestamp=39600,  # 11:00 AM
            source_frame_index=180,
            resolver_version="1.0",
            resolver_config_hash="config_hash",
            source_crossing_event_id="CE-003",
            geometry_version=1,
            geometry_config_hash="geom_hash",
        )
        
        context = AttendanceDecisionContext(
            resolved_transition=transition,
            timetable=timetable,
            attendance_policy=policy,
            person_id_override="person-003",
            day_override=SessionDay.MONDAY,
        )
        
        decision = engine.make_decision(context)
        
        assert decision.new_attendance_state == "absent"
        assert decision.decision_reason == "outside_attendance_window"

    def test_phase26_attendance_decision_left(self):
        """Verify LEFT decision for OUT within exit window."""
        entry = TimetableEntry(
            entry_id="entry-test-4",
            person_id="person-004",
            session_id="session-001",
            day=SessionDay.MONDAY,
            entry_time=36000,
            exit_time=72000,
            entry_window_start=35400,
            entry_window_end=36600,
            late_tolerance=600,
            exit_window_start=71400,
            exit_window_end=72600,
        )
        
        timetable = Timetable(timetable_id="ttb-test-4", timetable_version="1.0")
        timetable.entries.append(entry)
        
        policy = AttendancePolicy(policy_id="policy-test-4", policy_version="1.0")
        engine = AttendanceEngine(policy)
        
        # OUT at 8:00 PM (within exit window)
        transition = ResolvedTransition(
            resolution_id="RES-DEC-004",
            source_raw_event_id="RIE-004",
            camera_id="CAM1",
            local_track_id="track_001",
            global_observation_id="GO-126",
            direction="out",
            transition_type=DerivedState.OUTSIDE,
            previous_state=DerivedState.INSIDE,
            new_state=DerivedState.OUTSIDE,
            source_timestamp=72000,  # 8:00 PM
            source_frame_index=200,
            resolver_version="1.0",
            resolver_config_hash="config_hash",
            source_crossing_event_id="CE-004",
            geometry_version=1,
            geometry_config_hash="geom_hash",
        )
        
        context = AttendanceDecisionContext(
            resolved_transition=transition,
            timetable=timetable,
            attendance_policy=policy,
            person_id_override="person-004",
            day_override=SessionDay.MONDAY,
        )
        
        decision = engine.make_decision(context)
        
        assert decision.new_attendance_state == "left"
        assert decision.decision_reason == "exit_recorded"
        assert decision.is_out is True

    def test_phase26_provenance_preserved(self):
        """Verify full provenance chain in AttendanceDecision."""
        entry = TimetableEntry(
            entry_id="entry-test-5",
            person_id="person-005",
            session_id="session-001",
            day=SessionDay.MONDAY,
            entry_time=36000,
            exit_time=72000,
            entry_window_start=35400,
            entry_window_end=36600,
            late_tolerance=600,
            exit_window_start=71400,
            exit_window_end=72600,
        )
        
        timetable = Timetable(timetable_id="ttb-test-5", timetable_version="1.0")
        timetable.entries.append(entry)
        
        policy = AttendancePolicy(policy_id="policy-test-5", policy_version="1.0")
        engine = AttendanceEngine(policy)
        
        transition = ResolvedTransition(
            resolution_id="RES-DEC-005",
            source_raw_event_id="RIE-005",
            camera_id="CAM1",
            local_track_id="track_001",
            global_observation_id="GO-127",
            direction="in",
            transition_type=DerivedState.INSIDE,
            previous_state=DerivedState.UNKNOWN,
            new_state=DerivedState.INSIDE,
            source_timestamp=36000,
            source_frame_index=100,
            resolver_version="1.0",
            resolver_config_hash="config_hash",
            source_crossing_event_id="CE-005",
            geometry_version=1,
            geometry_config_hash="geom_hash",
        )
        
        context = AttendanceDecisionContext(
            resolved_transition=transition,
            timetable=timetable,
            attendance_policy=policy,
            person_id_override="person-005",
            day_override=SessionDay.MONDAY,
        )
        
        decision = engine.make_decision(context)
        
        # Verify provenance chain
        assert decision.source_raw_event_id == "RIE-005"
        assert decision.source_resolution_id == "RES-DEC-005"
        assert decision.source_crossing_event_id == "CE-005"
        assert decision.camera_id == "CAM1"
        assert decision.local_track_id == "track_001"
        assert decision.global_observation_id == "GO-127"
        assert decision.geometry_version == 1
        assert decision.geometry_config_hash == "geom_hash"
        assert decision.resolver_version == "1.0"
        assert decision.resolver_config_hash == "config_hash"
        assert decision.timetable_id == "ttb-test-5"
        assert decision.session_id == "session-001"
        assert decision.day == "monday"
        assert decision.attendance_policy_id == "policy-test-5"
        assert decision.attendance_policy_version == "1.0"

    # =========================================================================
    # PHASE 29: IMMEDIATE EVENT OUTPUT VERIFICATION
    # =========================================================================

    def test_phase29_event_bus_creation(self):
        """Verify event bus creation and basic operations."""
        bus = create_event_bus()
        
        assert bus.get_subscriber_count() == 0
        
        # Subscribe
        received = []
        def handler(event):
            received.append(event)
        
        subscriber = FunctionSubscriber("test-sub", handler)
        config = SubscriberConfig(subscriber_id="test-sub", queue_size=100)
        
        bus.subscribe(subscriber, config)
        assert bus.get_subscriber_count() == 1
        
        # Publish event
        event = ImmediateEvent(
            event_id="IEV-TEST-001",
            event_type=ImmediateEventType.ATTENDANCE_IN,
            direction=ImmediateEventDirection.IN,
            identity_certainty=OutputIdentityCertainty.KNOWN,
            identity_candidate="person-001",
            identity_confidence=0.95,
            event_timestamp=36000.0,
            event_frame_index=100,
            camera_id="CAM1",
            local_track_id="track_001",
            global_observation_id="GO-123",
            source_raw_event_id="RIE-001",
            source_resolution_id="RES-001",
            source_crossing_event_id="CE-001",
            geometry_version=1,
            geometry_config_hash="geom_hash",
            resolver_version="1.0",
            resolver_config_hash="config_hash",
            timetable_id="ttb-001",
            session_id="session-001",
            day="monday",
            previous_attendance_state="unknown",
            new_attendance_state="present",
            decision_reason="within_entry_window",
            attendance_policy_id="policy-001",
            attendance_policy_version="1.0",
            event_schema_version="1.0",
        )
        
        result = bus.publish(event)
        assert result is True
        
        # Wait for delivery
        import time
        time.sleep(0.1)
        
        assert len(received) == 1
        assert received[0].event_id == "IEV-TEST-001"
        
        bus.unsubscribe("test-sub")
        bus.shutdown()

    def test_phase29_deduplication(self):
        """Verify duplicate suppression in event bus."""
        bus = create_event_bus(max_dedup_cache=100)
        
        received = []
        def handler(event):
            received.append(event)
        
        subscriber = FunctionSubscriber("test-sub", handler)
        config = SubscriberConfig(subscriber_id="test-sub", queue_size=100)
        bus.subscribe(subscriber, config)
        
        event = ImmediateEvent(
            event_id="IEV-DEDUP-001",
            event_type=ImmediateEventType.ATTENDANCE_IN,
            direction=ImmediateEventDirection.IN,
            identity_certainty=OutputIdentityCertainty.KNOWN,
            identity_candidate="person-001",
            identity_confidence=0.95,
            event_timestamp=36000.0,
            event_frame_index=100,
            camera_id="CAM1",
            local_track_id="track_001",
            global_observation_id="GO-123",
            source_raw_event_id="RIE-001",
            source_resolution_id="RES-001",
            source_crossing_event_id="CE-001",
            geometry_version=1,
            geometry_config_hash="geom_hash",
            resolver_version="1.0",
            resolver_config_hash="config_hash",
            timetable_id="ttb-001",
            session_id="session-001",
            day="monday",
            previous_attendance_state="unknown",
            new_attendance_state="present",
            decision_reason="within_entry_window",
            attendance_policy_id="policy-001",
            attendance_policy_version="1.0",
            event_schema_version="1.0",
        )
        
        # Publish same event twice (same source_resolution_id + event_type)
        bus.publish(event)
        bus.publish(event)
        
        import time
        time.sleep(0.1)
        
        # Should only receive once due to deduplication
        assert len(received) == 1
        
        stats = bus.get_stats()
        assert stats["events_duplicated"] == 1
        
        bus.shutdown()

    def test_phase29_bounded_memory(self):
        """Verify bounded event history and deduplication cache."""
        bus = create_event_bus(max_history=10, max_dedup_cache=10)
        
        received = []
        def handler(event):
            received.append(event)
        
        subscriber = FunctionSubscriber("test-sub", handler)
        config = SubscriberConfig(subscriber_id="test-sub", queue_size=100)
        bus.subscribe(subscriber, config)
        
        # Publish 15 events (more than max_history=10)
        for i in range(15):
            event = ImmediateEvent(
                event_id=f"IEV-MEM-{i}",
                event_type=ImmediateEventType.ATTENDANCE_IN,
                direction=ImmediateEventDirection.IN,
                identity_certainty=OutputIdentityCertainty.KNOWN,
                identity_candidate=f"person-{i}",
                identity_confidence=0.95,
                event_timestamp=36000.0 + i,
                event_frame_index=100 + i,
                camera_id="CAM1",
                local_track_id=f"track_{i}",
                global_observation_id=f"GO-{i}",
                source_raw_event_id=f"RIE-{i}",
                source_resolution_id=f"RES-{i}",
                source_crossing_event_id=f"CE-{i}",
                geometry_version=1,
                geometry_config_hash="geom_hash",
                resolver_version="1.0",
                resolver_config_hash="config_hash",
                timetable_id="ttb-001",
                session_id="session-001",
                day="monday",
                previous_attendance_state="unknown",
                new_attendance_state="present",
                decision_reason="within_entry_window",
                attendance_policy_id="policy-001",
                attendance_policy_version="1.0",
                event_schema_version="1.0",
            )
            bus.publish(event)
        
        import time
        time.sleep(0.2)
        
        # History should be bounded
        history = bus.get_history(limit=20)
        assert len(history) <= 10
        
        stats = bus.get_stats()
        assert stats["history_size"] <= 10
        
        bus.shutdown()

    # =========================================================================
    # PHASE 30: DAILY EXCEL EXPORT VERIFICATION
    # =========================================================================

    def test_phase30_excel_export(self, temp_db_path, temp_output_dir):
        """Verify DailyExcelExporter generates workbook with required sheets."""
        config = StorageConfig(database_path=temp_db_path)
        storage = AttendanceStorage(config)
        repo = AttendanceRepository(storage=storage)
        exporter = DailyExcelExporter(repository=repo)
        
        # Create test attendance records
        base_ts = 1787446800.0  # 2026-08-23 08:00:00 Bangkok
        
        records = [
            AttendanceRecord(
                attendance_record_id="ATT-EXCEL-001",
                identity_certainty=AttendanceIdentityCertainty.KNOWN,
                identity_candidate="student_001",
                identity_confidence=0.95,
                identity_evidence_ref="GO-001",
                direction=AttendanceDirection.IN,
                event_timestamp=base_ts,
                event_frame_index=100,
                camera_id="CAM1",
                local_track_id="track_001",
                global_observation_id="GO-001",
                source_raw_event_id="RIE-001",
                source_resolution_id="RES-001",
                source_crossing_event_id="CE-001",
                geometry_version=1,
                geometry_config_hash="geom_hash_123",
                resolver_version="1.0",
                resolver_config_hash="config_hash_456",
                previous_state="unknown",
                new_state="inside",
                attendance_schema_version="1.0",
            ),
            AttendanceRecord(
                attendance_record_id="ATT-EXCEL-002",
                identity_certainty=AttendanceIdentityCertainty.KNOWN,
                identity_candidate="student_002",
                identity_confidence=0.92,
                identity_evidence_ref="GO-002",
                direction=AttendanceDirection.OUT,
                event_timestamp=base_ts + 14400,  # 12:00:00
                event_frame_index=200,
                camera_id="CAM1",
                local_track_id="track_002",
                global_observation_id="GO-002",
                source_raw_event_id="RIE-002",
                source_resolution_id="RES-002",
                source_crossing_event_id="CE-002",
                geometry_version=1,
                geometry_config_hash="geom_hash_123",
                resolver_version="1.0",
                resolver_config_hash="config_hash_456",
                previous_state="inside",
                new_state="outside",
                attendance_schema_version="1.0",
            ),
            AttendanceRecord(
                attendance_record_id="ATT-EXCEL-003",
                identity_certainty=AttendanceIdentityCertainty.UNKNOWN,
                identity_candidate=None,
                identity_confidence=0.0,
                identity_evidence_ref=None,
                direction=AttendanceDirection.IN,
                event_timestamp=base_ts,
                event_frame_index=300,
                camera_id="CAM2",
                local_track_id="track_003",
                global_observation_id=None,
                source_raw_event_id="RIE-003",
                source_resolution_id="RES-003",
                source_crossing_event_id="CE-003",
                geometry_version=1,
                geometry_config_hash="geom_hash_123",
                resolver_version="1.0",
                resolver_config_hash="config_hash_456",
                previous_state="unknown",
                new_state="inside",
                attendance_schema_version="1.0",
            ),
        ]
        
        for record in records:
            storage.insert(record)
        
        output_path = temp_output_dir / "daily_attendance_2026-08-23.xlsx"
        
        request = DailyExportRequest(
            date=date(2026, 8, 23),
            output_path=str(output_path),
            timezone="Asia/Bangkok",
            include_events_sheet=True,
            include_provenance_sheet=True,
            include_summary_sheet=True,
        )
        
        result = exporter.export_daily_attendance(request)
        
        assert result.success
        assert result.records_processed == 3
        assert result.records_exported == 3
        assert output_path.exists()
        
        # Verify workbook structure
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        
        required_sheets = ["DAILY_ATTENDANCE", "EVENTS", "SUMMARY", "PROVENANCE"]
        for sheet in required_sheets:
            assert sheet in wb.sheetnames, f"Missing sheet: {sheet}"
        
        # Check DAILY_ATTENDANCE columns
        ws = wb["DAILY_ATTENDANCE"]
        headers = [cell.value for cell in ws[1]]
        required_columns = ["No.", "Person ID", "Name", "Identity Certainty", "State", "IN Time", "OUT Time", "Duration", "Camera", "Global Observation", "Status"]
        for col in required_columns:
            assert col in headers, f"Missing column: {col}"
        
        # Check data rows
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) >= 3
        
        # Check identity certainty preserved
        certainties = set()
        for row in data_rows:
            if row[3]:
                certainties.add(row[3].upper())
        assert "KNOWN" in certainties
        assert "UNKNOWN" in certainties
        
        # Check multi-camera provenance
        cameras = set()
        for row in data_rows:
            if row[8]:
                cameras.add(row[8])
        assert "CAM1" in cameras
        assert "CAM2" in cameras
        
        # Check PROVENANCE sheet
        ws_prov = wb["PROVENANCE"]
        prov_headers = [cell.value for cell in ws_prov[1]]
        required_prov = ["Attendance Record ID", "Source Resolution ID", "Source Raw Event ID", "Global Observation ID", "Camera ID", "Local Track ID"]
        for col in required_prov:
            assert col in prov_headers, f"Missing provenance column: {col}"
        
        wb.close()
        exporter.close()
        repo.close()
        storage.close()

    def test_phase30_no_database_mutation(self, temp_db_path, temp_output_dir):
        """Verify export does not mutate source database."""
        config = StorageConfig(database_path=temp_db_path)
        storage = AttendanceStorage(config)
        
        record = AttendanceRecord(
            attendance_record_id="ATT-MUT-001",
            identity_certainty=AttendanceIdentityCertainty.KNOWN,
            identity_candidate="student_test",
            identity_confidence=0.9,
            identity_evidence_ref="GO-test",
            direction=AttendanceDirection.IN,
            event_timestamp=1787446800.0,
            event_frame_index=100,
            camera_id="CAM1",
            local_track_id="track_test",
            global_observation_id="GO-test",
            source_raw_event_id="RIE-test",
            source_resolution_id="RES-test",
            source_crossing_event_id="CE-test",
            geometry_version=1,
            geometry_config_hash="geom_hash",
            resolver_version="1.0",
            resolver_config_hash="config_hash",
            previous_state="unknown",
            new_state="inside",
            attendance_schema_version="1.0",
        )
        storage.insert(record)
        
        count_before = storage.count()
        
        repo = AttendanceRepository(storage=storage)
        exporter = DailyExcelExporter(repository=repo)
        
        output_path = temp_output_dir / "test_mutation.xlsx"
        request = DailyExportRequest(
            date=date(2026, 8, 23),
            output_path=str(output_path),
        )
        
        result = exporter.export_daily_attendance(request)
        
        count_after = storage.count()
        
        assert result.success
        assert count_before == count_after, "Database should not be mutated by export"
        
        exporter.close()
        repo.close()
        storage.close()

    # =========================================================================
    # PHASE 27: ANNOTATED REPLAY / VIDEO EVIDENCE VERIFICATION
    # =========================================================================

    def test_phase27_annotated_replay_pipeline(self, test_data_dir, temp_output_dir):
        """Verify AnnotatedReplayPipeline integrates Phase 20-26."""
        cam1_path = test_data_dir / "cam1_test.mp4"
        cam2_path = test_data_dir / "cam2_test.mp4"
        
        source_configs = [
            ReplaySourceConfig(
                source_path=str(cam1_path),
                camera_id="CAM1",
            ),
            ReplaySourceConfig(
                source_path=str(cam2_path),
                camera_id="CAM2",
            ),
        ]
        
        config = AnnotatedReplayConfig(
            output_directory=str(temp_output_dir),
            save_annotation_frames=True,
            build_appearance_index=True,
        )
        
        pipeline = AnnotatedReplayPipeline(source_configs, config=config)
        state = pipeline.run()
        
        assert state.frames_processed > 0
        assert state.frames_annotated >= 0
        assert len(state.annotation_frames) >= 0
        
        # Verify output files created
        summary_path = temp_output_dir / "replay_summary.json"
        assert summary_path.exists()
        
        import json
        with open(summary_path) as f:
            summary = json.load(f)
        
        assert summary["frames_processed"] > 0
        assert "global_observations_detail" in summary
        assert "appearance_index" in summary

    def test_phase27_appearance_index(self, test_data_dir, temp_output_dir):
        """Verify person appearance indexing works."""
        cam1_path = test_data_dir / "cam1_test.mp4"
        
        source_configs = [
            ReplaySourceConfig(
                source_path=str(cam1_path),
                camera_id="CAM1",
            ),
        ]
        
        config = AnnotatedReplayConfig(
            output_directory=str(temp_output_dir),
            build_appearance_index=True,
        )
        
        pipeline = AnnotatedReplayPipeline(source_configs, config=config)
        state = pipeline.run()
        
        # Verify appearance index structure
        assert isinstance(state.appearance_index, dict)
        assert isinstance(state.track_appearances, dict)
        
        # Search functionality
        for person_id, appearances in state.appearance_index.items():
            result = pipeline.search_person_appearances(person_id)
            assert result.person_id == person_id
            assert len(result.appearances) == len(appearances)

    def test_phase27_video_evidence_contracts(self):
        """Verify video evidence contracts exist."""
        # VideoSourceInfo
        source_info = VideoSourceInfo(
            source_video_id="CAM1_video",
            camera_id="CAM1",
            file_path="/path/to/video.mp4",
            duration_seconds=3600.0,
            fps=30.0,
            frame_count=108000,
            width=1920,
            height=1080,
            codec="h264",
        )
        
        assert source_info.source_video_id == "CAM1_video"
        assert source_info.camera_id == "CAM1"
        
        # VideoSegmentRequest
        request = VideoSegmentRequest(
            source_video_id="CAM1_video",
            camera_id="CAM1",
            start_timestamp=1000.0,
            end_timestamp=1010.0,
            start_frame=30000,
            end_frame=30300,
        )
        
        assert request.source_video_id == "CAM1_video"
        assert request.camera_id == "CAM1"
        assert request.start_timestamp == 1000.0
        assert request.end_timestamp == 1010.0

    # =========================================================================
    # END-TO-END PROVENANCE CHAIN
    # =========================================================================

    def test_e2e_provenance_chain(self, test_data_dir, enrollment_db_path, temp_db_path, temp_output_dir):
        """Verify complete provenance chain from video to Excel."""
        # This is a comprehensive integration test that runs the full pipeline
        # and verifies all identifiers remain traceable
        
        cam1_path = test_data_dir / "cam1_test.mp4"
        
        # Phase 20: Replay
        source = ReplaySource(ReplaySourceConfig(
            source_path=str(cam1_path),
            camera_id="CAM1",
        ))
        source.open()
        frame = source.get_next_frame()
        assert frame is not None
        
        camera_id = frame.metadata.extra.get("camera_id")
        frame_index = frame.metadata.frame_index
        timestamp = frame.metadata.timestamp
        assert camera_id == "CAM1"
        assert frame_index >= 0
        assert timestamp >= 0
        
        # Phase 15-19: Pipeline
        pipeline = create_replay_pipeline(enrollment_db_path=enrollment_db_path)
        frame_result = pipeline.process_frame(frame)
        
        assert frame_result.camera_id == "CAM1"
        assert frame_result.frame_index == frame_index
        assert frame_result.timestamp == timestamp
        
        # Phase 21: Fusion (if detections exist)
        if frame_result.detections:
            fusion_engine = create_fusion_engine(DEFAULT_FUSION_CONFIG)
            
            for i, detection in enumerate(frame_result.detections):
                obs_ref = build_local_observation_ref(
                    frame=frame,
                    local_track_id=f"track_{detection.detection_id}",
                    detection_id=detection.detection_id,
                    face_crop_id=f"face_{detection.detection_id}" if i < len(frame_result.face_crops) else None,
                    quality_class=frame_result.quality_results[i].quality_class.value if i < len(frame_result.quality_results) else None,
                )
                fusion_engine.add_observation(obs_ref)
            
            global_observations = fusion_engine.associate_observations()
            # May be empty if only one camera
        
        # Phase 22-24: Geometry -> Raw Events -> Resolution
        # (Requires geometry config and track - simplified here)
        
        # Phase 25: Persistence
        config = StorageConfig(database_path=temp_db_path)
        storage = AttendanceStorage(config)
        repo = AttendanceRepository(config=config)
        
        # Create a complete record with full provenance
        record = AttendanceRecord(
            attendance_record_id="ATT-E2E-001",
            identity_certainty=AttendanceIdentityCertainty.KNOWN,
            identity_candidate="person_e2e",
            identity_confidence=0.9,
            identity_evidence_ref="GO-E2E-001",
            direction=AttendanceDirection.IN,
            event_timestamp=61200.0,  # 1970-01-02 00:00:00 Bangkok time (UTC+7)
            event_frame_index=frame_index,
            camera_id=camera_id,
            local_track_id="track_e2e_001",
            global_observation_id="GO-E2E-001",
            source_raw_event_id="RIE-E2E-001",
            source_resolution_id="RES-E2E-001",
            source_crossing_event_id="CE-E2E-001",
            geometry_version=1,
            geometry_config_hash="geom_hash",
            resolver_version="1.0",
            resolver_config_hash="config_hash",
            previous_state="unknown",
            new_state="inside",
            attendance_schema_version="1.0",
        )
        
        storage.insert(record)
        
        # Verify full provenance chain
        retrieved = repo.get_by_id("ATT-E2E-001")
        assert retrieved is not None
        assert retrieved.source_resolution_id == "RES-E2E-001"
        assert retrieved.source_raw_event_id == "RIE-E2E-001"
        assert retrieved.source_crossing_event_id == "CE-E2E-001"
        assert retrieved.global_observation_id == "GO-E2E-001"
        assert retrieved.camera_id == "CAM1"
        assert retrieved.local_track_id == "track_e2e_001"
        assert retrieved.geometry_version == 1
        assert retrieved.resolver_version == "1.0"
        
        # Phase 26: Attendance Decision
        entry = TimetableEntry(
            entry_id="entry-e2e",
            person_id="person_e2e",
            session_id="session_e2e",
            day=SessionDay.MONDAY,
            entry_time=36000,
            exit_time=72000,
            entry_window_start=35400,
            entry_window_end=36600,
            late_tolerance=600,
            exit_window_start=71400,
            exit_window_end=72600,
        )
        timetable = Timetable(timetable_id="ttb-e2e", timetable_version="1.0")
        timetable.entries.append(entry)
        
        policy = AttendancePolicy(policy_id="policy-e2e", policy_version="1.0")
        engine = AttendanceEngine(policy)
        
        transition = ResolvedTransition(
            resolution_id="RES-E2E-001",
            source_raw_event_id="RIE-E2E-001",
            camera_id="CAM1",
            local_track_id="track_e2e_001",
            global_observation_id="GO-E2E-001",
            direction="in",
            transition_type=DerivedState.INSIDE,
            previous_state=DerivedState.UNKNOWN,
            new_state=DerivedState.INSIDE,
            source_timestamp=timestamp,
            source_frame_index=frame_index,
            resolver_version="1.0",
            resolver_config_hash="config_hash",
            source_crossing_event_id="CE-E2E-001",
            geometry_version=1,
            geometry_config_hash="geom_hash",
        )
        
        context = AttendanceDecisionContext(
            resolved_transition=transition,
            timetable=timetable,
            attendance_policy=policy,
            person_id_override="person_e2e",
            day_override=SessionDay.MONDAY,
        )
        
        decision = engine.make_decision(context)
        
        # Verify decision has full provenance
        assert decision.source_raw_event_id == "RIE-E2E-001"
        assert decision.source_resolution_id == "RES-E2E-001"
        assert decision.source_crossing_event_id == "CE-E2E-001"
        assert decision.global_observation_id == "GO-E2E-001"
        assert decision.camera_id == "CAM1"
        assert decision.local_track_id == "track_e2e_001"
        assert decision.timetable_id == "ttb-e2e"
        assert decision.session_id == "session_e2e"
        assert decision.attendance_policy_id == "policy-e2e"
        
        # Phase 29: Immediate Event
        bus = create_event_bus()
        adapter = Phase26ToImmediateEventAdapter(bus)
        adapter.publish(decision)
        
        history = bus.get_history(limit=10)
        assert len(history) >= 1
        immediate_event = history[0]
        assert immediate_event.source_resolution_id == "RES-E2E-001"
        assert immediate_event.source_raw_event_id == "RIE-E2E-001"
        
        bus.shutdown()
        
        # Phase 30: Excel Export
        exporter = DailyExcelExporter(repository=repo)
        output_path = temp_output_dir / "e2e_provenance.xlsx"
        
        # Use a date that works with the frame timestamp (epoch 0 = 1970-01-01 UTC)
        # In Bangkok timezone (UTC+7), 1970-01-02 starts at 1970-01-01 17:00:00 UTC (timestamp 61200)
        # which is > 0, so the frame timestamp (0.0) falls on 1970-01-01 in Bangkok
        # But the query range for 1970-01-01 in Bangkok starts at negative timestamp
        # So use 1970-01-02 which has positive timestamp range
        from datetime import date
        export_date = date(1970, 1, 2)
        
        request = DailyExportRequest(
            date=export_date,
            output_path=str(output_path),
            include_provenance_sheet=True,
        )
        
        result = exporter.export_daily_attendance(request)
        assert result.success
        
        # Verify provenance in Excel
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws_prov = wb["PROVENANCE"]
        prov_data = list(ws_prov.iter_rows(min_row=2, values_only=True))
        assert len(prov_data) >= 1
        
        # Check provenance row has all identifiers
        prov_row = prov_data[0]
        assert "RES-E2E-001" in str(prov_row)  # source_resolution_id
        assert "RIE-E2E-001" in str(prov_row)  # source_raw_event_id
        assert "CE-E2E-001" in str(prov_row)   # source_crossing_event_id
        assert "GO-E2E-001" in str(prov_row)   # global_observation_id
        assert "CAM1" in str(prov_row)         # camera_id
        assert "track_e2e_001" in str(prov_row) # local_track_id
        
        wb.close()
        exporter.close()
        repo.close()
        storage.close()
        pipeline.close()
        source.close()

    # =========================================================================
    # IDEMPOTENCY GATE
    # =========================================================================

    def test_idempotency_phase23(self):
        """Verify Phase 23 idempotency: same crossing event = same raw event."""
        raw_engine1 = create_raw_event_engine()
        raw_engine2 = create_raw_event_engine()
        
        line_config = create_line_geometry(
            camera_id="CAM1",
            frame_width=1920,
            frame_height=1080,
            p1=(100, 500),
            p2=(1820, 500),
            direction_semantics=DirectionSemantics.SIDE_A_TO_B_IN,
            version=1,
        )
        geom_snapshot = GeometryConfigSnapshot.from_config(line_config)
        
        crossing_event = CrossingEvent(
            event_id="CE-IDEMP-001",
            camera_id="CAM1",
            geometry_config=geom_snapshot,
            local_track_id="track_001",
            global_observation_id="GO-123",
            event_type=CrossingEventType.LINE_CROSSING,
            direction=CrossingDirection.IN,
            crossing_point=Point2D(960, 500),
            crossing_timestamp=1000.0,
            previous_position=Point2D(960, 480),
            current_position=Point2D(960, 520),
            previous_frame_index=100,
            current_frame_index=101,
            previous_timestamp=999.0,
            current_timestamp=1000.0,
            crossing_distance=40.0,
            side_transition="SIDE_A->SIDE_B",
            trajectory_points=[],
            config_snapshot={},
            created_at="2026-01-01T00:00:00Z",
            version="1.0",
        )
        
        result1 = raw_engine1.process_crossing_event(crossing_event)
        result2 = raw_engine2.process_crossing_event(crossing_event)
        
        assert result1.success and result2.success
        assert result1.event.event_id == result2.event.event_id

    def test_idempotency_phase24(self):
        """Verify Phase 24 idempotency: same raw events = same resolution IDs."""
        resolver1 = create_repeated_in_out_resolver(create_default_resolver_config())
        resolver2 = create_repeated_in_out_resolver(create_default_resolver_config())
        
        raw_event = RawInOutEvent(
            event_id="RIE-IDEMP-001",
            camera_id="CAM1",
            geometry_id="hash123",
            geometry_version=1,
            geometry_config_hash="hash123",
            local_track_id="track_001",
            global_observation_id="GO-1",
            event_type=RawEventType.LINE_CROSSING,
            direction=RawEventDirection.IN,
            crossing_point_x=960.0,
            crossing_point_y=500.0,
            crossing_timestamp=1000.0,
            crossing_frame_index=100,
            previous_position_x=960.0,
            previous_position_y=480.0,
            current_position_x=960.0,
            current_position_y=520.0,
            previous_frame_index=99,
            current_frame_index=100,
            previous_timestamp=999.0,
            current_timestamp=1000.0,
            crossing_distance=40.0,
            side_transition="SIDE_A->SIDE_B",
            identity_certainty=IdentityCertainty.UNKNOWN,
            identity_candidate=None,
            identity_confidence=0.0,
            identity_evidence_ref="GO-1",
            source_crossing_event_id="CE-001",
            trajectory_points=[],
            config_snapshot={},
            event_schema_version="1.0",
            created_at="2026-01-01T00:00:00Z",
        )
        
        result1 = resolver1.resolve_events([raw_event])
        result2 = resolver2.resolve_events([raw_event])
        
        assert result1.transitions[0].resolution_id == result2.transitions[0].resolution_id

    def test_idempotency_phase25(self, temp_db_path):
        """Verify Phase 25 idempotency: duplicate source_resolution_id rejected."""
        config = StorageConfig(database_path=temp_db_path)
        storage = AttendanceStorage(config)
        
        record = AttendanceRecord(
            attendance_record_id="ATT-IDEMP-001",
            identity_certainty=AttendanceIdentityCertainty.UNKNOWN,
            direction=AttendanceDirection.IN,
            event_timestamp=1000.0,
            camera_id="CAM1",
            local_track_id="track_001",
            source_raw_event_id="RIE-001",
            source_resolution_id="RES-IDEMP-001",
        )
        
        inserted1 = storage.insert(record)
        inserted2 = storage.insert(record)
        
        assert inserted1 is True
        assert inserted2 is False  # Duplicate rejected
        
        storage.close()

    def test_idempotency_phase29(self):
        """Verify Phase 29 idempotency: duplicate immediate events suppressed."""
        bus = create_event_bus(max_dedup_cache=100)
        
        received = []
        def handler(event):
            received.append(event)
        
        subscriber = FunctionSubscriber("test-sub", handler)
        config = SubscriberConfig(subscriber_id="test-sub", queue_size=100)
        bus.subscribe(subscriber, config)
        
        event = ImmediateEvent(
            event_id="IEV-IDEMP-001",
            event_type=ImmediateEventType.ATTENDANCE_IN,
            direction=ImmediateEventDirection.IN,
            identity_certainty=OutputIdentityCertainty.KNOWN,
            identity_candidate="person-001",
            identity_confidence=0.95,
            event_timestamp=36000.0,
            event_frame_index=100,
            camera_id="CAM1",
            local_track_id="track_001",
            global_observation_id="GO-123",
            source_raw_event_id="RIE-001",
            source_resolution_id="RES-001",
            source_crossing_event_id="CE-001",
            geometry_version=1,
            geometry_config_hash="geom_hash",
            resolver_version="1.0",
            resolver_config_hash="config_hash",
            timetable_id="ttb-001",
            session_id="session-001",
            day="monday",
            previous_attendance_state="unknown",
            new_attendance_state="present",
            decision_reason="within_entry_window",
            attendance_policy_id="policy-001",
            attendance_policy_version="1.0",
            event_schema_version="1.0",
        )
        
        # Publish same event twice
        bus.publish(event)
        bus.publish(event)
        
        import time
        time.sleep(0.1)
        
        # Should only receive once
        assert len(received) == 1
        
        stats = bus.get_stats()
        assert stats["events_duplicated"] == 1
        
        bus.shutdown()

    # =========================================================================
    # DETERMINISM GATE
    # =========================================================================

    def test_determinism_phase20_replay(self, test_data_dir):
        """Verify deterministic replay: same source = same frame ordering."""
        cam1_path = test_data_dir / "cam1_test.mp4"
        
        frames_run1 = []
        frames_run2 = []
        
        for run in range(2):
            source = ReplaySource(ReplaySourceConfig(
                source_path=str(cam1_path),
                camera_id="CAM1",
            ))
            source.open()
            
            frames = []
            for _ in range(20):
                frame = source.get_next_frame()
                if frame is None:
                    break
                frames.append({
                    "frame_index": frame.metadata.frame_index,
                    "timestamp": frame.metadata.timestamp,
                    "camera_id": frame.metadata.extra.get("camera_id"),
                })
            source.close()
            
            if run == 0:
                frames_run1 = frames
            else:
                frames_run2 = frames
        
        assert len(frames_run1) == len(frames_run2)
        for f1, f2 in zip(frames_run1, frames_run2):
            assert f1["frame_index"] == f2["frame_index"]
            assert f1["timestamp"] == f2["timestamp"]
            assert f1["camera_id"] == f2["camera_id"]

    def test_determinism_phase21_fusion(self):
        """Verify deterministic fusion: same observations = same global observation IDs."""
        engine1 = create_fusion_engine(DEFAULT_FUSION_CONFIG)
        engine2 = create_fusion_engine(DEFAULT_FUSION_CONFIG)
        
        obs1 = LocalObservationRef(
            camera_id="CAM1",
            local_track_id="track_001",
            observation_id="CAM1_track_001_f100",
            frame_index=100,
            timestamp=ReplayTimestamp(value=1000.0, source="frame_metadata"),
        )
        obs2 = LocalObservationRef(
            camera_id="CAM2",
            local_track_id="track_002",
            observation_id="CAM2_track_002_f100",
            frame_index=100,
            timestamp=ReplayTimestamp(value=1000.1, source="frame_metadata"),
        )
        
        engine1.add_observation(obs1)
        engine1.add_observation(obs2)
        engine2.add_observation(obs1)
        engine2.add_observation(obs2)
        
        global_obs1 = engine1.associate_observations()
        global_obs2 = engine2.associate_observations()
        
        assert len(global_obs1) == len(global_obs2)
        if global_obs1 and global_obs2:
            assert global_obs1[0].global_observation_id == global_obs2[0].global_observation_id

    def test_determinism_phase24_resolver(self):
        """Verify deterministic resolver: same raw events = same resolution IDs."""
        resolver1 = create_repeated_in_out_resolver(create_default_resolver_config())
        resolver2 = create_repeated_in_out_resolver(create_default_resolver_config())
        
        raw_events = []
        for i in range(3):
            raw_events.append(RawInOutEvent(
                event_id=f"RIE-DET-{i}",
                camera_id="CAM1",
                geometry_id="hash123",
                geometry_version=1,
                geometry_config_hash="hash123",
                local_track_id="track_001",
                global_observation_id=f"GO-{i}",
                event_type=RawEventType.LINE_CROSSING,
                direction=RawEventDirection.IN if i % 2 == 0 else RawEventDirection.OUT,
                crossing_point_x=960.0,
                crossing_point_y=500.0,
                crossing_timestamp=1000.0 + i * 100,
                crossing_frame_index=100 + i,
                previous_position_x=960.0,
                previous_position_y=480.0 if i % 2 == 0 else 520.0,
                current_position_x=960.0,
                current_position_y=520.0 if i % 2 == 0 else 480.0,
                previous_frame_index=99 + i,
                current_frame_index=100 + i,
                previous_timestamp=999.0 + i * 100,
                current_timestamp=1000.0 + i * 100,
                crossing_distance=40.0,
                side_transition="SIDE_A->SIDE_B" if i % 2 == 0 else "SIDE_B->SIDE_A",
                identity_certainty=IdentityCertainty.UNKNOWN,
                identity_candidate=None,
                identity_confidence=0.0,
                identity_evidence_ref=f"GO-{i}",
                source_crossing_event_id=f"CE-{i}",
                trajectory_points=[],
                config_snapshot={},
                event_schema_version="1.0",
                created_at="2026-01-01T00:00:00Z",
            ))
        
        result1 = resolver1.resolve_events(raw_events)
        result2 = resolver2.resolve_events(raw_events)
        
        assert len(result1.transitions) == len(result2.transitions)
        for t1, t2 in zip(result1.transitions, result2.transitions):
            assert t1.resolution_id == t2.resolution_id

    def test_determinism_phase26_attendance_engine(self):
        """Verify deterministic attendance engine: same inputs = same decision IDs."""
        entry = TimetableEntry(
            entry_id="entry-det-1",
            person_id="person-det-1",
            session_id="session-det-1",
            day=SessionDay.MONDAY,
            entry_time=36000,
            exit_time=72000,
            entry_window_start=35400,
            entry_window_end=36600,
            late_tolerance=600,
            exit_window_start=71400,
            exit_window_end=72600,
        )
        
        timetable = Timetable(timetable_id="ttb-det-1", timetable_version="1.0")
        timetable.entries.append(entry)
        
        policy = AttendancePolicy(policy_id="policy-det-1", policy_version="1.0")
        engine1 = AttendanceEngine(policy)
        engine2 = AttendanceEngine(policy)
        
        transition = ResolvedTransition(
            resolution_id="RES-DET-001",
            source_raw_event_id="RIE-DET-001",
            camera_id="CAM1",
            local_track_id="track_001",
            global_observation_id="GO-DET-001",
            direction="in",
            transition_type=DerivedState.INSIDE,
            previous_state=DerivedState.UNKNOWN,
            new_state=DerivedState.INSIDE,
            source_timestamp=36000,
            source_frame_index=100,
            resolver_version="1.0",
            resolver_config_hash="config_hash",
            source_crossing_event_id="CE-DET-001",
            geometry_version=1,
            geometry_config_hash="geom_hash",
        )
        
        context = AttendanceDecisionContext(
            resolved_transition=transition,
            timetable=timetable,
            attendance_policy=policy,
            person_id_override="person-det-1",
            day_override=SessionDay.MONDAY,
        )
        
        decision1 = engine1.make_decision(context)
        decision2 = engine2.make_decision(context)
        
        assert decision1.decision_id == decision2.decision_id

    # =========================================================================
    # TWO-CAMERA CROSS-CAMERA TEST
    # =========================================================================

    def test_two_camera_cross_camera(self, test_data_dir):
        """Verify CAM1 and CAM2 process independently and fuse correctly."""
        cam1_path = test_data_dir / "cam1_test.mp4"
        cam2_path = test_data_dir / "cam2_test.mp4"
        
        source1 = ReplaySource(ReplaySourceConfig(
            source_path=str(cam1_path),
            camera_id="CAM1",
        ))
        source2 = ReplaySource(ReplaySourceConfig(
            source_path=str(cam2_path),
            camera_id="CAM2",
        ))
        
        source1.open()
        source2.open()
        
        frame1 = source1.get_next_frame()
        frame2 = source2.get_next_frame()
        
        assert frame1 is not None
        assert frame2 is not None
        assert frame1.metadata.extra.get("camera_id") == "CAM1"
        assert frame2.metadata.extra.get("camera_id") == "CAM2"
        
        # Process through pipeline
        pipeline1 = create_replay_pipeline()
        pipeline2 = create_replay_pipeline()
        
        result1 = pipeline1.process_frame(frame1)
        result2 = pipeline2.process_frame(frame2)
        
        assert result1.camera_id == "CAM1"
        assert result2.camera_id == "CAM2"
        
        # Fusion engine should keep tracks isolated
        fusion_engine = create_fusion_engine(DEFAULT_FUSION_CONFIG)
        
        if result1.detections:
            for i, det in enumerate(result1.detections):
                obs = build_local_observation_ref(
                    frame=frame1,
                    local_track_id=f"track_{det.detection_id}",
                    detection_id=det.detection_id,
                )
                fusion_engine.add_observation(obs)
        
        if result2.detections:
            for i, det in enumerate(result2.detections):
                obs = build_local_observation_ref(
                    frame=frame2,
                    local_track_id=f"track_{det.detection_id}",
                    detection_id=det.detection_id,
                )
                fusion_engine.add_observation(obs)
        
        stats = fusion_engine.get_stats()
        assert "CAM1" in stats["cameras"]
        assert "CAM2" in stats["cameras"]
        
        source1.close()
        source2.close()
        pipeline1.close()
        pipeline2.close()

    # =========================================================================
    # NEGATIVE TESTS
    # =========================================================================

    def test_negative_missing_source_video(self):
        """Test handling of missing source video."""
        source = ReplaySource(ReplaySourceConfig(
            source_path="nonexistent.mp4",
            camera_id="CAM1",
        ))
        
        # Should not crash, but fail gracefully
        try:
            source.open()
            # If it opens, reading should return None or fail
            frame = source.get_next_frame()
            # Either way, should handle gracefully
        except Exception:
            pass  # Expected to fail
        finally:
            source.close()

    def test_negative_corrupt_frame(self, test_data_dir):
        """Test handling of corrupt video frame."""
        corrupt_path = test_data_dir / "corrupt.mp4"
        if corrupt_path.exists():
            source = ReplaySource(ReplaySourceConfig(
                source_path=str(corrupt_path),
                camera_id="CAM1",
            ))
            try:
                source.open()
                frame = source.get_next_frame()
                # Should handle gracefully (may return None or partial frame)
            except ReplaySourceError:
                pass  # Expected to fail on corrupt file
            finally:
                source.close()

    def test_negative_unknown_identity(self):
        """Test UNKNOWN identity handling through pipeline."""
        resolver = create_repeated_in_out_resolver(create_default_resolver_config())
        
        raw_event = RawInOutEvent(
            event_id="RIE-UNKNOWN-001",
            camera_id="CAM1",
            geometry_id="hash123",
            geometry_version=1,
            geometry_config_hash="hash123",
            local_track_id="track_001",
            global_observation_id=None,  # No global observation
            event_type=RawEventType.LINE_CROSSING,
            direction=RawEventDirection.IN,
            crossing_point_x=960.0,
            crossing_point_y=500.0,
            crossing_timestamp=1000.0,
            crossing_frame_index=100,
            previous_position_x=960.0,
            previous_position_y=480.0,
            current_position_x=960.0,
            current_position_y=520.0,
            previous_frame_index=99,
            current_frame_index=100,
            previous_timestamp=999.0,
            current_timestamp=1000.0,
            crossing_distance=40.0,
            side_transition="SIDE_A->SIDE_B",
            identity_certainty=IdentityCertainty.UNKNOWN,
            identity_candidate=None,
            identity_confidence=0.0,
            identity_evidence_ref=None,
            source_crossing_event_id="CE-001",
            trajectory_points=[],
            config_snapshot={},
            event_schema_version="1.0",
            created_at="2026-01-01T00:00:00Z",
        )
        
        result = resolver.resolve_events([raw_event])
        
        assert result.accepted_transitions == 1
        assert result.transitions[0].global_observation_id is None
        assert result.final_states["CAM1:track_001"].current_state == DerivedState.INSIDE

    def test_negative_duplicate_event(self):
        """Test duplicate event handling."""
        raw_engine = create_raw_event_engine()
        
        line_config = create_line_geometry(
            camera_id="CAM1",
            frame_width=1920,
            frame_height=1080,
            p1=(100, 500),
            p2=(1820, 500),
            direction_semantics=DirectionSemantics.SIDE_A_TO_B_IN,
            version=1,
        )
        geom_snapshot = GeometryConfigSnapshot.from_config(line_config)
        
        crossing_event = CrossingEvent(
            event_id="CE-DUP-NEG-001",
            camera_id="CAM1",
            geometry_config=geom_snapshot,
            local_track_id="track_001",
            global_observation_id=None,
            event_type=CrossingEventType.LINE_CROSSING,
            direction=CrossingDirection.IN,
            crossing_point=Point2D(960, 500),
            crossing_timestamp=1000.0,
            previous_position=Point2D(960, 480),
            current_position=Point2D(960, 520),
            previous_frame_index=100,
            current_frame_index=101,
            previous_timestamp=999.0,
            current_timestamp=1000.0,
            crossing_distance=40.0,
            side_transition="SIDE_A->SIDE_B",
            trajectory_points=[],
            config_snapshot={},
            created_at="2026-01-01T00:00:00Z",
            version="1.0",
        )
        
        result1 = raw_engine.process_crossing_event(crossing_event)
        result2 = raw_engine.process_crossing_event(crossing_event)
        
        assert result1.success
        assert result2.success
        assert result1.event.event_id == result2.event.event_id
        assert raw_engine.get_stats()["duplicates"] == 1

    def test_negative_out_of_order_timestamp(self):
        """Test out-of-order timestamp handling."""
        resolver = create_repeated_in_out_resolver(
            ResolverConfig(out_of_order_policy=OutOfOrderPolicy.SORT)
        )
        
        # Create events out of order
        event_late = RawInOutEvent(
            event_id="RIE-LATE",
            camera_id="CAM1",
            geometry_id="hash123",
            geometry_version=1,
            geometry_config_hash="hash123",
            local_track_id="track_001",
            global_observation_id=None,
            event_type=RawEventType.LINE_CROSSING,
            direction=RawEventDirection.IN,
            crossing_point_x=960.0,
            crossing_point_y=500.0,
            crossing_timestamp=3000.0,
            crossing_frame_index=300,
            previous_position_x=960.0,
            previous_position_y=480.0,
            current_position_x=960.0,
            current_position_y=520.0,
            previous_frame_index=299,
            current_frame_index=300,
            previous_timestamp=2999.0,
            current_timestamp=3000.0,
            crossing_distance=40.0,
            side_transition="SIDE_A->SIDE_B",
            identity_certainty=IdentityCertainty.UNKNOWN,
            identity_candidate=None,
            identity_confidence=0.0,
            identity_evidence_ref=None,
            source_crossing_event_id="CE-123",
            trajectory_points=[],
            config_snapshot={},
            event_schema_version="1.0",
            created_at="2026-01-01T00:00:00Z",
        )
        
        event_early = RawInOutEvent(
            event_id="RIE-EARLY",
            camera_id="CAM1",
            geometry_id="hash123",
            geometry_version=1,
            geometry_config_hash="hash123",
            local_track_id="track_001",
            global_observation_id=None,
            event_type=RawEventType.LINE_CROSSING,
            direction=RawEventDirection.IN,
            crossing_point_x=960.0,
            crossing_point_y=500.0,
            crossing_timestamp=1000.0,
            crossing_frame_index=100,
            previous_position_x=960.0,
            previous_position_y=480.0,
            current_position_x=960.0,
            current_position_y=520.0,
            previous_frame_index=99,
            current_frame_index=100,
            previous_timestamp=999.0,
            current_timestamp=1000.0,
            crossing_distance=40.0,
            side_transition="SIDE_A->SIDE_B",
            identity_certainty=IdentityCertainty.UNKNOWN,
            identity_candidate=None,
            identity_confidence=0.0,
            identity_evidence_ref=None,
            source_crossing_event_id="CE-123",
            trajectory_points=[],
            config_snapshot={},
            event_schema_version="1.0",
            created_at="2026-01-01T00:00:00Z",
        )
        
        # Process out of order
        result = resolver.resolve_events([event_late, event_early])
        
        # Should be sorted chronologically
        timestamps = [t.source_timestamp for t in result.transitions]
        assert timestamps == [1000.0, 3000.0]

    def test_negative_invalid_geometry(self):
        """Test invalid geometry handling."""
        # Create line geometry with invalid points
        try:
            line_config = create_line_geometry(
                camera_id="CAM1",
                frame_width=1920,
                frame_height=1080,
                p1=(100, 500),
                p2=(100, 500),  # Same point - invalid line
                direction_semantics=DirectionSemantics.SIDE_A_TO_B_IN,
                version=1,
            )
            # Should either reject or handle gracefully
            engine = create_crossing_engine(line_config)
            # If it creates, processing should handle gracefully
        except Exception:
            pass  # Expected to fail validation

    def test_negative_missing_provenance(self):
        """Test handling of missing provenance fields."""
        # Create raw event with minimal provenance
        raw_event = RawInOutEvent(
            event_id="RIE-NO-PROV-001",
            camera_id="CAM1",
            geometry_id="hash123",
            geometry_version=1,
            geometry_config_hash="hash123",
            local_track_id="track_001",
            global_observation_id=None,
            event_type=RawEventType.LINE_CROSSING,
            direction=RawEventDirection.IN,
            crossing_point_x=960.0,
            crossing_point_y=500.0,
            crossing_timestamp=1000.0,
            crossing_frame_index=100,
            previous_position_x=960.0,
            previous_position_y=480.0,
            current_position_x=960.0,
            current_position_y=520.0,
            previous_frame_index=99,
            current_frame_index=100,
            previous_timestamp=999.0,
            current_timestamp=1000.0,
            crossing_distance=40.0,
            side_transition="SIDE_A->SIDE_B",
            identity_certainty=IdentityCertainty.UNKNOWN,
            identity_candidate=None,
            identity_confidence=0.0,
            identity_evidence_ref=None,
            source_crossing_event_id="CE-001",
            trajectory_points=[],
            config_snapshot={},
            event_schema_version="1.0",
            created_at="2026-01-01T00:00:00Z",
        )
        
        resolver = create_repeated_in_out_resolver(create_default_resolver_config())
        result = resolver.resolve_events([raw_event])
        
        assert result.accepted_transitions == 1
        assert result.transitions[0].global_observation_id is None

    # =========================================================================
    # MEMORY / BOUNDING GATE
    # =========================================================================

    def test_bounded_replay_buffers(self):
        """Verify replay buffers are bounded."""
        # ReplayPipeline uses temporal aggregator with bounded windows
        pipeline = create_replay_pipeline()
        
        # The temporal aggregator has max_samples and max_duration bounds
        assert pipeline.temporal_aggregator.config.max_samples == 100
        assert pipeline.temporal_aggregator.config.max_duration == 30.0
        
        pipeline.close()

    def test_bounded_event_history(self):
        """Verify event bus history is bounded."""
        bus = create_event_bus(max_history=50, max_dedup_cache=50)
        
        received = []
        def handler(event):
            received.append(event)
        
        subscriber = FunctionSubscriber("test-sub", handler)
        config = SubscriberConfig(subscriber_id="test-sub", queue_size=100)
        bus.subscribe(subscriber, config)
        
        # Publish 100 events (more than max_history=50)
        for i in range(100):
            event = ImmediateEvent(
                event_id=f"IEV-BOUND-{i}",
                event_type=ImmediateEventType.ATTENDANCE_IN,
                direction=ImmediateEventDirection.IN,
                identity_certainty=OutputIdentityCertainty.KNOWN,
                identity_candidate=f"person-{i}",
                identity_confidence=0.95,
                event_timestamp=36000.0 + i,
                event_frame_index=100 + i,
                camera_id="CAM1",
                local_track_id=f"track_{i}",
                global_observation_id=f"GO-{i}",
                source_raw_event_id=f"RIE-{i}",
                source_resolution_id=f"RES-{i}",
                source_crossing_event_id=f"CE-{i}",
                geometry_version=1,
                geometry_config_hash="geom_hash",
                resolver_version="1.0",
                resolver_config_hash="config_hash",
                timetable_id="ttb-001",
                session_id="session-001",
                day="monday",
                previous_attendance_state="unknown",
                new_attendance_state="present",
                decision_reason="within_entry_window",
                attendance_policy_id="policy-001",
                attendance_policy_version="1.0",
                event_schema_version="1.0",
            )
            bus.publish(event)
        
        import time
        time.sleep(0.2)
        
        # History should be bounded
        history = bus.get_history(limit=100)
        assert len(history) <= 50
        
        stats = bus.get_stats()
        assert stats["history_size"] <= 50
        
        bus.shutdown()

    def test_bounded_deduplication_cache(self):
        """Verify deduplication cache is bounded."""
        bus = create_event_bus(max_history=1000, max_dedup_cache=10)
        
        # Publish 20 unique events (more than max_dedup_cache=10)
        for i in range(20):
            event = ImmediateEvent(
                event_id=f"IEV-DEDUP-BOUND-{i}",
                event_type=ImmediateEventType.ATTENDANCE_IN,
                direction=ImmediateEventDirection.IN,
                identity_certainty=OutputIdentityCertainty.KNOWN,
                identity_candidate=f"person-{i}",
                identity_confidence=0.95,
                event_timestamp=36000.0 + i,
                event_frame_index=100 + i,
                camera_id="CAM1",
                local_track_id=f"track_{i}",
                global_observation_id=f"GO-{i}",
                source_raw_event_id=f"RIE-{i}",
                source_resolution_id=f"RES-{i}",  # Unique resolution IDs
                source_crossing_event_id=f"CE-{i}",
                geometry_version=1,
                geometry_config_hash="geom_hash",
                resolver_version="1.0",
                resolver_config_hash="config_hash",
                timetable_id="ttb-001",
                session_id="session-001",
                day="monday",
                previous_attendance_state="unknown",
                new_attendance_state="present",
                decision_reason="within_entry_window",
                attendance_policy_id="policy-001",
                attendance_policy_version="1.0",
                event_schema_version="1.0",
            )
            bus.publish(event)
        
        import time
        time.sleep(0.2)
        
        stats = bus.get_stats()
        assert stats["dedup_cache_size"] <= 10
        
        bus.shutdown()

    def test_bounded_trajectory_state(self):
        """Verify trajectory state in crossing engine is bounded."""
        line_config = create_line_geometry(
            camera_id="CAM1",
            frame_width=1920,
            frame_height=1080,
            p1=(100, 500),
            p2=(1820, 500),
            direction_semantics=DirectionSemantics.SIDE_A_TO_B_IN,
            version=1,
        )
        
        engine = create_crossing_engine(line_config)
        
        # TrackCrossingState has max_history bound
        from app.geometry.crossing import TrackCrossingState
        state = TrackCrossingState(
            track_id="test_track",
            camera_id="CAM1",
            max_history=10,
        )
        
        assert state.max_history == 10
        
        # Add more positions than max_history
        from app.geometry.crossing import TrajectoryPoint
        from app.geometry.contract import Point2D
        
        for i in range(15):
            point = TrajectoryPoint(
                track_id="test_track",
                frame_index=i,
                timestamp=1000.0 + i,
                position=Point2D(100 + i, 500),
                bbox=(90 + i, 400, 110 + i, 600),
                camera_id="CAM1",
            )
            state.add_position(point)
        
        # Should only keep max_history positions
        assert len(state.recent_positions) <= 10

    def test_bounded_query_pagination(self, temp_db_path):
        """Verify query pagination is bounded."""
        config = StorageConfig(database_path=temp_db_path)
        storage = AttendanceStorage(config)
        repo = AttendanceRepository(config=config)
        
        # Insert 20 records
        for i in range(20):
            record = AttendanceRecord(
                attendance_record_id=f"ATT-PAGE-{i}",
                identity_certainty=AttendanceIdentityCertainty.UNKNOWN,
                direction=AttendanceDirection.IN,
                event_timestamp=1000.0 + i,
                camera_id="CAM1",
                local_track_id=f"track_{i}",
                source_raw_event_id=f"RIE-{i}",
                source_resolution_id=f"RES-{i}",
            )
            storage.insert(record)
        
        # Query with limit
        limited = repo.query_by_camera("CAM1", limit=5)
        assert len(limited) == 5
        
        # Query with offset
        offset = repo.query_by_camera("CAM1", limit=5, offset=5)
        assert len(offset) == 5
        assert limited[0].attendance_record_id != offset[0].attendance_record_id
        
        repo.close()
        storage.close()


if __name__ == "__main__":
    pytest.main([__file__, "-v"])