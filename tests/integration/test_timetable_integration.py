"""
Phase 37A — Integration Tests for Timetable System.

Tests the full integration of:
- Timetable loading
- Calendar/exception handling
- Daily expected-student resolver
- Identity resolution
- Day resolution
- AttendanceEngine integration
- Excel export integration
"""

import tempfile
import os
from datetime import date, datetime, time, timezone
from pathlib import Path

import openpyxl
import pytest
import pytz

from app.attendance.timetable_loader import (
    TimetableLoader,
    load_timetable_from_excel,
    create_sample_timetable_excel,
)
from app.attendance.timetable import (
    Timetable,
    TimetableEntry,
    SessionDay,
    SessionType,
    AttendanceState,
)
from app.attendance.calendar import (
    CalendarEngine,
    CalendarConfig,
    DayType,
    ScheduleException,
    ExceptionType,
    create_calendar_engine,
)
from app.attendance.daily_resolver import (
    DailyExpectedResolver,
    IdentityResolver,
    DayResolver,
    ExpectedStatus,
    create_daily_resolver,
    create_identity_resolver,
    create_day_resolver,
)
from app.attendance.engine import (
    AttendanceEngine,
    AttendancePolicy,
    AttendanceDecisionContext,
    create_attendance_engine_with_resolvers,
)
from app.attendance.policy import (
    DecisionReason,
    IdentityHandlingPolicy,
)
from app.in_out.resolver_contract import (
    ResolvedTransition,
    DerivedState,
    TransitionType,
    ResolutionStatus,
)
from app.replay.fusion import (
    GlobalObservation,
    LocalObservationRef,
    AssociationState,
    AssociationEvidence,
    ReplayTimestamp,
)
from app.attendance.daily_excel import (
    DailyExcelExporter,
    DailyExportRequest,
    create_daily_excel_exporter,
)
from app.attendance.repository import AttendanceRepository, create_attendance_repository
from app.attendance.storage import StorageConfig, create_attendance_storage


class TestCalendarIntegration:
    """Test calendar and exception integration."""
    
    def test_calendar_school_day(self):
        """Test calendar identifies school days correctly."""
        config = CalendarConfig(
            timezone="Asia/Bangkok",
            default_school_days=(0, 1, 2, 3, 4),  # Mon-Fri
            holidays=("2026-01-01",),
        )
        engine = CalendarEngine(config)
        
        # Monday (school day)
        monday = date(2026, 1, 5)  # Monday
        day = engine.get_day(monday)
        assert day.is_school_day is True
        assert day.day_type == DayType.SCHOOL_DAY
        
        # Saturday (weekend)
        saturday = date(2026, 1, 10)  # Saturday
        day = engine.get_day(saturday)
        assert day.is_school_day is False
        assert day.day_type == DayType.HOLIDAY
        
        # Holiday
        holiday = date(2026, 1, 1)  # New Year
        day = engine.get_day(holiday)
        assert day.is_school_day is False
        assert day.day_type == DayType.HOLIDAY
    
    def test_calendar_exception_late_start(self):
        """Test calendar exception for late start."""
        config = CalendarConfig(
            timezone="Asia/Bangkok",
            default_school_days=(0, 1, 2, 3, 4),
            exceptions=(
                {
                    "exception_id": "EXC-001",
                    "date": "2026-01-05",
                    "session_id": "MATH101_MON",
                    "exception_type": "late_start",
                    "description": "Assembly",
                    "new_entry_time": 28800,  # 08:00:00
                    "new_entry_window_start": 28500,
                    "new_entry_window_end": 29100,
                },
            ),
        )
        engine = CalendarEngine(config)
        
        target_date = date(2026, 1, 5)
        exc = engine.get_applicable_exception(target_date, "MATH101_MON")
        
        assert exc is not None
        assert exc.exception_type == ExceptionType.LATE_START
        assert exc.new_entry_time == 28800
    
    def test_calendar_exception_student_specific(self):
        """Test calendar exception for specific student."""
        config = CalendarConfig(
            timezone="Asia/Bangkok",
            default_school_days=(0, 1, 2, 3, 4),
            exceptions=(
                {
                    "exception_id": "EXC-002",
                    "date": "2026-01-05",
                    "session_id": "MATH101_MON",
                    "exception_type": "early_dismissal",
                    "description": "Doctor appointment",
                    "new_exit_time": 36000,  # 10:00:00
                    "student_id": "HS001",
                },
            ),
        )
        engine = CalendarEngine(config)
        
        target_date = date(2026, 1, 5)
        
        # Exception applies to HS001
        exc = engine.get_applicable_exception(target_date, "MATH101_MON", student_id="HS001")
        assert exc is not None
        assert exc.exception_type == ExceptionType.EARLY_DISMISSAL
        
        # Exception does NOT apply to HS002
        exc = engine.get_applicable_exception(target_date, "MATH101_MON", student_id="HS002")
        assert exc is None
    
    def test_calendar_resolve_session_times(self):
        """Test calendar resolves session times with exceptions."""
        config = CalendarConfig(
            timezone="Asia/Bangkok",
            default_school_days=(0, 1, 2, 3, 4),
            exceptions=(
                {
                    "exception_id": "EXC-003",
                    "date": "2026-01-05",
                    "session_id": "MATH101_MON",
                    "exception_type": "late_start",
                    "description": "Assembly",
                    "new_entry_time": 28800,  # 08:00:00
                    "new_entry_window_start": 28500,
                    "new_entry_window_end": 29100,
                },
            ),
        )
        engine = CalendarEngine(config)
        
        target_date = date(2026, 1, 5)
        
        # Base times
        base_entry = 27000  # 07:30:00
        base_exit = 43200   # 12:00:00
        base_entry_ws = 26700
        base_entry_we = 27300
        base_late = 600
        base_exit_ws = 42900
        base_exit_we = 43500
        
        resolved = engine.resolve_session_times(
            target_date, "MATH101_MON",
            base_entry, base_exit,
            base_entry_ws, base_entry_we,
            base_late,
            base_exit_ws, base_exit_we,
        )
        
        # Entry time should be updated to 08:00:00
        assert resolved[0] == 28800
        assert resolved[2] == 28500
        assert resolved[3] == 29100
        # Exit time unchanged
        assert resolved[1] == 43200
    
    def test_calendar_timezone_conversion(self):
        """Test calendar converts UTC timestamp to local date correctly."""
        config = CalendarConfig(timezone="Asia/Bangkok")  # UTC+7
        engine = CalendarEngine(config)
        
        # 2026-01-05 00:00:00 UTC = 2026-01-05 07:00:00 Bangkok
        timestamp = datetime(2026, 1, 5, 0, 0, 0, tzinfo=pytz.UTC).timestamp()
        local_date = engine.get_date_from_timestamp(timestamp)
        assert local_date == date(2026, 1, 5)
        
        # 2026-01-04 17:00:00 UTC = 2026-01-05 00:00:00 Bangkok
        timestamp = datetime(2026, 1, 4, 17, 0, 0, tzinfo=pytz.UTC).timestamp()
        local_date = engine.get_date_from_timestamp(timestamp)
        assert local_date == date(2026, 1, 5)
        
        session_day = engine.get_session_day_enum_from_timestamp(timestamp)
        assert session_day == SessionDay.MONDAY


class TestDailyResolverIntegration:
    """Test daily expected-student resolver integration."""
    
    def create_sample_timetable(self) -> Timetable:
        """Create a sample timetable for testing."""
        entries = [
            TimetableEntry(
                entry_id="ENT-1",
                person_id="HS001",
                session_id="MATH101_MON",
                session_type=SessionType.MORNING,
                day=SessionDay.MONDAY,
                class_name="Math 101",
                person_name="Student One",
                entry_time=27000,      # 07:30:00
                exit_time=43200,       # 12:00:00
                entry_window_start=26700,
                entry_window_end=27300,
                late_tolerance=600,
                exit_window_start=42900,
                exit_window_end=43500,
            ),
            TimetableEntry(
                entry_id="ENT-2",
                person_id="HS002",
                session_id="PHYS101_MON",
                session_type=SessionType.MORNING,
                day=SessionDay.MONDAY,
                class_name="Physics 101",
                person_name="Student Two",
                entry_time=30600,      # 08:30:00
                exit_time=46800,       # 13:00:00
                entry_window_start=30300,
                entry_window_end=30900,
                late_tolerance=600,
                exit_window_start=46500,
                exit_window_end=47100,
            ),
            TimetableEntry(
                entry_id="ENT-3",
                person_id="HS001",
                session_id="ENG101_TUE",
                session_type=SessionType.AFTERNOON,
                day=SessionDay.TUESDAY,
                class_name="English 101",
                person_name="Student One",
                entry_time=46800,      # 13:00:00
                exit_time=63000,       # 17:30:00
                entry_window_start=46500,
                entry_window_end=47100,
                late_tolerance=600,
                exit_window_start=62700,
                exit_window_end=63300,
            ),
        ]
        return Timetable(
            timetable_id="TTB-test",
            timetable_version="1.0",
            entries=entries,
        )
    
    def test_daily_resolver_school_day(self):
        """Test daily resolver on a school day."""
        timetable = self.create_sample_timetable()
        calendar = create_calendar_engine(timezone="Asia/Bangkok")
        resolver = create_daily_resolver(timetable, calendar, ["HS001", "HS002", "HS003"])
        
        # Monday 2026-01-05
        target_date = date(2026, 1, 5)
        result = resolver.resolve_for_date(target_date)
        
        assert result.date == target_date
        assert result.session_day == SessionDay.MONDAY
        assert result.day_type == DayType.SCHOOL_DAY
        assert result.total_scheduled == 2  # HS001 and HS002
        assert result.total_not_scheduled == 1  # HS003
        
        # Check HS001
        hs001 = next(s for s in result.expected_students if s.student_id == "HS001")
        assert hs001.status == ExpectedStatus.SCHEDULED
        assert len(hs001.sessions) == 1
        assert hs001.sessions[0].session_id == "MATH101_MON"
        
        # Check HS002
        hs002 = next(s for s in result.expected_students if s.student_id == "HS002")
        assert hs002.status == ExpectedStatus.SCHEDULED
        assert len(hs002.sessions) == 1
        assert hs002.sessions[0].session_id == "PHYS101_MON"
        
        # Check HS003 (enrolled but no sessions Monday)
        hs003 = next(s for s in result.expected_students if s.student_id == "HS003")
        assert hs003.status == ExpectedStatus.NOT_SCHEDULED
        assert len(hs003.sessions) == 0
    
    def test_daily_resolver_holiday(self):
        """Test daily resolver on a holiday."""
        timetable = self.create_sample_timetable()
        calendar = create_calendar_engine(
            timezone="Asia/Bangkok",
            holidays=["2026-01-01"],
        )
        resolver = create_daily_resolver(timetable, calendar, ["HS001", "HS002"])
        
        # Holiday
        target_date = date(2026, 1, 1)
        result = resolver.resolve_for_date(target_date)
        
        assert result.day_type == DayType.HOLIDAY
        for student in result.expected_students:
            assert student.status == ExpectedStatus.HOLIDAY
    
    def test_daily_resolver_exception_late_start(self):
        """Test daily resolver with late start exception."""
        timetable = self.create_sample_timetable()
        calendar = create_calendar_engine(
            timezone="Asia/Bangkok",
            exceptions=[
                {
                    "exception_id": "EXC-001",
                    "date": "2026-01-05",
                    "session_id": "MATH101_MON",
                    "exception_type": "late_start",
                    "description": "Assembly",
                    "new_entry_time": 28800,  # 08:00:00
                    "new_entry_window_start": 28500,
                    "new_entry_window_end": 29100,
                },
            ],
        )
        resolver = create_daily_resolver(timetable, calendar, ["HS001", "HS002"])
        
        target_date = date(2026, 1, 5)
        result = resolver.resolve_for_date(target_date)
        
        assert result.day_type == DayType.EXCEPTION
        
        hs001 = next(s for s in result.expected_students if s.student_id == "HS001")
        assert hs001.status == ExpectedStatus.LATER_START
        assert hs001.sessions[0].exception is not None
        assert hs001.sessions[0].exception.exception_type == ExceptionType.LATE_START
        assert hs001.sessions[0].effective_entry_time == 28800
    
    def test_daily_resolver_exception_student_specific(self):
        """Test daily resolver with student-specific exception."""
        timetable = self.create_sample_timetable()
        calendar = create_calendar_engine(
            timezone="Asia/Bangkok",
            exceptions=[
                {
                    "exception_id": "EXC-002",
                    "date": "2026-01-05",
                    "session_id": "MATH101_MON",
                    "exception_type": "early_dismissal",
                    "description": "Doctor appointment",
                    "new_exit_time": 36000,  # 10:00:00
                    "student_id": "HS001",
                },
            ],
        )
        resolver = create_daily_resolver(timetable, calendar, ["HS001", "HS002"])
        
        target_date = date(2026, 1, 5)
        result = resolver.resolve_for_date(target_date)
        
        hs001 = next(s for s in result.expected_students if s.student_id == "HS001")
        assert hs001.status == ExpectedStatus.EARLIER_DEPARTURE
        assert hs001.sessions[0].effective_exit_time == 36000
        
        # HS002 should not be affected (their session is PHYS101_MON, not MATH101_MON)
        hs002 = next(s for s in result.expected_students if s.student_id == "HS002")
        assert hs002.status == ExpectedStatus.SCHEDULED
        # HS002's session is PHYS101_MON with exit_time=46800
        assert hs002.sessions[0].effective_exit_time == 46800
    
    def test_daily_resolver_is_student_expected_at(self):
        """Test checking if student is expected at specific timestamp."""
        timetable = self.create_sample_timetable()
        calendar = create_calendar_engine(timezone="Asia/Bangkok")
        resolver = create_daily_resolver(timetable, calendar, ["HS001", "HS002"])
        
        target_date = date(2026, 1, 5)
        
        # HS001 entry window: 07:25-07:35 (26700-27300)
        # HS001 exit window: 11:55-12:05 (42900-43500)
        
        # Before entry window
        expected, session, reason = resolver.is_student_expected_at(
            target_date, "HS001", 26000  # 07:13:20
        )
        assert expected is True
        assert reason == "before_entry_window"
        
        # Within entry window
        expected, session, reason = resolver.is_student_expected_at(
            target_date, "HS001", 27000  # 07:30:00
        )
        assert expected is True
        assert reason == "within_session_window"
        
        # After exit window
        expected, session, reason = resolver.is_student_expected_at(
            target_date, "HS001", 44000  # 12:13:20
        )
        assert expected is True
        assert reason == "after_exit_window"
        
        # HS003 not scheduled
        expected, session, reason = resolver.is_student_expected_at(
            target_date, "HS003", 27000
        )
        assert expected is False


class TestIdentityResolverIntegration:
    """Test identity resolver integration with GlobalObservation."""
    
    def create_mock_global_observation(self, person_id: str = "HS001") -> GlobalObservation:
        """Create a mock GlobalObservation for testing."""
        from app.replay.fusion import IdentityHypothesis, HypothesisState
        
        # Create local observation refs
        obs1 = LocalObservationRef(
            camera_id="CAM1",
            local_track_id="track_001",
            observation_id="CAM1_track_001_f100",
            frame_index=100,
            timestamp=ReplayTimestamp(value=1000.0, source="frame"),
            identity_hypothesis=IdentityHypothesis(
                hypothesis_id="HYP-1",
                candidate_identity=person_id,
                weighted_score=8.5,
                best_similarity=0.85,
                state=HypothesisState.CONFIDENT,
            ),
        )
        
        go = GlobalObservation(
            global_observation_id="GO-test-123",
            observations=(obs1,),
            association_state=AssociationState.ASSOCIATED,
            association_evidence=AssociationEvidence(
                timestamp_delta=0.1,
                timestamp_compatible=True,
                timestamp_tolerance=1.0,
                identity_evidence_support=0.85,
                identity_candidates=[person_id],
                camera_ids=("CAM1",),
            ),
            temporal_start=ReplayTimestamp(value=1000.0, source="fusion"),
            temporal_end=ReplayTimestamp(value=1000.1, source="fusion"),
            temporal_span=0.1,
            camera_ids=("CAM1",),
            local_track_ids=("CAM1:track_001",),
            primary_identity_candidate=person_id,
            identity_confidence=0.85,
            identity_state=HypothesisState.CONFIDENT,
        )
        return go
    
    def test_identity_resolver_known(self):
        """Test identity resolver with known identity."""
        enrollment_ids = ["HS001", "HS002", "HS003"]
        resolver = create_identity_resolver(enrollment_ids)
        
        go = self.create_mock_global_observation("HS001")
        resolution = resolver.resolve_from_global_observation(go)
        
        assert resolution.person_id == "HS001"
        assert resolution.student_id == "HS001"
        assert resolution.identity_certainty == "known"
        assert resolution.identity_confidence == 0.85
        assert resolution.resolution_method == "global_observation_primary_candidate"
    
    def test_identity_resolver_unknown(self):
        """Test identity resolver with unknown identity."""
        enrollment_ids = ["HS001", "HS002", "HS003"]
        resolver = create_identity_resolver(enrollment_ids)
        
        go = self.create_mock_global_observation("HS999")  # Not enrolled
        resolution = resolver.resolve_from_global_observation(go)
        
        assert resolution.person_id == "HS999"
        assert resolution.student_id is None
        assert resolution.identity_certainty == "unknown"
    
    def test_identity_resolver_no_candidate(self):
        """Test identity resolver with no identity candidate."""
        enrollment_ids = ["HS001", "HS002", "HS003"]
        resolver = create_identity_resolver(enrollment_ids)
        
        # Create GO without identity
        obs1 = LocalObservationRef(
            camera_id="CAM1",
            local_track_id="track_001",
            observation_id="CAM1_track_001_f100",
            frame_index=100,
            timestamp=ReplayTimestamp(value=1000.0, source="frame"),
        )
        
        go = GlobalObservation(
            global_observation_id="GO-test-456",
            observations=(obs1,),
            association_state=AssociationState.ASSOCIATED,
            association_evidence=AssociationEvidence(
                timestamp_delta=0.1,
                timestamp_compatible=True,
                timestamp_tolerance=1.0,
                identity_evidence_support=0.0,
                identity_candidates=[],
                camera_ids=("CAM1",),
            ),
            temporal_start=ReplayTimestamp(value=1000.0, source="fusion"),
            temporal_end=ReplayTimestamp(value=1000.1, source="fusion"),
            temporal_span=0.1,
            camera_ids=("CAM1",),
            local_track_ids=("CAM1:track_001",),
            primary_identity_candidate=None,
            identity_confidence=0.0,
            identity_state=None,
        )
        
        resolution = resolver.resolve_from_global_observation(go)
        
        assert resolution.person_id is None
        assert resolution.student_id is None
        assert resolution.identity_certainty == "insufficient"


class TestDayResolverIntegration:
    """Test day resolver integration."""
    
    def test_day_resolver_from_timestamp(self):
        """Test day resolver converts timestamp to SessionDay."""
        calendar = create_calendar_engine(timezone="Asia/Bangkok")
        resolver = create_day_resolver(calendar)
        
        # Monday 2026-01-05 07:30:00 Bangkok = 2026-01-05 00:30:00 UTC
        timestamp = datetime(2026, 1, 5, 0, 30, 0).timestamp()
        day = resolver.resolve_day(timestamp)
        assert day == SessionDay.MONDAY
        
        # Tuesday 2026-01-06 08:30:00 Bangkok = 2026-01-06 01:30:00 UTC
        timestamp = datetime(2026, 1, 6, 1, 30, 0).timestamp()
        day = resolver.resolve_day(timestamp)
        assert day == SessionDay.TUESDAY
    
    def test_day_resolver_date_from_timestamp(self):
        """Test day resolver converts timestamp to date."""
        calendar = create_calendar_engine(timezone="Asia/Bangkok")
        resolver = create_day_resolver(calendar)
        
        # 2026-01-04 17:00:00 UTC = 2026-01-05 00:00:00 Bangkok
        timestamp = datetime(2026, 1, 4, 17, 0, 0, tzinfo=pytz.UTC).timestamp()
        target_date = resolver.resolve_date(timestamp)
        assert target_date == date(2026, 1, 5)


class TestAttendanceEngineIntegration:
    """Test AttendanceEngine integration with Phase 37A resolvers."""
    
    def create_sample_timetable(self) -> Timetable:
        """Create a sample timetable for testing."""
        entries = [
            TimetableEntry(
                entry_id="ENT-1",
                person_id="HS001",
                session_id="MATH101_MON",
                session_type=SessionType.MORNING,
                day=SessionDay.MONDAY,
                class_name="Math 101",
                entry_time=27000,      # 07:30:00
                exit_time=43200,       # 12:00:00
                entry_window_start=26700,
                entry_window_end=27300,
                late_tolerance=600,
                exit_window_start=42900,
                exit_window_end=43500,
            ),
        ]
        return Timetable(
            timetable_id="TTB-test",
            timetable_version="1.0",
            entries=entries,
        )
    
    def create_mock_resolved_transition(self, direction: str = "in") -> ResolvedTransition:
        """Create a mock ResolvedTransition for testing."""
        # 2026-01-05 07:30:00 Bangkok = 2026-01-05 00:30:00 UTC
        # Use pytz.UTC for proper UTC handling
        dt_utc = pytz.UTC.localize(datetime(2026, 1, 5, 0, 30, 0))
        timestamp = dt_utc.timestamp()
        return ResolvedTransition(
            resolution_id="RES-test-123",
            source_raw_event_id="RAW-test-123",
            camera_id="CAM1",
            local_track_id="track_001",
            global_observation_id="GO-test-123",
            direction=direction,
            transition_type=TransitionType.IN if direction == "in" else TransitionType.OUT,
            previous_state=DerivedState.OUTSIDE if direction == "in" else DerivedState.INSIDE,
            new_state=DerivedState.INSIDE if direction == "in" else DerivedState.OUTSIDE,
            source_timestamp=timestamp,
            source_frame_index=100,
            resolver_version="1.0",
            resolver_config_hash="abc123",
            resolution_status=ResolutionStatus.ACCEPTED,
            source_crossing_event_id="CROSS-test-123",
            geometry_version=1,
            geometry_config_hash="geom123",
        )
    
    def create_mock_global_observation(self) -> GlobalObservation:
        """Create a mock GlobalObservation for testing."""
        from app.replay.fusion import IdentityHypothesis, HypothesisState
        
        obs1 = LocalObservationRef(
            camera_id="CAM1",
            local_track_id="track_001",
            observation_id="CAM1_track_001_f100",
            frame_index=100,
            timestamp=ReplayTimestamp(value=1000.0, source="frame"),
            identity_hypothesis=IdentityHypothesis(
                hypothesis_id="HYP-1",
                candidate_identity="HS001",
                weighted_score=8.5,
                best_similarity=0.85,
                state=HypothesisState.CONFIDENT,
            ),
        )
        
        return GlobalObservation(
            global_observation_id="GO-test-123",
            observations=(obs1,),
            association_state=AssociationState.ASSOCIATED,
            association_evidence=AssociationEvidence(
                timestamp_delta=0.1,
                timestamp_compatible=True,
                timestamp_tolerance=1.0,
                identity_evidence_support=0.85,
                identity_candidates=["HS001"],
                camera_ids=("CAM1",),
            ),
            temporal_start=ReplayTimestamp(value=1000.0, source="fusion"),
            temporal_end=ReplayTimestamp(value=1000.1, source="fusion"),
            temporal_span=0.1,
            camera_ids=("CAM1",),
            local_track_ids=("CAM1:track_001",),
            primary_identity_candidate="HS001",
            identity_confidence=0.85,
            identity_state=HypothesisState.CONFIDENT,
        )
    
    def test_attendance_engine_auto_decision(self):
        """Test AttendanceEngine make_decision_auto method."""
        policy = AttendancePolicy(
            policy_id="POL-test",
            policy_version="1.0",
        )
        timetable = self.create_sample_timetable()
        calendar = create_calendar_engine(timezone="Asia/Bangkok")
        enrollment_ids = ["HS001", "HS002"]
        identity_resolver = create_identity_resolver(enrollment_ids)
        
        engine, _, _, _ = create_attendance_engine_with_resolvers(
            policy, timetable, calendar, enrollment_ids
        )
        
        transition = self.create_mock_resolved_transition("in")
        go = self.create_mock_global_observation()
        
        decision = engine.make_decision_auto(
            resolved_transition=transition,
            timetable=timetable,
            calendar_engine=calendar,
            identity_resolver=identity_resolver,
            global_observation=go,
        )
        
        assert decision.identity_candidate == "HS001"
        assert decision.identity_certainty == "known"
        # When person_id_override is provided, identity_confidence is set to 1.0
        assert decision.identity_confidence == 1.0
        assert decision.new_attendance_state == "present"
        assert decision.decision_reason == "within_entry_window"
        assert decision.day == "monday"
    
    def test_attendance_engine_auto_decision_late(self):
        """Test AttendanceEngine auto decision for late arrival."""
        policy = AttendancePolicy(
            policy_id="POL-test",
            policy_version="1.0",
        )
        timetable = self.create_sample_timetable()
        calendar = create_calendar_engine(timezone="Asia/Bangkok")
        enrollment_ids = ["HS001", "HS002"]
        identity_resolver = create_identity_resolver(enrollment_ids)
        
        engine, _, _, _ = create_attendance_engine_with_resolvers(
            policy, timetable, calendar, enrollment_ids
        )
        
        # Late arrival: 07:40:00 Bangkok = 2026-01-05 00:40:00 UTC
        dt_utc = pytz.UTC.localize(datetime(2026, 1, 5, 0, 40, 0))
        late_timestamp = dt_utc.timestamp()
        
        transition = ResolvedTransition(
            resolution_id="RES-test-456",
            source_raw_event_id="RAW-test-456",
            camera_id="CAM1",
            local_track_id="track_001",
            global_observation_id="GO-test-123",
            direction="in",
            transition_type=TransitionType.IN,
            previous_state=DerivedState.OUTSIDE,
            new_state=DerivedState.INSIDE,
            source_timestamp=late_timestamp,
            source_frame_index=100,
            resolver_version="1.0",
            resolver_config_hash="abc123",
            resolution_status=ResolutionStatus.ACCEPTED,
            source_crossing_event_id="CROSS-test-456",
            geometry_version=1,
            geometry_config_hash="geom123",
        )
        go = self.create_mock_global_observation()
        
        decision = engine.make_decision_auto(
            resolved_transition=transition,
            timetable=timetable,
            calendar_engine=calendar,
            identity_resolver=identity_resolver,
            global_observation=go,
        )
        
        assert decision.new_attendance_state == "late"
        assert decision.decision_reason == "late_within_tolerance"
    
    def test_attendance_engine_with_daily_resolver(self):
        """Test AttendanceEngine make_decision_with_daily_resolver method."""
        policy = AttendancePolicy(
            policy_id="POL-test",
            policy_version="1.0",
        )
        timetable = self.create_sample_timetable()
        calendar = create_calendar_engine(timezone="Asia/Bangkok")
        enrollment_ids = ["HS001", "HS002"]
        
        engine, identity_resolver, day_resolver, daily_resolver = create_attendance_engine_with_resolvers(
            policy, timetable, calendar, enrollment_ids
        )
        
        transition = self.create_mock_resolved_transition("in")
        go = self.create_mock_global_observation()
        
        decision = engine.make_decision_with_daily_resolver(
            resolved_transition=transition,
            timetable=timetable,
            daily_resolver=daily_resolver,
            global_observation=go,
        )
        
        assert decision.identity_candidate == "HS001"
        assert decision.new_attendance_state == "present"


class TestExcelExportIntegration:
    """Test Excel export integration with timetable data."""
    
    def create_sample_timetable(self) -> Timetable:
        """Create a sample timetable for testing."""
        entries = [
            TimetableEntry(
                entry_id="ENT-1",
                person_id="HS001",
                session_id="MATH101_MON",
                session_type=SessionType.MORNING,
                day=SessionDay.MONDAY,
                class_name="Math 101",
                person_name="Student One",
                entry_time=27000,
                exit_time=43200,
                entry_window_start=26700,
                entry_window_end=27300,
                late_tolerance=600,
                exit_window_start=42900,
                exit_window_end=43500,
            ),
            TimetableEntry(
                entry_id="ENT-2",
                person_id="HS002",
                session_id="PHYS101_MON",
                session_type=SessionType.MORNING,
                day=SessionDay.MONDAY,
                class_name="Physics 101",
                person_name="Student Two",
                entry_time=30600,
                exit_time=46800,
                entry_window_start=30300,
                entry_window_end=30900,
                late_tolerance=600,
                exit_window_start=46500,
                exit_window_end=47100,
            ),
        ]
        return Timetable(
            timetable_id="TTB-test",
            timetable_version="1.0",
            entries=entries,
        )
    
    def test_excel_export_with_timetable(self):
        """Test Excel export with timetable creates expected schedule sheet."""
        # Create temp database
        temp_db = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        temp_db.close()
        
        # Create temp output
        temp_output = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp_output.close()
        
        try:
            # Setup storage and repository
            config = StorageConfig(database_path=temp_db.name)
            storage = create_attendance_storage(config)
            repository = create_attendance_repository(storage)
            
            # Create exporter
            exporter = create_daily_excel_exporter(repository)
            
            # Create timetable
            timetable = self.create_sample_timetable()
            
            # Create export request with timetable
            request = DailyExportRequest(
                date=date(2026, 1, 5),
                output_path=temp_output.name,
                timezone="Asia/Bangkok",
                timetable=timetable,
                include_events_sheet=True,
                include_provenance_sheet=True,
                include_summary_sheet=True,
            )
            
            result = exporter.export_daily_attendance(request)
            
            assert result.success is True
            assert "DAILY_ATTENDANCE" in result.sheets_created
            assert "EXPECTED_SCHEDULE" in result.sheets_created
            assert "EVENTS" in result.sheets_created
            assert "SUMMARY" in result.sheets_created
            assert "PROVENANCE" in result.sheets_created
            
            # Verify output file exists
            assert os.path.exists(temp_output.name)
            
            # Verify content
            wb = openpyxl.load_workbook(temp_output.name)
            assert "EXPECTED_SCHEDULE" in wb.sheetnames
            
            ws = wb["EXPECTED_SCHEDULE"]
            assert ws.cell(row=1, column=1).value == "No."
            assert ws.cell(row=1, column=2).value == "Student ID"
            assert ws.max_row >= 3  # Header + 2 students
            
        finally:
            # Close exporter to release database connection
            if 'exporter' in locals():
                exporter.close()
            os.unlink(temp_db.name)
            if os.path.exists(temp_output.name):
                os.unlink(temp_output.name)
    
    def test_excel_export_without_timetable(self):
        """Test Excel export without timetable (backward compatibility)."""
        temp_db = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        temp_db.close()
        
        temp_output = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp_output.close()
        
        try:
            config = StorageConfig(database_path=temp_db.name)
            storage = create_attendance_storage(config)
            repository = create_attendance_repository(storage)
            
            exporter = create_daily_excel_exporter(repository)
            
            request = DailyExportRequest(
                date=date(2026, 1, 5),
                output_path=temp_output.name,
                timezone="Asia/Bangkok",
                timetable=None,  # No timetable
            )
            
            result = exporter.export_daily_attendance(request)
            
            assert result.success is True
            assert "DAILY_ATTENDANCE" in result.sheets_created
            assert "EXPECTED_SCHEDULE" not in result.sheets_created
            
        finally:
            # Close exporter to release database connection
            if 'exporter' in locals():
                exporter.close()
            os.unlink(temp_db.name)
            if os.path.exists(temp_output.name):
                os.unlink(temp_output.name)


class TestEnrollmentNpyCompatibility:
    """Test enrollment/.npy compatibility with timetable."""
    
    def test_enrollment_person_ids_match_timetable(self):
        """Test that timetable student_ids can be validated against enrollment."""
        # Load actual enrollment metadata
        import json
        metadata_path = "data/enrollment_db/embeddings.npy.metadata.json"
        
        if os.path.exists(metadata_path):
            with open(metadata_path, "r") as f:
                metadata = json.load(f)
            
            enrollment_person_ids = metadata.get("person_ids", [])
            assert "HS001" in enrollment_person_ids
            assert "HS002" in enrollment_person_ids
            assert "HS003" in enrollment_person_ids
            
            # Create timetable with valid student_ids
            entries = [
                TimetableEntry(
                    entry_id="ENT-1",
                    person_id="HS001",
                    session_id="MATH101_MON",
                    session_type=SessionType.MORNING,
                    day=SessionDay.MONDAY,
                    class_name="Math 101",
                    entry_time=27000,
                    exit_time=43200,
                    entry_window_start=26700,
                    entry_window_end=27300,
                    late_tolerance=600,
                    exit_window_start=42900,
                    exit_window_end=43500,
                ),
            ]
            timetable = Timetable(
                timetable_id="TTB-test",
                timetable_version="1.0",
                entries=entries,
            )
            
            # Load with enrollment validation
            loader = TimetableLoader(enrollment_person_ids=enrollment_person_ids)
            
            # Create temp Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Timetable"
            headers = ["student_id", "class_name", "day", "session_type", "entry_time", "exit_time"]
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            ws.cell(row=2, column=1, value="HS001")
            ws.cell(row=2, column=2, value="Math 101")
            ws.cell(row=2, column=3, value="monday")
            ws.cell(row=2, column=4, value="morning")
            ws.cell(row=2, column=5, value="07:30:00")
            ws.cell(row=2, column=6, value="12:00:00")
            
            temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            wb.save(temp_file.name)
            temp_file.close()
            
            try:
                result = loader.load_from_excel(temp_file.name)
                assert result.success is True
            finally:
                os.unlink(temp_file.name)
            
            # Test with invalid student_id
            ws.cell(row=2, column=1, value="HS999")
            temp_file2 = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            wb.save(temp_file2.name)
            temp_file2.close()
            
            try:
                result = loader.load_from_excel(temp_file2.name)
                assert result.success is False
                assert any("not found in enrollment database" in e.message for e in result.errors)
            finally:
                os.unlink(temp_file2.name)


class TestVideoProvenanceMapping:
    """Test video/provenance identity mapping."""
    
    def test_global_observation_preserves_provenance_chain(self):
        """Test GlobalObservation preserves camera->track->person->student chain."""
        from app.replay.fusion import IdentityHypothesis, HypothesisState
        
        # Create observations from multiple cameras
        obs1 = LocalObservationRef(
            camera_id="CAM1",
            local_track_id="track_A17",
            observation_id="CAM1_track_A17_f100",
            frame_index=100,
            timestamp=ReplayTimestamp(value=1000.0, source="frame"),
            detection_id="DET-1",
            face_crop_id="CROP-1",
            quality_class="GOOD",
            identity_hypothesis=IdentityHypothesis(
                hypothesis_id="HYP-1",
                candidate_identity="HS001",
                weighted_score=8.5,
                best_similarity=0.85,
                state=HypothesisState.CONFIDENT,
            ),
        )
        
        obs2 = LocalObservationRef(
            camera_id="CAM2",
            local_track_id="track_B04",
            observation_id="CAM2_track_B04_f105",
            frame_index=105,
            timestamp=ReplayTimestamp(value=1000.1, source="frame"),
            detection_id="DET-2",
            face_crop_id="CROP-2",
            quality_class="GOOD",
            identity_hypothesis=IdentityHypothesis(
                hypothesis_id="HYP-2",
                candidate_identity="HS001",
                weighted_score=8.2,
                best_similarity=0.82,
                state=HypothesisState.CONFIDENT,
            ),
        )
        
        go = GlobalObservation(
            global_observation_id="GO-cross-cam-123",
            observations=(obs1, obs2),
            association_state=AssociationState.ASSOCIATED,
            association_evidence=AssociationEvidence(
                timestamp_delta=0.1,
                timestamp_compatible=True,
                timestamp_tolerance=1.0,
                identity_evidence_support=0.85,
                identity_candidates=["HS001"],
                camera_ids=("CAM1", "CAM2"),
            ),
            temporal_start=ReplayTimestamp(value=1000.0, source="fusion"),
            temporal_end=ReplayTimestamp(value=1000.1, source="fusion"),
            temporal_span=0.1,
            camera_ids=("CAM1", "CAM2"),
            local_track_ids=("CAM1:track_A17", "CAM2:track_B04"),
            primary_identity_candidate="HS001",
            identity_confidence=0.85,
            identity_state=HypothesisState.CONFIDENT,
        )
        
        # Verify provenance chain
        assert go.camera_ids == ("CAM1", "CAM2")
        assert go.local_track_ids == ("CAM1:track_A17", "CAM2:track_B04")
        assert go.primary_identity_candidate == "HS001"
        assert len(go.observations) == 2
        
        # Verify each observation preserves its camera-local identity
        cam1_obs = next(o for o in go.observations if o.camera_id == "CAM1")
        assert cam1_obs.local_track_id == "track_A17"
        assert cam1_obs.detection_id == "DET-1"
        assert cam1_obs.face_crop_id == "CROP-1"
        
        cam2_obs = next(o for o in go.observations if o.camera_id == "CAM2")
        assert cam2_obs.local_track_id == "track_B04"
        assert cam2_obs.detection_id == "DET-2"
        assert cam2_obs.face_crop_id == "CROP-2"


class TestRegressionPhase26:
    """Regression tests for Phase 26 functionality."""
    
    def test_timetable_entry_serialization(self):
        """Test TimetableEntry serialization round-trip."""
        entry = TimetableEntry(
            entry_id="ENT-1",
            person_id="HS001",
            session_id="MATH101_MON",
            session_type=SessionType.MORNING,
            day=SessionDay.MONDAY,
            class_name="Math 101",
            person_name="Student One",
            entry_time=27000,
            exit_time=43200,
            entry_window_start=26700,
            entry_window_end=27300,
            late_tolerance=600,
            exit_window_start=42900,
            exit_window_end=43500,
            timetable_version="1.0",
        )
        
        # Dict round-trip
        entry_dict = entry.to_dict()
        entry2 = TimetableEntry.from_dict(entry_dict)
        assert entry2.entry_id == entry.entry_id
        assert entry2.person_id == entry.person_id
        assert entry2.day == entry.day
        assert entry2.entry_time == entry.entry_time
        
        # JSON round-trip
        entry_json = entry.to_json()
        entry3 = TimetableEntry.from_json(entry_json)
        assert entry3.entry_id == entry.entry_id
    
    def test_timetable_serialization(self):
        """Test Timetable serialization round-trip."""
        entries = [
            TimetableEntry(
                entry_id="ENT-1",
                person_id="HS001",
                session_id="MATH101_MON",
                session_type=SessionType.MORNING,
                day=SessionDay.MONDAY,
                class_name="Math 101",
                entry_time=27000,
                exit_time=43200,
                entry_window_start=26700,
                entry_window_end=27300,
                late_tolerance=600,
                exit_window_start=42900,
                exit_window_end=43500,
            ),
        ]
        timetable = Timetable(
            timetable_id="TTB-test",
            timetable_version="1.0",
            entries=entries,
        )
        
        # Dict round-trip
        tt_dict = timetable.to_dict()
        tt2 = Timetable.from_dict(tt_dict)
        assert tt2.timetable_id == timetable.timetable_id
        assert len(tt2.entries) == 1
        assert tt2.entries[0].person_id == "HS001"
        
        # JSON round-trip
        tt_json = timetable.to_json()
        tt3 = Timetable.from_json(tt_json)
        assert tt3.timetable_id == timetable.timetable_id
    
    def test_attendance_policy_serialization(self):
        """Test AttendancePolicy serialization round-trip."""
        policy = AttendancePolicy(
            policy_id="POL-test",
            policy_version="1.0",
            unknown_identity_policy=IdentityHandlingPolicy.UNRESOLVED,
            ambiguous_identity_policy=IdentityHandlingPolicy.PENDING_REVIEW,
            default_entry_window_seconds=300,
            default_late_tolerance_seconds=600,
            default_exit_window_seconds=300,
        )
        
        policy_dict = policy.to_dict()
        policy2 = AttendancePolicy.from_dict(policy_dict)
        assert policy2.policy_id == policy.policy_id
        assert policy2.unknown_identity_policy == policy.unknown_identity_policy
    
    def test_attendance_decision_serialization(self):
        """Test AttendanceDecision serialization round-trip."""
        from app.attendance.policy import AttendanceDecision, DecisionReason
        
        decision = AttendanceDecision(
            decision_id="DEC-test",
            identity_certainty="known",
            identity_candidate="HS001",
            identity_confidence=0.9,
            identity_evidence_ref="GO-123",
            direction="in",
            event_timestamp=27000.0,
            event_frame_index=100,
            camera_id="CAM1",
            local_track_id="track_001",
            global_observation_id="GO-123",
            source_raw_event_id="RAW-123",
            source_resolution_id="RES-123",
            source_crossing_event_id="CROSS-123",
            geometry_version=1,
            geometry_config_hash="geom123",
            resolver_version="1.0",
            resolver_config_hash="resolver123",
            timetable_id="TTB-test",
            timetable_version="1.0",
            session_id="MATH101_MON",
            day="monday",
            previous_attendance_state="unknown",
            new_attendance_state="present",
            decision_reason=DecisionReason.WITHIN_ENTRY_WINDOW.value,
            attendance_policy_id="POL-test",
            attendance_policy_version="1.0",
            decision_schema_version="1.0",
        )
        
        decision_dict = decision.to_dict()
        decision2 = AttendanceDecision.from_dict(decision_dict)
        assert decision2.decision_id == decision.decision_id
        assert decision2.identity_candidate == decision.identity_candidate
        assert decision2.new_attendance_state == decision.new_attendance_state


if __name__ == "__main__":
    pytest.main([__file__, "-v"])