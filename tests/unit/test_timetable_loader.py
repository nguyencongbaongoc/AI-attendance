"""
Phase 37A — Unit Tests for Timetable Loader.

Tests the TimetableLoader functionality including:
- Valid timetable loading
- Malformed rows handling
- Invalid student_id validation
- Invalid day validation
- Invalid time validation
- Duplicate/conflicting entries detection
"""

import tempfile
import os
from pathlib import Path
from datetime import date

import openpyxl
import pytest

from app.attendance.timetable_loader import (
    TimetableLoader,
    TimetableLoadResult,
    TimetableValidationError,
    load_timetable_from_excel,
    parse_time_value,
    parse_day_value,
    parse_session_type_value,
    create_sample_timetable_excel,
)
from app.attendance.timetable import (
    Timetable,
    TimetableEntry,
    SessionDay,
    SessionType,
)
from app.vision.enrollment_contract import EnrollmentDatabaseMetadata


class TestTimeParsing:
    """Test time parsing utilities."""
    
    def test_parse_time_value_seconds(self):
        assert parse_time_value(3600) == 3600
        assert parse_time_value(3661) == 3661
        assert parse_time_value(0) == 0
    
    def test_parse_time_value_string_hms(self):
        assert parse_time_value("01:00:00") == 3600
        assert parse_time_value("01:01:01") == 3661
        assert parse_time_value("23:59:59") == 86399
    
    def test_parse_time_value_string_hm(self):
        assert parse_time_value("01:00") == 3600
        assert parse_time_value("01:30") == 5400
    
    def test_parse_time_value_time_object(self):
        from datetime import time
        assert parse_time_value(time(1, 0, 0)) == 3600
        assert parse_time_value(time(1, 1, 1)) == 3661
    
    def test_parse_time_value_invalid(self):
        with pytest.raises(ValueError):
            parse_time_value("invalid")
        with pytest.raises(ValueError):
            parse_time_value("25:00:00")


class TestDayParsing:
    """Test day parsing utilities."""
    
    def test_parse_day_value_full(self):
        assert parse_day_value("monday") == SessionDay.MONDAY
        assert parse_day_value("tuesday") == SessionDay.TUESDAY
        assert parse_day_value("sunday") == SessionDay.SUNDAY
    
    def test_parse_day_value_abbreviated(self):
        assert parse_day_value("mon") == SessionDay.MONDAY
        assert parse_day_value("tue") == SessionDay.TUESDAY
        assert parse_day_value("wed") == SessionDay.WEDNESDAY
        assert parse_day_value("thu") == SessionDay.THURSDAY
        assert parse_day_value("fri") == SessionDay.FRIDAY
        assert parse_day_value("sat") == SessionDay.SATURDAY
        assert parse_day_value("sun") == SessionDay.SUNDAY
    
    def test_parse_day_value_case_insensitive(self):
        assert parse_day_value("MONDAY") == SessionDay.MONDAY
        assert parse_day_value("Monday") == SessionDay.MONDAY
    
    def test_parse_day_value_invalid(self):
        with pytest.raises(ValueError):
            parse_day_value("invalid")
        with pytest.raises(ValueError):
            parse_day_value("funday")


class TestSessionTypeParsing:
    """Test session type parsing utilities."""
    
    def test_parse_session_type_value_full(self):
        assert parse_session_type_value("morning") == SessionType.MORNING
        assert parse_session_type_value("afternoon") == SessionType.AFTERNOON
        assert parse_session_type_value("full_day") == SessionType.FULL_DAY
        assert parse_session_type_value("evening") == SessionType.EVENING
    
    def test_parse_session_type_value_abbreviated(self):
        assert parse_session_type_value("am") == SessionType.MORNING
        assert parse_session_type_value("pm") == SessionType.AFTERNOON
        assert parse_session_type_value("full") == SessionType.FULL_DAY
        assert parse_session_type_value("eve") == SessionType.EVENING
    
    def test_parse_session_type_value_invalid(self):
        with pytest.raises(ValueError):
            parse_session_type_value("invalid")


class TestTimetableLoader:
    """Test TimetableLoader functionality."""
    
    def create_test_excel(self, data: list, headers: list = None) -> str:
        """Create a temporary Excel file with test data."""
        if headers is None:
            headers = [
                "student_id", "class_name", "day", "session_type",
                "entry_time", "exit_time", "session_id", "person_name",
                "entry_window_start", "entry_window_end", "late_tolerance",
                "exit_window_start", "exit_window_end", "timetable_version"
            ]
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timetable"
        
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))
        
        temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(temp_file.name)
        temp_file.close()
        return temp_file.name
    
    def test_load_valid_timetable(self):
        """Test loading a valid timetable."""
        data = [
            {
                "student_id": "HS001",
                "class_name": "Math 101",
                "day": "monday",
                "session_type": "morning",
                "entry_time": "07:30:00",
                "exit_time": "12:00:00",
                "session_id": "MATH101_MON",
                "person_name": "Student One",
            },
            {
                "student_id": "HS002",
                "class_name": "Physics 101",
                "day": "monday",
                "session_type": "morning",
                "entry_time": "08:30:00",
                "exit_time": "13:00:00",
                "session_id": "PHYS101_MON",
                "person_name": "Student Two",
            },
        ]
        
        file_path = self.create_test_excel(data)
        try:
            loader = TimetableLoader(enrollment_person_ids=["HS001", "HS002"])
            result = loader.load_from_excel(file_path)
            
            assert result.success is True
            assert result.timetable is not None
            assert isinstance(result.timetable, Timetable)
            assert len(result.timetable.entries) == 2
            assert result.rows_valid == 2
            assert result.rows_invalid == 0
        finally:
            os.unlink(file_path)
    
    def test_load_timetable_with_defaults(self):
        """Test loading timetable with default values for optional fields."""
        data = [
            {
                "student_id": "HS001",
                "class_name": "Math 101",
                "day": "monday",
                "session_type": "morning",
                "entry_time": "07:30:00",
                "exit_time": "12:00:00",
            },
        ]
        
        file_path = self.create_test_excel(data)
        try:
            loader = TimetableLoader(enrollment_person_ids=["HS001"])
            result = loader.load_from_excel(file_path)
            
            assert result.success is True
            assert len(result.timetable.entries) == 1
            entry = result.timetable.entries[0]
            assert entry.session_id == "Math 101_monday"  # Auto-generated
            assert entry.entry_window_start == 26700  # 07:30 - 300 = 07:25
            assert entry.entry_window_end == 27300    # 07:30 + 300 = 07:35
            assert entry.late_tolerance == 600
            assert entry.exit_window_start == 42900   # 12:00 - 300 = 11:55
            assert entry.exit_window_end == 43500     # 12:00 + 300 = 12:05
        finally:
            os.unlink(file_path)
    
    def test_load_timetable_missing_required_column(self):
        """Test loading timetable with missing required column (student_id)."""
        data = [
            {
                # Missing student_id
                "class_name": "Math 101",
                "day": "monday",
                "session_type": "morning",
                "entry_time": "07:30:00",
                "exit_time": "12:00:00",
            },
        ]
        
        file_path = self.create_test_excel(data)
        try:
            loader = TimetableLoader()
            result = loader.load_from_excel(file_path)
            
            assert result.success is False
            assert any("student_id is required" in e.message for e in result.errors)
        finally:
            os.unlink(file_path)
    
    def test_load_timetable_invalid_student_id(self):
        """Test loading timetable with student_id not in enrollment."""
        data = [
            {
                "student_id": "HS999",  # Not in enrollment
                "class_name": "Math 101",
                "day": "monday",
                "session_type": "morning",
                "entry_time": "07:30:00",
                "exit_time": "12:00:00",
            },
        ]
        
        file_path = self.create_test_excel(data)
        try:
            loader = TimetableLoader(enrollment_person_ids=["HS001", "HS002"])
            result = loader.load_from_excel(file_path)
            
            assert result.success is False
            assert any("not found in enrollment database" in e.message for e in result.errors)
        finally:
            os.unlink(file_path)
    
    def test_load_timetable_invalid_day(self):
        """Test loading timetable with invalid day."""
        data = [
            {
                "student_id": "HS001",
                "class_name": "Math 101",
                "day": "funday",  # Invalid day
                "session_type": "morning",
                "entry_time": "07:30:00",
                "exit_time": "12:00:00",
            },
        ]
        
        file_path = self.create_test_excel(data)
        try:
            loader = TimetableLoader(enrollment_person_ids=["HS001"])
            result = loader.load_from_excel(file_path)
            
            assert result.success is False
            assert any("Invalid day" in e.message for e in result.errors)
        finally:
            os.unlink(file_path)
    
    def test_load_timetable_invalid_time(self):
        """Test loading timetable with invalid time format."""
        data = [
            {
                "student_id": "HS001",
                "class_name": "Math 101",
                "day": "monday",
                "session_type": "morning",
                "entry_time": "invalid_time",
                "exit_time": "12:00:00",
            },
        ]
        
        file_path = self.create_test_excel(data)
        try:
            loader = TimetableLoader(enrollment_person_ids=["HS001"])
            result = loader.load_from_excel(file_path)
            
            assert result.success is False
            assert any("Cannot parse time value" in e.message for e in result.errors)
        finally:
            os.unlink(file_path)
    
    def test_load_timetable_entry_after_exit(self):
        """Test loading timetable with entry_time >= exit_time."""
        data = [
            {
                "student_id": "HS001",
                "class_name": "Math 101",
                "day": "monday",
                "session_type": "morning",
                "entry_time": "12:00:00",
                "exit_time": "07:30:00",  # Before entry
            },
        ]
        
        file_path = self.create_test_excel(data)
        try:
            loader = TimetableLoader(enrollment_person_ids=["HS001"])
            result = loader.load_from_excel(file_path)
            
            assert result.success is False
            assert any("entry_time must be < exit_time" in e.message for e in result.errors)
        finally:
            os.unlink(file_path)
    
    def test_load_timetable_duplicate_entries(self):
        """Test loading timetable with duplicate entries."""
        data = [
            {
                "student_id": "HS001",
                "class_name": "Math 101",
                "day": "monday",
                "session_type": "morning",
                "entry_time": "07:30:00",
                "exit_time": "12:00:00",
                "session_id": "MATH101_MON",
            },
            {
                "student_id": "HS001",
                "class_name": "Math 101",
                "day": "monday",
                "session_type": "morning",
                "entry_time": "07:30:00",
                "exit_time": "12:00:00",
                "session_id": "MATH101_MON",  # Duplicate
            },
        ]
        
        file_path = self.create_test_excel(data)
        try:
            loader = TimetableLoader(enrollment_person_ids=["HS001"])
            result = loader.load_from_excel(file_path)
            
            assert result.success is False
            assert any("Duplicate timetable entry" in e.message for e in result.errors)
        finally:
            os.unlink(file_path)
    
    def test_load_timetable_overlapping_sessions(self):
        """Test loading timetable with overlapping sessions for same student/day."""
        data = [
            {
                "student_id": "HS001",
                "class_name": "Math 101",
                "day": "monday",
                "session_type": "morning",
                "entry_time": "07:30:00",
                "exit_time": "12:00:00",
                "session_id": "MATH101_MON",
            },
            {
                "student_id": "HS001",
                "class_name": "Physics 101",
                "day": "monday",
                "session_type": "morning",
                "entry_time": "10:00:00",  # Overlaps with Math
                "exit_time": "14:00:00",
                "session_id": "PHYS101_MON",
            },
        ]
        
        file_path = self.create_test_excel(data)
        try:
            loader = TimetableLoader(enrollment_person_ids=["HS001"])
            result = loader.load_from_excel(file_path)
            
            assert result.success is False
            assert any("Overlapping sessions" in e.message for e in result.errors)
        finally:
            os.unlink(file_path)
    
    def test_load_timetable_non_overlapping_sessions(self):
        """Test loading timetable with non-overlapping sessions for same student/day."""
        data = [
            {
                "student_id": "HS001",
                "class_name": "Math 101",
                "day": "monday",
                "session_type": "morning",
                "entry_time": "07:30:00",
                "exit_time": "10:00:00",
                "session_id": "MATH101_MON",
            },
            {
                "student_id": "HS001",
                "class_name": "Physics 101",
                "day": "monday",
                "session_type": "afternoon",
                "entry_time": "10:00:00",  # Exactly at previous exit
                "exit_time": "14:00:00",
                "session_id": "PHYS101_MON",
            },
        ]
        
        file_path = self.create_test_excel(data)
        try:
            loader = TimetableLoader(enrollment_person_ids=["HS001"])
            result = loader.load_from_excel(file_path)
            
            assert result.success is True
            assert len(result.timetable.entries) == 2
        finally:
            os.unlink(file_path)
    
    def test_load_timetable_invalid_window(self):
        """Test loading timetable with invalid window (start > end)."""
        data = [
            {
                "student_id": "HS001",
                "class_name": "Math 101",
                "day": "monday",
                "session_type": "morning",
                "entry_time": "07:30:00",
                "exit_time": "12:00:00",
                "entry_window_start": "08:00:00",  # After entry_window_end
                "entry_window_end": "07:00:00",
            },
        ]
        
        file_path = self.create_test_excel(data)
        try:
            loader = TimetableLoader(enrollment_person_ids=["HS001"])
            result = loader.load_from_excel(file_path)
            
            assert result.success is False
            assert any("entry_window_start must be <= entry_window_end" in e.message for e in result.errors)
        finally:
            os.unlink(file_path)
    
    def test_load_timetable_empty_file(self):
        """Test loading empty timetable file."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timetable"
        
        temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(temp_file.name)
        temp_file.close()
        
        try:
            loader = TimetableLoader()
            result = loader.load_from_excel(temp_file.name)
            
            assert result.success is False
            assert any("Required columns not found" in e.message for e in result.errors)
        finally:
            os.unlink(temp_file.name)
    
    def test_load_timetable_file_not_found(self):
        """Test loading non-existent file."""
        loader = TimetableLoader()
        result = loader.load_from_excel("/nonexistent/path/file.xlsx")
        
        assert result.success is False
        assert any("File not found" in e.message for e in result.errors)
    
    def test_load_timetable_unsupported_format(self):
        """Test loading unsupported file format."""
        temp_file = tempfile.NamedTemporaryFile(suffix=".csv", delete=False)
        temp_file.write(b"student_id,class_name\nHS001,Math")
        temp_file.close()
        
        try:
            loader = TimetableLoader()
            result = loader.load_from_excel(temp_file.name)
            
            assert result.success is False
            assert any("Unsupported file format" in e.message for e in result.errors)
        finally:
            os.unlink(temp_file.name)


class TestTimetableVersioning:
    """Test timetable versioning functionality."""
    
    def test_compute_timetable_version_info(self):
        """Test computing version info for a timetable."""
        from app.attendance.timetable_loader import compute_timetable_version_info
        
        # Create a sample timetable
        entries = [
            TimetableEntry(
                entry_id="ENT-1",
                person_id="HS001",
                session_id="MATH101_MON",
                session_type=SessionType.MORNING,
                day=SessionDay.MONDAY,
                class_name="Math 101",
                entry_time=27000,
                exit_time=43200,
                entry_window_start=26700,
                entry_window_end=27300,
                late_tolerance=600,
                exit_window_start=42900,
                exit_window_end=43500,
            ),
        ]
        timetable = Timetable(
            timetable_id="TTB-test",
            timetable_version="1.0",
            entries=entries,
        )
        
        # Create a temp file for source
        temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp_file.write(b"test content")
        temp_file.close()
        
        try:
            version_info = compute_timetable_version_info(timetable, temp_file.name)
            
            assert version_info.timetable_id == "TTB-test"
            assert version_info.timetable_version == "1.0"
            assert version_info.source_file == temp_file.name
            assert version_info.entry_count == 1
            assert version_info.person_ids == ["HS001"]
            assert version_info.days_covered == ["monday"]
            assert version_info.sessions_covered == ["MATH101_MON"]
        finally:
            os.unlink(temp_file.name)
    
    def test_detect_timetable_change(self):
        """Test detecting timetable source file changes."""
        from app.attendance.timetable_loader import (
            compute_timetable_version_info,
            detect_timetable_change,
            TimetableVersionInfo,
        )
        
        # Create a temp file
        temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp_file.write(b"original content")
        temp_file.close()
        
        try:
            # Create version info
            version_info = TimetableVersionInfo(
                timetable_id="TTB-test",
                timetable_version="1.0",
                source_file=temp_file.name,
                source_hash="abc123",  # Wrong hash
                loaded_at="2026-01-01T00:00:00Z",
                entry_count=1,
                person_ids=["HS001"],
                days_covered=["monday"],
                sessions_covered=["MATH101_MON"],
            )
            
            # Should detect change
            changed, new_hash = detect_timetable_change(version_info, temp_file.name)
            assert changed is True
            assert new_hash is not None
            
            # Update with correct hash
            version_info2 = TimetableVersionInfo(
                timetable_id="TTB-test",
                timetable_version="1.0",
                source_file=temp_file.name,
                source_hash=new_hash,
                loaded_at="2026-01-01T00:00:00Z",
                entry_count=1,
                person_ids=["HS001"],
                days_covered=["monday"],
                sessions_covered=["MATH101_MON"],
            )
            
            # Should not detect change
            changed2, _ = detect_timetable_change(version_info2, temp_file.name)
            assert changed2 is False
        finally:
            os.unlink(temp_file.name)


class TestConvenienceFunctions:
    """Test convenience functions."""
    
    def test_load_timetable_from_excel(self):
        """Test convenience function for loading timetable."""
        data = [
            {
                "student_id": "HS001",
                "class_name": "Math 101",
                "day": "monday",
                "session_type": "morning",
                "entry_time": "07:30:00",
                "exit_time": "12:00:00",
            },
        ]
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timetable"
        headers = ["student_id", "class_name", "day", "session_type", "entry_time", "exit_time"]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))
        
        temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(temp_file.name)
        temp_file.close()
        
        try:
            result = load_timetable_from_excel(temp_file.name, enrollment_person_ids=["HS001"])
            
            assert result.success is True
            assert result.timetable is not None
        finally:
            os.unlink(temp_file.name)
    
    def test_create_sample_timetable_excel(self):
        """Test creating sample timetable Excel."""
        temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp_file.close()
        
        try:
            create_sample_timetable_excel(temp_file.name)
            
            # Verify file exists and can be loaded
            assert os.path.exists(temp_file.name)
            
            wb = openpyxl.load_workbook(temp_file.name)
            ws = wb.active
            assert ws.title == "Timetable"
            assert ws.cell(row=1, column=1).value == "student_id"
            assert ws.max_row >= 4  # Header + 3 sample rows
        finally:
            os.unlink(temp_file.name)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])